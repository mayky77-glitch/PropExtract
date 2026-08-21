#!/usr/bin/env python3
"""Windows-only offline smoke for PropExtract's exact portable runtime.

The workflow starts this module with the installed app-local Python.  Fixtures
are generated under ``TemporaryDirectory`` and never leave the runner.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import queue
import subprocess
import sys
import tempfile
import threading
import time
from http.client import HTTPConnection
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PERMIT_NUMBER = "38-1-1-2026"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _pdf_literal(value: str) -> bytes:
    """Escape the ASCII fixture text used in a minimal, text-layer PDF."""
    return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)").encode("ascii")


def _write_text_pdf(path: Path, lines: list[str]) -> None:
    """Write a deterministic one-page PDF without a generator dependency."""
    if not lines or any(not line.isascii() for line in lines):
        raise ValueError("synthetic PDF lines must be non-empty ASCII")
    commands = [b"BT", b"/F1 12 Tf", b"72 720 Td"]
    for index, line in enumerate(lines):
        if index:
            commands.append(b"0 -18 Td")
        commands.append(b"(" + _pdf_literal(line) + b") Tj")
    commands.append(b"ET")
    stream = b"\n".join(commands) + b"\n"
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"endstream",
    ]
    payload = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for number, body in enumerate(objects, start=1):
        offsets.append(len(payload))
        payload.extend(f"{number} 0 obj\n".encode("ascii"))
        payload.extend(body)
        payload.extend(b"\nendobj\n")
    xref = len(payload)
    payload.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    payload.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        payload.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    payload.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode("ascii")
    )
    path.write_bytes(payload)


def _require_portable_windows_runtime() -> None:
    if os.name != "nt":
        raise RuntimeError("Windows smoke must run on Windows")
    lock = json.loads((PROJECT_ROOT / "windows-runtime.lock.json").read_text(encoding="utf-8"))
    expected = (
        PROJECT_ROOT
        / ".runtime"
        / "windows"
        / f"python-{lock['artifacts']['python']['version']}"
        / str(lock["pythonTree"]["executablePath"])
    ).resolve()
    if Path(sys.executable).resolve() != expected:
        raise RuntimeError("Windows smoke must use the installed app-local portable Python")


def _readline_with_timeout(process: subprocess.Popen[str], timeout: float = 5.0) -> str:
    if process.stdout is None:
        raise RuntimeError("native cancel probe stdout is unavailable")
    lines: queue.Queue[str] = queue.Queue()
    reader = threading.Thread(target=lambda: lines.put(process.stdout.readline()), daemon=True)
    reader.start()
    try:
        return lines.get(timeout=timeout).strip()
    except queue.Empty as error:
        raise RuntimeError("native cancel probe did not become ready") from error


def _qualify_native_cancel_listener(root: Path) -> str:
    """Exercise the exact helper listener over a native redirected stdin pipe."""
    if os.name != "nt":
        raise RuntimeError("native cancel listener qualification requires Windows")
    helper = (PROJECT_ROOT / "scripts" / "windows_excel_insert.ps1").read_text(encoding="utf-8")
    source_start = helper.index("using System;\n", helper.index("Add-Type -ReferencedAssemblies"))
    source_end = helper.index("\n'@\n\nfunction Test-NativeCancel", source_start)
    listener = helper[source_start:source_end]
    probe = root / "native-cancel-listener-probe.ps1"
    probe.write_text(
        "param([int]$Iterations=300)\nAdd-Type -ReferencedAssemblies 'System.Runtime.Serialization.dll' -TypeDefinition @'\n"
        + listener
        + "\n'@\n"
        + "$l=New-Object NativeCancelListener;$l.Start();if(-not $l.WaitForOpen(5000)){exit 6};"
        + "[Console]::Out.WriteLine('ready');"
        + "for($i=0;$i -lt $Iterations;$i++){$l.Poll();if($l.Failed){exit 5};"
        + "if($l.IsCancellationRequested){[Console]::Out.WriteLine('cancelled');exit 4};"
        + "Start-Sleep -Milliseconds 10};[Console]::Out.WriteLine('complete');exit 0\n",
        encoding="utf-8",
    )

    def run(*, cancel: bool) -> tuple[int, str]:
        process = subprocess.Popen(
            ["powershell.exe", "-NoLogo", "-NoProfile", "-NonInteractive", "-File", str(probe), "-Iterations", "300" if cancel else "30"],
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        try:
            if process.stdin is None:
                raise RuntimeError("native cancel probe stdin is unavailable")
            process.stdin.write('{"command":"open"}\n')
            process.stdin.flush()
            if _readline_with_timeout(process) != "ready":
                raise RuntimeError("native cancel probe returned an invalid ready marker")
            if cancel:
                process.stdin.write('{"command":"cancel"}\n')
                process.stdin.flush()
            code = process.wait(timeout=4.0)
            if process.stdout is None:
                raise RuntimeError("native cancel probe stdout disappeared")
            return code, process.stdout.read().strip()
        finally:
            if process.poll() is None:
                process.terminate()
                process.wait(timeout=1.0)

    if run(cancel=False) != (0, "complete"):
        raise RuntimeError("native no-cancel checkpoints blocked or failed")
    if run(cancel=True) != (4, "cancelled"):
        raise RuntimeError("native post-open cancel was not consumed within grace")
    return "native-powershell-live-stdin"


def _request(port: int, method: str, path: str, payload: dict[str, object] | None = None) -> tuple[int, dict[str, object], str | None]:
    connection = HTTPConnection("127.0.0.1", port, timeout=5)
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8") if payload is not None else None
    headers = {"Content-Type": "application/json"} if body is not None else {}
    connection.request(method, path, body=body, headers=headers)
    response = connection.getresponse()
    raw = response.read()
    instance = response.getheader("X-PropExtract-Instance")
    connection.close()
    return response.status, json.loads(raw.decode("utf-8")), instance


def _wait_for_job(port: int, job_id: str) -> dict[str, object]:
    for _ in range(120):
        status, job, _ = _request(port, "GET", f"/api/jobs/{job_id}")
        if status != 200:
            raise RuntimeError(f"job polling failed: HTTP {status}")
        if job.get("status") in {"done", "error"}:
            return job
        time.sleep(0.25)
    raise RuntimeError("synthetic import did not complete within 30 seconds")


def _make_register(path: Path) -> None:
    from openpyxl import Workbook
    from rns_import_server.workbook import HEADERS, SHEET

    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    for label, column in HEADERS.items():
        sheet.cell(3, column).value = label
    sheet["Y3"] = "service formula one"
    sheet["Z3"] = "service formula two"
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    sheet["Z4"] = '=IF(F4<>"",ROW(),"")'
    book.save(path)


def _assert_published_workbook(register: Path, permit: Path, report: dict[str, object]) -> None:
    """Assert the observable Excel publication contract, not just HTTP success."""
    from openpyxl import load_workbook
    from rns_import_server.workbook import HEADERS, SHEET, STATUS_COLUMN, STATUS_HEADER

    changes = report.get("changes")
    verification = report.get("verification")
    if not isinstance(changes, list) or len(changes) != 1:
        raise RuntimeError("synthetic publication did not report exactly one row")
    change = changes[0]
    if not isinstance(change, dict) or change.get("number") != PERMIT_NUMBER or change.get("outcome") != "added":
        raise RuntimeError("synthetic publication did not classify its inserted row")
    if change.get("written") != ["F4", "G4"] or change.get("status") is not None:
        raise RuntimeError("synthetic publication wrote unexpected fields or review status")
    if not isinstance(verification, dict) or verification.get("x14_preserved") is not True or verification.get("native_cf_preserved") is not True:
        raise RuntimeError("synthetic publication did not verify native Excel conditional formatting")

    book = load_workbook(register, data_only=False)
    sheet = book[SHEET]
    if (
        sheet["F4"].value != PERMIT_NUMBER
        or sheet["G4"].value.strftime("%d.%m.%Y") != "01.02.2026"
        or sheet["Y4"].value != '=IF(A4<>"",ROW(),"")'
        or sheet["Z4"].value != '=IF(F4<>"",ROW(),"")'
    ):
        raise RuntimeError("synthetic publication changed data or Y/Z formulas")
    link = sheet.cell(4, HEADERS["Ссылка на документ"])
    if link.value != permit.name or not link.hyperlink or link.hyperlink.target != permit.resolve().as_uri():
        raise RuntimeError("synthetic publication did not write the source PDF hyperlink")
    if sheet.cell(3, STATUS_COLUMN).value != STATUS_HEADER or sheet.cell(4, STATUS_COLUMN).value is not None:
        raise RuntimeError("synthetic publication did not preserve the status/review contract")


def run() -> dict[str, object]:
    _require_portable_windows_runtime()
    from rns_import_server.app import run as import_run
    from rns_import_server.runtime import runtime_status
    from rns_import_server.server import create_server, project_instance_id

    runtime = runtime_status()
    if not runtime.get("ready") or not all(runtime.get("commands", {}).values()):
        raise RuntimeError("portable Poppler/Tesseract runtime is not ready")
    field_contracts = _verify_field_contracts()
    admin_contracts = _verify_admin_edit_contracts()

    with tempfile.TemporaryDirectory(prefix="propextract-e2e-") as temporary_name:
        temporary = Path(temporary_name) / "Проверка_Юникод"
        pdf_dir = temporary / "PDF"
        pdf_dir.mkdir(parents=True)
        permit = pdf_dir / "permit_38-1-1-2026.pdf"
        _write_text_pdf(permit, [PERMIT_NUMBER, "1.1: 01.02.2026"])
        register = temporary / "register.xlsx"
        _make_register(register)
        original_register_hash = _sha256(register)
        original_pdf_hash = _sha256(permit)

        server = create_server("127.0.0.1", 0, import_run)
        port = int(server.server_address[1])
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            status, health, instance = _request(port, "GET", "/health")
            if status != 200 or health.get("status") != "ok" or instance != project_instance_id():
                raise RuntimeError("server health identity mismatch")

            status, started, _ = _request(port, "POST", "/api/jobs", {"pdf_dir": str(pdf_dir), "xlsx": str(register), "dpi": 180})
            if status != 202 or not isinstance(started.get("id"), str):
                raise RuntimeError("first synthetic job was not accepted")
            first = _wait_for_job(port, str(started["id"]))
            if first.get("status") != "done":
                raise RuntimeError(f"first synthetic job failed: {first.get('error')}")
            if _sha256(permit) != original_pdf_hash:
                raise RuntimeError("source PDF hash changed during publication")
            first_hash = _sha256(register)
            first_mtime = register.stat().st_mtime_ns
            report = register.with_name(f"{register.stem} — отчет PropExtract.json")
            first_report = json.loads(report.read_text(encoding="utf-8"))
            if first_report.get("input_hashes") != {
                "xlsx": original_register_hash,
                "pdfs": {permit.name: original_pdf_hash},
            }:
                raise RuntimeError("first publication did not preserve its source hashes")
            _assert_published_workbook(register, permit, first_report)
            backups = register.parent / "Резервные копии PropExtract"
            first_backups = sorted(item.name for item in backups.glob("*.xlsx")) if backups.is_dir() else []
            if not first_backups:
                raise RuntimeError("published XLSX did not create a verified backup")
            if any(_sha256(backups / name) != original_register_hash for name in first_backups):
                raise RuntimeError("Excel backup is not byte-identical to the pre-publication source")

            status, started, _ = _request(port, "POST", "/api/jobs", {"pdf_dir": str(pdf_dir), "xlsx": str(register), "dpi": 180})
            if status != 202 or not isinstance(started.get("id"), str):
                raise RuntimeError("second synthetic job was not accepted")
            second = _wait_for_job(port, str(started["id"]))
            if second.get("status") != "done":
                raise RuntimeError(f"second synthetic job failed: {second.get('error')}")
            if _sha256(register) != first_hash or register.stat().st_mtime_ns != first_mtime:
                raise RuntimeError("identical no-op changed XLSX bytes or mtime")
            second_report = json.loads(report.read_text(encoding="utf-8"))
            if second_report.get("input_hashes") != {
                "xlsx": first_hash,
                "pdfs": {permit.name: original_pdf_hash},
            }:
                raise RuntimeError("no-op did not preserve its source hashes")
            second_backups = sorted(item.name for item in backups.glob("*.xlsx")) if backups.is_dir() else []
            if second_backups != first_backups:
                raise RuntimeError("identical no-op created another Excel backup")
            if _sha256(permit) != original_pdf_hash:
                raise RuntimeError("source PDF hash changed during no-op")
            verification = first_report.get("verification")
            return {
                "status": "ok",
                "portable_python": str(Path(sys.executable).resolve()),
                "unicode_fixture": temporary.name,
                "first_backup_count": len(first_backups),
                "no_op_backup_count": len(second_backups),
                "runtime_commands": runtime["commands"],
                "x14_preserved": isinstance(verification, dict) and verification.get("x14_preserved") is True,
                "native_cf_preserved": isinstance(verification, dict) and verification.get("native_cf_preserved") is True,
                "field_contracts": field_contracts,
                "admin_contracts": admin_contracts,
            }
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=10)
            if thread.is_alive():
                raise RuntimeError("synthetic server did not stop")


def _verify_field_contracts() -> str:
    # Keep these checks direct and dependency-free beyond the app's portable
    # runtime: they exercise regression contracts before the full HTTP smoke.
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))
    from rns_import_server import app
    from rns_import_server.normalization import field_comparison_equal
    from rns_import_server.ocr import OCRLine, OCRText, OCRWord
    from rns_import_server.rns_adapter import extract, norm
    from rns_import_server.workbook import _change_outcome

    number = "38-03-06-2025"

    def record(filename: str, *, end: str, changed: str) -> dict[str, object]:
        return {
            "number": number,
            "filename": filename,
            "pdf": filename,
            "end": end,
            "changed": changed,
            "field_provenance": {"end": "ocr", "changed": "ocr"},
            "warnings": [],
        }

    permit = record("permit.pdf", end="16.12.2026", changed="01.01.2026")
    change = record("amendment-a.pdf", end="16.12.2026", changed="03.01.2026")
    extension = record("extension-b.pdf", end="17.12.2026", changed="02.01.2026")
    forwards = app._merge_group(number, [permit, change, extension], [])
    backwards = app._merge_group(number, [permit, extension, change], [])
    if (
        forwards is None
        or backwards is None
        or forwards["end"] != "17.12.2026"
        or forwards["filename"] != "amendment-a.pdf"
        or (forwards["filename"], forwards["pdf"])
        != (backwards["filename"], backwards["pdf"])
    ):
        raise RuntimeError("amendment merge is not later-date or order deterministic")

    directive_only = app._merge_group(number, [change, extension], [])
    if directive_only is None or not directive_only.get("existing_only") or directive_only.get("end") != "17.12.2026":
        raise RuntimeError("directive-only merge did not keep its newest deadline")
    if (
        app._identity_retry_page_limit(100, 14) != 10
        or app._identity_retry_page_limit(4, 14) != 4
    ):
        raise RuntimeError("identity retry page limit is not bounded")
    try:
        app.collect(Path("."), 180, -4, pdfs=[])
    except ValueError as error:
        if "отрицательным" not in str(error):
            raise RuntimeError("negative page limit did not keep its Russian diagnostic") from error
    else:
        raise RuntimeError("negative page limit reached the OCR pipeline")
    issuer_a = record("amendment-issuer-a.pdf", end="16.12.2026", changed="03.01.2026")
    issuer_b = record("amendment-issuer-b.pdf", end="16.12.2026", changed="03.01.2026")
    issuer_a["issuer"], issuer_b["issuer"] = "Орган А", "Орган Б"
    quarantined = app._merge_group(number, [issuer_a, issuer_b], [])
    if (
        quarantined is None
        or quarantined.get("issuer") is not None
        or not any(issue.get("field") == "issuer" for issue in quarantined.get("merge_issues", []))
    ):
        raise RuntimeError("conflicting directive issuer was not quarantined for review")
    if _change_outcome(True, ["D4"], ["спор документов"]) != "review":
        raise RuntimeError("new row with merge issues is not classified for review")
    javascript = (PROJECT_ROOT / "rns_import_server" / "static" / "app.js").read_text(encoding="utf-8")
    if not all(
        marker in javascript
        for marker in (
            "function renderReviewCard",
            ".filter(item => item.needs_review",
            "item.details ||",
            "item.review_details",
            "item.review_details || (item.status !== \"approved\" && !manuallyResolvedStatuses.has(item.status))",
            "item.status === \"approved\" && !item.review_details",
            "Строки для проверки",
        )
    ):
        raise RuntimeError("admin UI no longer renders non-proposal review rows in the decision group")

    quote_equal = field_comparison_equal("Застройщик", 'ПАО "Тест"', "ПАО «Тест»")
    prefix_equal = field_comparison_equal(
        "Наименование объекта", "ПС Северная. Этап 1", "ПС Южная. Этап 1"
    )
    if quote_equal is not True or prefix_equal is not False:
        raise RuntimeError("field comparison lost quote or meaningful-prefix distinction")
    if not field_comparison_equal(
        "Наименование объекта",
        "ПС Восточная 110 кВ, мощность 16 МВА",
        "ПС Восточная 110 KB, мощность 16МВА",
    ):
        raise RuntimeError("numeric engineering-unit comparison is not stable")

    garbage_value = "а Иркутской области разрешения"
    garbage_lines = (
        OCRLine(
            1,
            1000,
            1400,
            tuple(
                OCRWord(token, 20 + index * 45, 30, 35, 16, 96.1)
                for index, token in enumerate(
                    f"Номер разрешения на строительство: {number}".split()
                )
            ),
        ),
        OCRLine(
            1,
            1000,
            1400,
            tuple(
                OCRWord(token, 20 + index * 45, 60, 35, 16, 96.1)
                for index, token in enumerate(f"Муниципальный район: {garbage_value}".split())
            ),
        ),
    )
    garbage = extract(
        Path("synthetic.pdf"),
        OCRText(
            f"Номер разрешения на строительство: {number}\n"
            f"Муниципальный район: {garbage_value}",
            garbage_lines,
            source="raster",
        ),
    )
    if (
        garbage is None
        or garbage.get("field_quality", {}).get("district", {}).get("status") != "review"
    ):
        raise RuntimeError("garbled high-confidence district is not review-only")

    before_label_line = OCRLine(
        1,
        1000,
        1400,
        tuple(
            OCRWord(token, 20 + index * 45, 60, 35, 16, 96.1)
            for index, token in enumerate(
                "Жигаловский район справочно Муниципальный район:".split()
            )
        ),
    )
    before_label = extract(
        Path("synthetic.pdf"),
        OCRText(
            f"Номер разрешения на строительство: {number}\n"
            "Жигаловский район справочно Муниципальный район:",
            (garbage_lines[0], before_label_line),
            source="raster",
        ),
    )
    if (
        before_label is None
        or before_label.get("field_quality", {}).get("district", {}).get("status") != "review"
    ):
        raise RuntimeError("text before a field label became actionable")

    trace = app._trace_for_read(
        OCRText(
            "no identity",
            source="text_layer",
            trace={
                "effective_dpi": 400,
                "processed_pages": 3,
                "total_pages": 3,
                "tesseract_calls": 0,
            },
        ),
        3,
        400,
    )
    if not all(key in trace for key in ("route", "tesseract_started", "dpi", "pages", "ocr_calls", "fallback_reason")):
        raise RuntimeError("Windows OCR trace compatibility fields are absent")
    if trace["dpi"] != 400 or trace["pages"] != 3 or trace["ocr_calls"] != 0:
        raise RuntimeError("Windows OCR trace compatibility fields are inconsistent")

    with tempfile.TemporaryDirectory(prefix="propextract-retry-contract-") as retry_name:
        retry_root = Path(retry_name)
        permit_pdf = retry_root / "scan.pdf"
        permit_pdf.write_bytes(b"synthetic")
        calls: list[tuple[int, int, bool]] = []

        def bad_layer_then_raster(
            source: Path,
            dpi: int,
            pages: int,
            *,
            force_ocr: bool = False,
        ) -> tuple[OCRText, int]:
            calls.append((dpi, pages, force_ocr))
            if force_ocr:
                return OCRText(
                    f"Номер разрешения на строительство: {number}",
                    source="raster",
                ), 1
            return OCRText("Номер не читается", source="text_layer"), 16

        original_read_ocr = app.read_ocr
        try:
            app.read_ocr = bad_layer_then_raster
            recovered, documents = app.collect(retry_root, 400, 0, pdfs=[permit_pdf])
        finally:
            app.read_ocr = original_read_ocr
        if list(recovered) != [number] or calls != [(400, 0, False), (400, 10, True)]:
            raise RuntimeError("direct-400 bad text layer did not get one bounded forced-raster retry")
        if documents[0].get("outcome") != "processed_rns":
            raise RuntimeError("generic direct-400 fixture did not recover its permit number")

        calls.clear()

        permit_context_pdf = retry_root / "Разрешение по ГПЗУ.pdf"
        permit_context_pdf.write_bytes(b"synthetic")

        def permit_context_then_raster(
            source: Path,
            dpi: int,
            pages: int,
            *,
            force_ocr: bool = False,
        ) -> tuple[OCRText, int]:
            calls.append((dpi, pages, force_ocr))
            if force_ocr:
                return OCRText(
                    f"Номер разрешения на строительство: {number}",
                    source="raster",
                ), 16
            return OCRText("ГПЗУ: градостроительный план земельного участка", source="text_layer"), 16

        try:
            app.read_ocr = permit_context_then_raster
            recovered, documents = app.collect(retry_root, 180, 0, pdfs=[permit_context_pdf])
        finally:
            app.read_ocr = original_read_ocr
        if list(recovered) != [number] or calls != [(180, 0, False), (180, 10, True)]:
            raise RuntimeError("permit basename lost precedence over a strong GPZU text-layer title")
        if documents[0].get("outcome") != "processed_rns":
            raise RuntimeError("permit/context fixture did not recover its permit number")

        calls.clear()

        def raster_400_miss(
            source: Path,
            dpi: int,
            pages: int,
            *,
            force_ocr: bool = False,
        ) -> tuple[OCRText, int]:
            calls.append((dpi, pages, force_ocr))
            return OCRText("нет номера", source="raster"), 1

        try:
            app.read_ocr = raster_400_miss
            app.collect(retry_root, 400, 0, pdfs=[permit_pdf])
        finally:
            app.read_ocr = original_read_ocr
        if calls != [(400, 0, False)]:
            raise RuntimeError("an actual 400-DPI raster miss was processed twice")

        context_pdf = retry_root / f"ГПЗУ для РНС {number}.pdf"
        context_pdf.write_bytes(b"synthetic")
        calls.clear()
        try:
            app.read_ocr = bad_layer_then_raster
            context_records, context_documents = app.collect(
                retry_root, 400, 0, pdfs=[context_pdf]
            )
        finally:
            app.read_ocr = original_read_ocr
        if calls or context_records or context_documents[0].get("outcome") != "out_of_scope":
            raise RuntimeError("leading context document type lost to a later RNS reference")

    projected = app.safe_report_projection(
        {
            "documents": [
                {
                    "file": r"C:\Private Folder\permit.pdf",
                    "ocr_text": "private OCR text",
                    "technical_error": "Failed /Users/operator/Secret Project/permit.pdf",
                    "ocr_trace": trace,
                }
            ],
            "capability": "private capability",
        }
    )
    if (
        projected["documents"][0].get("file") != "permit.pdf"
        or "ocr_text" in projected["documents"][0]
        or "capability" in projected
        or "Secret Project" in projected["documents"][0].get("technical_error", "")
    ):
        raise RuntimeError("disk report projection retained private OCR/path data")

    labeled = (
        f"Номер разрешения на строительство: {number}\n"
        "Дата выдачи: 01.02.2025\nСправочный номер: RU-99999999-99-2026"
    )
    ambiguous = (
        f"Номер разрешения на строительство: {number}\n"
        "Номер разрешения на строительство: RU-99999999-99-2026"
    )
    if (
        norm(Path("synthetic.pdf"), labeled) != number
        or norm(Path("synthetic.pdf"), ambiguous) is not None
    ):
        raise RuntimeError("labeled permit identity selection is invalid")

    reordered = f"|-2. строительство: Номер разрешения на [{number}"
    narrative_reference = (
        f"Для справки указан номер разрешения прежнего строительства: {number}.\n"
        "Основной реквизит документа: RU-99999999-99-2026."
    )
    repeated_reordered = f"{reordered}\n2. строительство: Номер разрешения на [{number}"
    if (
        norm(Path("synthetic.pdf"), reordered) != number
        or norm(Path("synthetic.pdf"), narrative_reference) is not None
        or norm(Path("synthetic.pdf"), repeated_reordered) is not None
    ):
        raise RuntimeError("reordered identity grammar accepted a narrative reference or repeated form label")

    terminal = extract(
        Path("synthetic.pdf"),
        f"Номер разрешения на строительство: {number}\nМуниципальный район: Тестовый район составе",
    )
    retained = extract(
        Path("synthetic.pdf"),
        f"Номер разрешения на строительство: {number}\n"
        "Муниципальный район: Тестовый район в составе муниципального образования",
    )
    if (
        terminal is None
        or retained is None
        or terminal.get("district") != "Тестовый район"
        or retained.get("district") != "Тестовый район в составе муниципального образования"
    ):
        raise RuntimeError("district cleanup does not preserve intended boundaries")

    return "merge,comparison,identity,district,quality,ocr-retry,ocr-trace,report-privacy,review"


def _verify_admin_edit_contracts() -> str:
    """Exercise portable invalid-date, quality, edit, and OOXML contracts."""
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))
    from datetime import datetime

    from openpyxl import Workbook, load_workbook
    from rns_import_server import app
    from rns_import_server.audit import sha256
    from rns_import_server.ocr import OCRLine, OCRText, OCRWord
    from rns_import_server.rns_adapter import extract
    from rns_import_server.server import JobManager
    from rns_import_server.workbook import SHEET, apply

    invalid = extract(Path(f"Разрешение {PERMIT_NUMBER} от 03.02.2026.pdf"), f"Номер разрешения на строительство: {PERMIT_NUMBER}\nДата выдачи: 11.41.2025")
    if not invalid or invalid.get("issue") != "03.02.2026" or "invalid_date:issue" not in invalid.get("warnings", []):
        raise RuntimeError("invalid date was not quarantined at field scope")
    if app._document_outcome(Path("ГРО.pdf"), "ГРО", None) != "out_of_scope" or app._document_outcome(Path("Разрешение.pdf"), "Номер разрешения на строительство", None) != "unidentified_permit":
        raise RuntimeError("typed document outcomes are not stable")
    quality_text = f"Номер разрешения на строительство: {PERMIT_NUMBER}\nРазработчик ПД: ООО Проект"
    words = tuple(OCRWord(token, 10 + index * 30, 20, 20, 10, 24.0) for index, token in enumerate(quality_text.replace("\n", " ").split()))
    quality = extract(Path("permit.pdf"), OCRText(quality_text, (OCRLine(1, 10000, 1000, words),), source="raster"))
    if not quality or quality.get("field_quality", {}).get("developer", {}).get("status") != "review":
        raise RuntimeError("low-quality raster field is not review-only")

    with tempfile.TemporaryDirectory(prefix="propextract-admin-contract-") as temporary_name:
        root = Path(temporary_name)
        pdf_dir = root / "pdf"; pdf_dir.mkdir()
        old_pdf, pdf = pdf_dir / "old.pdf", pdf_dir / "new.pdf"
        old_pdf.write_bytes(b"old"); pdf.write_bytes(b"new")
        target = root / "register.xlsx"
        book = Workbook(); sheet = book.active; sheet.title = SHEET
        sheet["F4"], sheet["D4"], sheet["H4"] = PERMIT_NUMBER, "Старый объект", datetime(2025, 12, 31)
        sheet["W4"] = old_pdf.name; sheet["W4"].hyperlink = old_pdf.as_uri()
        sheet["Y4"], sheet["Z4"] = '=IF(A4<>"",ROW(),"")', '=IF(F4<>"",ROW(),"")'
        book.save(target)
        quality_status = "review"
        record_object = "Новый объект"

        def runner(pdf_root: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict[str, object]:
            record = {"number": PERMIT_NUMBER, "filename": pdf.name, "pdf": str(pdf), "stage": None, "object": record_object, "issue": None, "end": None, "changed": None, "issuer": None, "builder": None, "region": None, "district": None, "developer": None, "merge_issues": [{"message": "Требуется сверка."}], "field_quality": {"object": {"status": quality_status, "reason": "low_ocr_confidence"}}}
            result = apply({PERMIT_NUMBER: record}, xlsx, output, sha256(xlsx))
            result.update(input_hashes={"xlsx": sha256(xlsx), "pdfs": {pdf.name: sha256(pdf)}}, documents=[{"file": str(pdf)}], logical_records=[PERMIT_NUMBER], selected_records={PERMIT_NUMBER: {**record, "field_sources": {"object": pdf.name}, "field_quality": {"object": {"status": quality_status, "reason": "low_ocr_confidence"}}}})
            return result

        manager = JobManager(runner, error_log=root / "error.log")
        job_id = str(manager.start(str(pdf_dir), str(target))["id"])
        for _ in range(100):
            job = manager.get(job_id)
            if job and job.get("status") in {"done", "error"}:
                break
            time.sleep(0.02)
        else:
            raise RuntimeError("manual edit contract job did not finish")
        public = manager.public(job_id)
        if not public or public.get("status") != "done" or public.get("proposals") and public["proposals"][0].get("id"):
            raise RuntimeError("low-quality proposal received one-click transfer authority")
        card = public["row_cards"][0]
        before_link = load_workbook(target, data_only=False)[SHEET]["W4"].hyperlink.target
        updated = manager.edit(job_id, str(card["edit_id"]), public["capability"], {"object": "Исправленный объект", "end": "2027-12-31"})
        saved = load_workbook(target, data_only=False)[SHEET]
        if updated["row_cards"][0].get("edited") is not True or saved["D4"].value != "Исправленный объект" or saved["H4"].value != datetime(2027, 12, 31):
            raise RuntimeError("manual edit did not publish typed values")
        if updated["row_cards"][0].get("object") != "Исправленный объект" or updated["proposals"][0].get("status") != "resolved_manual" or updated["proposals"][0].get("manual_value") != "Исправленный объект":
            raise RuntimeError("manual edit did not refresh the public review card")
        if saved["F4"].value != PERMIT_NUMBER or saved["W4"].hyperlink.target != before_link or saved["Y4"].value != '=IF(A4<>"",ROW(),"")' or saved["Z4"].value != '=IF(F4<>"",ROW(),"")':
            raise RuntimeError("manual edit changed identity, W, or Y:Z invariants")
        # D-001/D-005: a weak OCR value equal to the manual value is not a
        # workbook mutation.  Existing review remains visible but no weak
        # generated AA line, backup, hash, or mtime change is allowed.
        record_object = "Исправленный объект"
        before_hash, before_mtime = sha256(target), target.stat().st_mtime_ns
        manual_aa = str(load_workbook(target, data_only=False)[SHEET]["AA4"].value)
        if "Исправлено вручную" not in manual_aa:
            raise RuntimeError("manual edit did not persist AA audit marker")
        backups = target.parent / "Резервные копии PropExtract"
        before_backups = sorted(item.name for item in backups.glob("*.xlsx"))
        repeated_id = str(manager.start(str(pdf_dir), str(target))["id"])
        for _ in range(100):
            repeated = manager.get(repeated_id)
            if repeated and repeated.get("status") in {"done", "error"}:
                break
            time.sleep(0.02)
        else:
            raise RuntimeError("stable low-quality no-op job did not finish")
        if not repeated or repeated.get("published") is not False or repeated.get("backup") is not None:
            raise RuntimeError("manual-equivalent low-quality review republished the workbook")
        summary = repeated.get("summary", {})
        if summary.get("changed_rows") != 0 or summary.get("review_rows") != 1:
            raise RuntimeError("stable review summary does not separate physical and review rows")
        if sha256(target) != before_hash or target.stat().st_mtime_ns != before_mtime or sorted(item.name for item in backups.glob("*.xlsx")) != before_backups:
            raise RuntimeError("manual-equivalent low-quality review changed workbook or backups")
        stable_aa = str(load_workbook(target, data_only=False)[SHEET]["AA4"].value)
        if stable_aa != manual_aa or "значение PDF требует ручной сверки" in stable_aa:
            raise RuntimeError("manual-equivalent low-quality review recreated generated AA warning")

        book = Workbook(); sheet = book.active; sheet.title = SHEET
        sheet["F4"], sheet["D4"], sheet["H4"] = PERMIT_NUMBER, "Старый объект", datetime(2025, 12, 31)
        sheet["W4"] = old_pdf.name; sheet["W4"].hyperlink = old_pdf.as_uri()
        sheet["Y4"], sheet["Z4"] = '=IF(A4<>"",ROW(),"")', '=IF(F4<>"",ROW(),"")'
        book.save(target)
        quality_status = "actionable"
        actionable = JobManager(runner, error_log=root / "actionable-error.log")
        actionable_id = str(actionable.start(str(pdf_dir), str(target))["id"])
        for _ in range(100):
            job = actionable.get(actionable_id)
            if job and job.get("status") in {"done", "error"}:
                break
            time.sleep(0.02)
        else:
            raise RuntimeError("actionable proposal contract job did not finish")
        public = actionable.public(actionable_id)
        if not public or public.get("status") != "done" or not public.get("proposals") or public["proposals"][0].get("status") != "pending":
            raise RuntimeError("actionable proposal did not expose initial pending status")
        report = target.with_name(f"{target.stem} — отчет PropExtract.json")
        initial_report = json.loads(report.read_text(encoding="utf-8"))
        if not initial_report.get("input_hashes") or not initial_report.get("verification") or initial_report.get("final_state", {}).get("actions") != []:
            raise RuntimeError("initial final action report is incomplete")
        edit_id = str(public["row_cards"][0]["edit_id"])
        actionable.approve(actionable_id, str(public["proposals"][0]["id"]), public["capability"])
        approved = actionable.public(actionable_id)
        if not approved or approved["proposals"][0].get("status") != "approved":
            raise RuntimeError("actionable proposal approval did not refresh public status")
        actionable.edit(actionable_id, edit_id, public["capability"], {"object": "Финальное ручное значение"})
        final_report = json.loads(report.read_text(encoding="utf-8"))
        final = final_report.get("final_state", {})
        expected_actions = [
            {"type": "proposal_approved", "row": 4, "field": "Наименование объекта", "status": "approved"},
            {"type": "manual_edit", "row": 4, "field": "Наименование объекта", "status": "edited"},
        ]
        if final.get("actions") != expected_actions or final.get("summary", {}).get("changed_rows") is None:
            raise RuntimeError("final action report did not retain ordered final state")
        report_text = json.dumps(final_report, ensure_ascii=False)
        if str(target) in report_text or public["capability"] in report_text or "Финальное ручное значение" in report_text:
            raise RuntimeError("final action report leaked private path, capability, or raw edit")
    return "invalid-date,outcomes,quality,manual-edit,proposal-status,report-final-state,review-summary,ooxml"


def self_test() -> dict[str, object]:
    field_contracts = _verify_field_contracts()
    admin_contracts = _verify_admin_edit_contracts()

    with tempfile.TemporaryDirectory(prefix="propextract-e2e-self-test-") as temporary_name:
        fixture = Path(temporary_name) / "fixture.pdf"
        _write_text_pdf(fixture, [PERMIT_NUMBER, "1.1: 01.02.2026"])
        payload = fixture.read_bytes()
        if not payload.startswith(b"%PDF-1.4\n") or b"xref\n0 6\n" not in payload or not payload.endswith(b"%%EOF\n"):
            raise RuntimeError("synthetic PDF structure is invalid")
        return {
            "status": "ok",
            "fixture_sha256": _sha256(fixture),
            "temporary_fixture_deleted": True,
            "field_contracts": field_contracts,
            "admin_contracts": admin_contracts,
            "native_cancel_listener": _qualify_native_cancel_listener(Path(temporary_name)),
        }


def main() -> None:
    parser = argparse.ArgumentParser(description="Offline Windows portable-runtime smoke")
    parser.add_argument("--self-test", action="store_true", help="validate the dependency-free fixture generator only")
    options = parser.parse_args()
    result = self_test() if options.self_test else run()
    # Windows PowerShell 5.1 can give a redirected portable Python process a
    # legacy console encoding. JSON escapes keep the complete Unicode payload
    # representable without depending on that code page.
    print(json.dumps(result, ensure_ascii=True, sort_keys=True))


if __name__ == "__main__":
    main()
