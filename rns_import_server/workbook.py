"""Safe, style-preserving publication of an RNS import workbook."""
from __future__ import annotations

import os
import re
import shutil
import tempfile
import zipfile
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

try:
    from rns_import_server.audit import sha256
    from rns_import_server.normalization import canonical_rns_identities, field_comparison_equal
except ModuleNotFoundError:  # Direct ``python rns_import_server/app.py`` invocation.
    from audit import sha256
    from normalization import canonical_rns_identities, field_comparison_equal

SHEET = "Реестр РНС"
STATUS_COLUMN = 27  # AA; Y:Z are occupied by the register's service formulas.
STATUS_HEADER = "Статус переноса"
HEADERS = {
    "Номер этапа": 2, "Наименование объекта": 4, "Номер РНС": 6,
    "Дата выдачи": 7, "Срок действия": 8, "Дата последн. измен.": 9,
    "Орган выдачи": 10, "Застройщик": 11, "Субъект РФ": 12,
    "Муниципальный р-н": 13, "Разработчик ПД": 14, "Ссылка на документ": 23,
    "Примечание": 24,
}
DATE_FMT = "dd\\.mm\\.yyyy"
_STANDARD_CF = re.compile(rb"<conditionalFormatting\b.*?</conditionalFormatting>", re.DOTALL)
_SYSTEM_STATUS_PREFIXES = (
    "Не перенесено «",
    "Не подтверждено «",
    "В документе не найдено ни одного подтверждённого поля для переноса.",
    "Связанные изменения содержат разные значения поля «",
)


def iso_date(value: object) -> datetime | None:
    if not isinstance(value, str) or not value:
        return None
    for pattern in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, pattern)
        except ValueError:
            continue
    raise ValueError("invalid_date")


def _value_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def _same_value(label: str, existing: object, proposed: object) -> bool:
    """Use one conservative equality rule for outcome and proposal decisions."""
    return field_comparison_equal(label, _value_text(existing), _value_text(proposed))


def transfer_issue(label: str, existing: object, proposed: object) -> str | None:
    """Explain why a mapped field was not transferred, or return no issue."""
    existing_empty = existing in (None, "")
    if proposed is None:
        if existing_empty:
            return f"Не перенесено «{label}»: значение не найдено в PDF."
        return (
            f"Не подтверждено «{label}»: значение не найдено в PDF; "
            f"значение Excel «{_value_text(existing)}» сохранено."
        )
    if existing_empty or _same_value(label, existing, proposed):
        return None
    return (
        f"Не перенесено «{label}»: в Excel — «{_value_text(existing)}», "
        f"в PDF — «{_value_text(proposed)}»; значение Excel сохранено."
    )


def _change_outcome(new: bool, written: list[str], issues: list[str]) -> str:
    if issues:
        return "review"
    if new:
        return "added"
    if written:
        return "updated"
    return "already_present"


def _status_value(existing: object, issues: list[str]) -> str | None:
    """Replace generated review lines without discarding operator audit notes."""
    preserved = [
        line for line in str(existing).splitlines()
        if not line.startswith(_SYSTEM_STATUS_PREFIXES)
    ] if existing not in (None, "") else []
    lines = list(dict.fromkeys([*issues, *preserved]))
    return "\n".join(lines) or None


def _set_status_presentation(status, reference) -> bool:
    """Apply the standard AA presentation and report whether it changed."""
    changed = False
    for attribute in ("font", "fill", "border", "number_format", "protection"):
        if copy(getattr(status, attribute)) != copy(getattr(reference, attribute)):
            setattr(status, attribute, copy(getattr(reference, attribute)))
            changed = True
    alignment = copy(reference.alignment)
    alignment.wrap_text = True
    alignment.vertical = "top"
    if copy(status.alignment) != alignment:
        status.alignment = alignment
        changed = True
    return changed


_TRANSFER_FIELDS = ("stage", "object", "issue", "end", "changed", "issuer", "builder", "region", "district", "developer")
EDITABLE_FIELDS = {
    "stage": "Номер этапа",
    "object": "Наименование объекта",
    "issue": "Дата выдачи",
    "end": "Срок действия",
    "changed": "Дата последн. измен.",
    "issuer": "Орган выдачи",
    "builder": "Застройщик",
    "region": "Субъект РФ",
    "district": "Муниципальный р-н",
    "developer": "Разработчик ПД",
}
DATE_EDIT_FIELDS = frozenset({"issue", "end", "changed"})


def _sheet_xml_path(book: Path, sheet_name: str) -> str:
    main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    with zipfile.ZipFile(book) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationship_id = next(sheet.attrib[f"{{{rel}}}id"] for sheet in workbook.findall(f"{{{main}}}sheets/{{{main}}}sheet") if sheet.attrib["name"] == sheet_name)
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    target = next(item.attrib["Target"] for item in relationships if item.attrib["Id"] == relationship_id)
    return "xl/" + target.removeprefix("/").removeprefix("xl/")


def _extension_block(source: Path, sheet_path: str) -> bytes | None:
    with zipfile.ZipFile(source) as archive:
        xml = archive.read(sheet_path)
    start, end = xml.find(b"<extLst>"), xml.rfind(b"</extLst>")
    return xml[start:end + len(b"</extLst>")] if start >= 0 and end >= start else None


def _standard_cf_blocks(book: Path, sheet_path: str) -> tuple[bytes, ...]:
    """Return native worksheet CF blocks, excluding x14 rules in ``extLst``."""
    with zipfile.ZipFile(book) as archive:
        return tuple(_STANDARD_CF.findall(archive.read(sheet_path)))


def _reinject_native_conditional_formatting(source: Path, staged: Path) -> None:
    """Restore native CF bytes at a valid worksheet position after openpyxl saves."""
    sheet_path = _sheet_xml_path(source, SHEET)
    with zipfile.ZipFile(source) as archive:
        source_sheet = archive.read(sheet_path)
    source_blocks = _STANDARD_CF.findall(source_sheet)
    if not source_blocks:
        return
    with zipfile.ZipFile(staged) as archive:
        payload = {entry.filename: archive.read(entry.filename) for entry in archive.infolist()}
    sheet = _STANDARD_CF.sub(b"", payload[sheet_path])
    last = None
    for last in _STANDARD_CF.finditer(source_sheet):
        pass
    assert last is not None
    # Find the first schema-valid sibling that followed source CF, then restore
    # immediately before its equivalent in staged XML. This retains the proper
    # location after sheetProtection/autoFilter/mergeCells when they precede CF.
    following = (
        b"<dataValidations", b"<hyperlinks", b"<printOptions", b"<pageMargins", b"<pageSetup",
        b"<headerFooter", b"<rowBreaks", b"<colBreaks", b"<customProperties", b"<cellWatches",
        b"<ignoredErrors", b"<smartTags", b"<drawing", b"<legacyDrawing", b"<legacyDrawingHF",
        b"<picture", b"<oleObjects", b"<controls", b"<webPublishItems", b"<tableParts", b"<extLst",
        b"</worksheet>",
    )
    anchor = next((marker for marker in following if source_sheet.find(marker, last.end()) >= 0), None)
    position = sheet.find(anchor) if anchor else -1
    if position < 0:
        raise RuntimeError("native_conditional_formatting_position_missing")
    payload[sheet_path] = sheet[:position] + b"".join(source_blocks) + sheet[position:]
    descriptor, temporary_name = tempfile.mkstemp(prefix=".rns-native-cf-", suffix=".xlsx", dir=staged.parent)
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as archive:
            for name, data in payload.items():
                archive.writestr(name, data)
        os.replace(temporary, staged)
    finally:
        temporary.unlink(missing_ok=True)


def _reinject_extensions(source: Path, staged: Path) -> None:
    sheet_path = _sheet_xml_path(source, SHEET)
    block = _extension_block(source, sheet_path)
    if not block:
        return
    with zipfile.ZipFile(staged) as archive:
        payload = {entry.filename: archive.read(entry.filename) for entry in archive.infolist()}
    sheet = payload[sheet_path]
    if b"<extLst>" not in sheet:
        with zipfile.ZipFile(source) as archive:
            source_sheet = archive.read(sheet_path)
        source_start = source_sheet.find(b"<worksheet")
        staged_start = sheet.find(b"<worksheet")
        source_open = source_sheet[source_start:source_sheet.find(b">", source_start) + 1]
        staged_open = sheet[staged_start:sheet.find(b">", staged_start) + 1]
        namespaces = re.findall(rb'\s+xmlns:([A-Za-z0-9]+)="([^"]+)"', source_open)
        missing = b"".join(b' xmlns:' + prefix + b'="' + uri + b'"' for prefix, uri in namespaces if b"xmlns:" + prefix + b"=" not in staged_open)
        if missing:
            sheet = sheet.replace(staged_open, staged_open[:-1] + missing + b">", 1)
        payload[sheet_path] = sheet.replace(b"</worksheet>", block + b"</worksheet>")
    descriptor, temporary_name = tempfile.mkstemp(prefix=".rns-reinject-", suffix=".xlsx", dir=staged.parent)
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as archive:
            for name, data in payload.items():
                archive.writestr(name, data)
        os.replace(temporary, staged)
    finally:
        temporary.unlink(missing_ok=True)


def _rows_by_number(sheet, number: str) -> list[int]:
    """Find only cells containing one unambiguous canonical RNS identity."""
    rows: list[int] = []
    for row in range(4, sheet.max_row + 1):
        identities = canonical_rns_identities(sheet.cell(row, HEADERS["Номер РНС"]).value)
        if len(identities) == 1 and identities[0] == number:
            rows.append(row)
    return rows


def _row_by_number(sheet, number: str) -> int | None:
    rows = _rows_by_number(sheet, number)
    return rows[0] if len(rows) == 1 else None


def _ambiguous_rows_by_number(sheet, number: str) -> list[int]:
    return [
        row
        for row in range(4, sheet.max_row + 1)
        if number in canonical_rns_identities(sheet.cell(row, HEADERS["Номер РНС"]).value)
        if len(canonical_rns_identities(sheet.cell(row, HEADERS["Номер РНС"]).value)) != 1
    ]


def _copy_row_style(sheet, source: int, target: int) -> None:
    for column in range(1, sheet.max_column + 1):
        before, after = sheet.cell(source, column), sheet.cell(target, column)
        if before.has_style:
            after._style = copy(before._style)
        if isinstance(before.value, str) and before.value.startswith("="):
            after.value = Translator(before.value, origin=before.coordinate).translate_formula(after.coordinate)
    sheet.row_dimensions[target].height = sheet.row_dimensions[source].height


def _next_data_row(sheet) -> int:
    last = 3
    for row in range(4, sheet.max_row + 1):
        if any(sheet.cell(row, column).value not in (None, "") for column in range(1, 25)):
            last = row
    return last + 1


def _put(cell, value: object, label: str, number: str, conflicts: list[dict[str, str]]) -> bool:
    if value is None:
        return False
    if cell.value in (None, ""):
        cell.value = value
        if isinstance(value, str):
            cell.data_type = "s"  # OCR text must never become an Excel formula.
        return True
    if not _same_value(label, cell.value, value):
        conflicts.append({"number": number, "cell": cell.coordinate, "field": label, "existing": _value_text(cell.value), "pdf": _value_text(value), "action": "Перенести изменения"})
    return False


def _validate(
    source: Path,
    staged: Path,
    records: dict[str, dict[str, object]],
    statuses: dict[str, str | None],
    outcomes: dict[str, str],
) -> dict[str, Any]:
    source_book, output_book = load_workbook(source, data_only=False), load_workbook(staged, data_only=False)
    source_sheet, output_sheet = source_book[SHEET], output_book[SHEET]
    intended_link_cells = {
        f"{source_sheet.cell(row, HEADERS['Ссылка на документ']).coordinate}"
        for number in records
        if outcomes[number] != "already_present"
        if (row := _row_by_number(source_sheet, number)) is not None
        if any(
            source_sheet.cell(row, HEADERS[label]).value != output_sheet.cell(row, HEADERS[label]).value
            for label in HEADERS
            if label not in {"Номер РНС", "Ссылка на документ", "Примечание"}
        )
    }
    intended_status_cells = {
        f"{output_sheet.cell(row, STATUS_COLUMN).coordinate}"
        for number in records
        if outcomes[number] != "already_present"
        if (row := _row_by_number(output_sheet, number)) is not None
    }
    field_headers = {
        "stage": "Номер этапа", "object": "Наименование объекта", "issue": "Дата выдачи", "end": "Срок действия",
        "changed": "Дата последн. измен.", "issuer": "Орган выдачи", "builder": "Застройщик", "region": "Субъект РФ",
        "district": "Муниципальный р-н", "developer": "Разработчик ПД",
    }
    intended_data_style_cells = {
        source_sheet.cell(row, HEADERS[label]).coordinate
        for number, record in records.items()
        if (row := _row_by_number(source_sheet, number)) is not None
        for field, label in field_headers.items()
        if record.get(field) is not None
    }
    new_record_rows = {
        row
        for number in records
        if _row_by_number(source_sheet, number) is None
        if (row := _row_by_number(output_sheet, number)) is not None
    }
    already_present_rows = {
        row
        for number in records
        if outcomes[number] == "already_present"
        if (row := _row_by_number(source_sheet, number)) is not None
    }
    intended_style_cells = intended_link_cells | intended_status_cells | intended_data_style_cells | {output_sheet.cell(3, STATUS_COLUMN).coordinate}
    for row in range(4, source_sheet.max_row + 1):
        for column in range(1, source_sheet.max_column + 1):
            before, after = source_sheet.cell(row, column), output_sheet.cell(row, column)
            if (
                row not in new_record_rows
                and before.coordinate not in intended_style_cells
                and before._style != after._style
            ):
                raise RuntimeError(f"style_changed:{before.coordinate}")
            if row in already_present_rows and before.value != after.value:
                raise RuntimeError(f"data_changed:{before.coordinate}")
            if isinstance(before.value, str) and before.value.startswith("=") and before.value != after.value:
                raise RuntimeError(f"formula_changed:{before.coordinate}")
    for number, record in records.items():
        if outcomes[number] == "review_conflict" or not any(record.get(field) is not None for field in _TRANSFER_FIELDS):
            continue
        row = _row_by_number(output_sheet, number)
        if row is None:
            raise RuntimeError(f"record_missing:{number}")
        link = output_sheet.cell(row, HEADERS["Ссылка на документ"])
        if outcomes[number] == "already_present":
            source_row = _row_by_number(source_sheet, number)
            if source_row is None:
                raise RuntimeError(f"record_missing:{number}")
            source_link = source_sheet.cell(source_row, HEADERS["Ссылка на документ"])
            source_target = source_link.hyperlink.target if source_link.hyperlink else None
            output_target = link.hyperlink.target if link.hyperlink else None
            if link.value != source_link.value or output_target != source_target:
                raise RuntimeError(f"link_changed:{number}")
        elif outcomes[number] == "review":
            source_row = _row_by_number(source_sheet, number)
            if source_row is not None:
                source_link = source_sheet.cell(source_row, HEADERS["Ссылка на документ"])
                source_target = source_link.hyperlink.target if source_link.hyperlink else None
                output_target = link.hyperlink.target if link.hyperlink else None
                changed_data = any(
                    source_sheet.cell(source_row, HEADERS[label]).value != output_sheet.cell(source_row, HEADERS[label]).value
                    for label in EDITABLE_FIELDS.values()
                )
                if not changed_data and (link.value != source_link.value or output_target != source_target):
                    raise RuntimeError(f"link_changed:{number}")
                if changed_data and (not link.hyperlink or link.hyperlink.target != Path(str(record["pdf"])).as_uri()):
                    raise RuntimeError(f"link_invalid:{number}")
        elif not link.hyperlink or link.hyperlink.target != Path(str(record["pdf"])).as_uri():
            raise RuntimeError(f"link_invalid:{number}")
        if output_sheet.cell(row, STATUS_COLUMN).value != statuses[number]:
            raise RuntimeError(f"status_invalid:{number}")
    if output_sheet.cell(3, STATUS_COLUMN).value != STATUS_HEADER:
        raise RuntimeError("status_header_invalid")
    sheet_path = _sheet_xml_path(source, SHEET)
    source_ext, output_ext = _extension_block(source, sheet_path), _extension_block(staged, sheet_path)
    if source_ext != output_ext:
        raise RuntimeError("x14_extensions_changed")
    source_standard_cf = _standard_cf_blocks(source, sheet_path)
    output_standard_cf = _standard_cf_blocks(staged, sheet_path)
    if source_standard_cf != output_standard_cf:
        raise RuntimeError("native_conditional_formatting_changed")
    return {
        "x14_preserved": True,
        "native_cf_preserved": True,
        "standard_cf_blocks": len(output_standard_cf),
        "h_are_dates": {
            f"H{_row_by_number(output_sheet, number)}": isinstance(
                output_sheet.cell(_row_by_number(output_sheet, number), 8).value,
                datetime,
            )
            for number in records
            if outcomes[number] != "review_conflict"
            if any(record.get(field) is not None for field in _TRANSFER_FIELDS)
        },
    }


def apply(records: dict[str, dict[str, object]], source: Path, output: Path, source_hash: str) -> dict[str, Any]:
    """Write and validate a staged workbook, then atomically publish it."""
    output.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{output.stem}.", suffix=".staged.xlsx", dir=output.parent)
    os.close(descriptor)
    staged = Path(temporary_name)
    try:
        workbook = load_workbook(source)
        sheet = workbook[SHEET]
        mutated = False
        existing_status_header = sheet.cell(3, STATUS_COLUMN).value
        if existing_status_header not in (None, "", STATUS_HEADER):
            raise RuntimeError(f"status_column_occupied:{sheet.cell(3, STATUS_COLUMN).coordinate}")
        status_header = sheet.cell(3, STATUS_COLUMN)
        if status_header.value != STATUS_HEADER:
            status_header.value = STATUS_HEADER
            mutated = True
        header_reference = sheet.cell(3, HEADERS["Ссылка на документ"])
        header_style = copy(header_reference._style)
        if (header_reference.has_style and status_header._style != header_style) or (
            not header_reference.has_style and status_header.has_style
        ):
            status_header._style = header_style
            mutated = True
        if sheet.column_dimensions[status_header.column_letter].width != 58:
            sheet.column_dimensions[status_header.column_letter].width = 58
            mutated = True
        conflicts, changes = [], []
        statuses: dict[str, str | None] = {}
        outcomes: dict[str, str] = {}
        for number, record in records.items():
            # Keep publication truth per row.  A discovered review is useful to
            # an operator, but is not itself a workbook mutation.
            row_mutated = False
            merge_issue_messages = [
                str(issue["message"])
                for issue in record.get("merge_issues", [])
                if isinstance(issue, dict) and issue.get("message")
            ]
            matches = _rows_by_number(sheet, number)
            ambiguous = _ambiguous_rows_by_number(sheet, number)
            if len(matches) > 1 or ambiguous:
                rows = matches if len(matches) > 1 else ambiguous
                existing = "несколько строк Excel" if len(matches) > 1 else "несколько номеров РНС в ячейке Excel"
                conflicts.append({"number": number, "cell": ",".join(f"F{row}" for row in rows), "field": "Номер РНС", "existing": existing, "pdf": number, "action": "review_conflict"})
                outcomes[number] = "review_conflict"
                statuses[number] = None
                changes.append({"number": number, "row": None, "new": False, "outcome": "review_conflict", "written": [], "physical_mutation": False, "document": record["filename"], "end": record.get("end"), "status": None, "issues": ["Несколько строк Excel содержат этот номер РНС; перенос не выполнен."]})
                continue
            row = matches[0] if matches else None
            if record.get("existing_only") and row is None:
                conflicts.append({"number": number, "cell": "F", "field": "Номер РНС", "existing": "строка Excel не найдена", "pdf": number, "action": "review_conflict"})
                outcomes[number] = "review_conflict"
                statuses[number] = None
                changes.append({"number": number, "row": None, "new": False, "outcome": "review_conflict", "written": [], "physical_mutation": False, "document": record["filename"], "end": record.get("end"), "status": None, "issues": ["Изменение/продление содержит номер РНС, но строка Excel не найдена; новая строка не создана."]})
                continue
            new = row is None
            if not any(record.get(field) is not None for field in _TRANSFER_FIELDS):
                outcomes[number] = "review"
                statuses[number] = sheet.cell(row, STATUS_COLUMN).value if row is not None else None
                issues = merge_issue_messages or [
                    "В документе не найдено ни одного подтверждённого поля для переноса."
                ]
                if row is not None:
                    status = sheet.cell(row, STATUS_COLUMN)
                    status_changed = _set_status_presentation(status, sheet.cell(row, HEADERS["Примечание"]))
                    mutated = status_changed or mutated
                    row_mutated = status_changed or row_mutated
                    value = _status_value(status.value, issues)
                    if status.value != value:
                        status.value = value
                        mutated = True
                        row_mutated = True
                    statuses[number] = status.value
                changes.append({"number": number, "row": row, "new": False, "outcome": "review", "written": [], "physical_mutation": row_mutated, "document": record["filename"], "end": record.get("end"), "status": statuses[number], "issues": issues})
                continue
            if new:
                row = _next_data_row(sheet)
                _copy_row_style(sheet, row - 1, row)
                ids = [sheet.cell(item, 1).value for item in range(4, row) if isinstance(sheet.cell(item, 1).value, int)]
                sheet.cell(row, 1).value = max(ids, default=0) + 1
                mutated = True
                row_mutated = True
            mapping = {"Номер этапа": record.get("stage"), "Наименование объекта": record.get("object"), "Номер РНС": number if new else None, "Дата выдачи": iso_date(record.get("issue")), "Срок действия": iso_date(record.get("end")), "Дата последн. измен.": iso_date(record.get("changed")), "Орган выдачи": record.get("issuer"), "Застройщик": record.get("builder"), "Субъект РФ": record.get("region"), "Муниципальный р-н": record.get("district"), "Разработчик ПД": record.get("developer")}
            issues, written = list(merge_issue_messages), []
            quality = record.get("field_quality", {})
            for label, value in mapping.items():
                if label == "Номер РНС" and not new:
                    continue
                if value is None:
                    continue
                target = sheet.cell(row, HEADERS[label])
                field_key = next((key for key, mapped_label in EDITABLE_FIELDS.items() if mapped_label == label), None)
                quality_item = quality.get(field_key, {}) if isinstance(quality, dict) and field_key else {}
                if isinstance(quality_item, dict) and quality_item.get("status") != "actionable" and quality_item.get("status") is not None:
                    # A manually confirmed value can match a weak OCR value.
                    # Do not recreate a generated warning (or publish AA) for
                    # that semantic no-op; preserve the operator marker.
                    if target.value in (None, "") or not _same_value(label, target.value, value):
                        issues.append(f"Не перенесено «{label}»: значение PDF требует ручной сверки.")
                    # Keep the usual conflict projection for an occupied,
                    # substantively different cell. Server-side quality policy
                    # then exposes it as review-only, never as a proposal.
                    if target.value not in (None, ""):
                        wrote = _put(target, value, label, number, conflicts)
                        mutated = wrote or mutated
                        row_mutated = wrote or row_mutated
                    continue
                if issue := transfer_issue(label, target.value, value):
                    issues.append(issue)
                if _put(target, value, label, number, conflicts):
                    written.append(target.coordinate)
                    mutated = True
                    row_mutated = True
            outcome = _change_outcome(new, written, issues)
            outcomes[number] = outcome
            link = sheet.cell(row, HEADERS["Ссылка на документ"])
            status = sheet.cell(row, STATUS_COLUMN)
            # A review with no transferable values must not overwrite the
            # provenance link of an existing row. The review source is exposed
            # by the job capability instead; an explicit proposal approval may
            # still update W later.
            if outcome != "already_present" and (new or written):
                link_value, link_target = str(record["filename"]), Path(str(record["pdf"])).as_uri()
                if link.value != link_value or not link.hyperlink or link.hyperlink.target != link_target or link.style != "Hyperlink":
                    link.value, link.hyperlink, link.style = link_value, link_target, "Hyperlink"
                    mutated = True
                    row_mutated = True
            if outcome != "already_present":
                status_changed = _set_status_presentation(status, sheet.cell(row, HEADERS["Примечание"]))
                mutated = status_changed or mutated
                row_mutated = status_changed or row_mutated
                value = _status_value(status.value, issues)
                if status.value != value:
                    status.value = value
                    mutated = True
                    row_mutated = True
            statuses[number] = status.value
            changes.append({
                "number": number,
                "row": row,
                "new": new,
                "outcome": outcome,
                "written": written,
                "physical_mutation": row_mutated,
                "document": record["filename"],
                "end": record.get("end"),
                "status": status.value,
                "issues": issues,
            })
        if not mutated:
            if sha256(source) != source_hash:
                raise RuntimeError("source_xlsx_changed")
            verification = _validate(source, source, records, statuses, outcomes)
            if source.resolve() != output.resolve():
                shutil.copy2(source, staged)
                if sha256(staged) != source_hash:
                    raise RuntimeError("no_op_copy_changed")
                os.replace(staged, output)
            return {"changes": changes, "conflicts": conflicts, "verification": verification, "published": False}
        workbook.save(staged)
        _reinject_native_conditional_formatting(source, staged)
        _reinject_extensions(source, staged)
        if sha256(source) != source_hash:
            raise RuntimeError("source_xlsx_changed")
        verification = _validate(source, staged, records, statuses, outcomes)
        os.replace(staged, output)
        return {"changes": changes, "conflicts": conflicts, "verification": verification, "published": True}
    finally:
        staged.unlink(missing_ok=True)


def apply_proposal(source: Path, output: Path, source_hash: str, number: str, field: str, value: object, pdf: Path) -> None:
    """Stage exactly one approved existing-row value without weakening validation."""
    if field not in HEADERS:
        raise RuntimeError("proposal_field_invalid")
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{output.stem}.", suffix=".staged.xlsx", dir=output.parent)
    os.close(descriptor)
    staged = Path(temporary_name)
    try:
        workbook = load_workbook(source)
        sheet = workbook[SHEET]
        rows = _rows_by_number(sheet, number)
        if len(rows) != 1:
            raise RuntimeError("proposal_row_unavailable")
        row = rows[0]
        cell = sheet.cell(row, HEADERS[field])
        cell.value = value
        if isinstance(value, str):
            cell.data_type = "s"
        link = sheet.cell(row, HEADERS["Ссылка на документ"])
        link.value, link.hyperlink, link.style = pdf.name, pdf.as_uri(), "Hyperlink"
        status = sheet.cell(row, STATUS_COLUMN)
        if isinstance(status.value, str):
            marker = f"Не перенесено «{field}»:"
            status.value = "\n".join(line for line in status.value.splitlines() if not line.startswith(marker)) or None
        workbook.save(staged)
        _reinject_native_conditional_formatting(source, staged)
        _reinject_extensions(source, staged)
        if sha256(source) != source_hash:
            raise RuntimeError("proposal_target_stale")
        source_book, staged_book = load_workbook(source, data_only=False), load_workbook(staged, data_only=False)
        source_sheet, staged_sheet = source_book[SHEET], staged_book[SHEET]
        for current_row in range(4, source_sheet.max_row + 1):
            for column in range(1, source_sheet.max_column + 1):
                old, new = source_sheet.cell(current_row, column), staged_sheet.cell(current_row, column)
                allowed = current_row == row and old.coordinate in {cell.coordinate, link.coordinate, status.coordinate}
                old_link = old.hyperlink.target if old.hyperlink else None
                new_link = new.hyperlink.target if new.hyperlink else None
                if not allowed and (old.value != new.value or old._style != new._style or old_link != new_link):
                    raise RuntimeError(f"proposal_changed:{old.coordinate}")
        sheet_path = _sheet_xml_path(source, SHEET)
        if _extension_block(source, sheet_path) != _extension_block(staged, sheet_path):
            raise RuntimeError("x14_extensions_changed")
        if _standard_cf_blocks(source, sheet_path) != _standard_cf_blocks(staged, sheet_path):
            raise RuntimeError("native_conditional_formatting_changed")
        os.replace(staged, output)
    finally:
        staged.unlink(missing_ok=True)


def editable_field_values(source: Path, row: int, number: str) -> dict[str, str]:
    """Read a compact, allowlisted value contract for one exact RNS row."""
    workbook = load_workbook(source, data_only=False)
    sheet = workbook[SHEET]
    if row < 4 or _row_by_number(sheet, number) != row:
        raise RuntimeError("manual_row_unavailable")
    return {key: _value_text(sheet.cell(row, HEADERS[label]).value) if sheet.cell(row, HEADERS[label]).value not in (None, "") else "" for key, label in EDITABLE_FIELDS.items()}


def apply_manual_edit(
    source: Path,
    output: Path,
    source_hash: str,
    row: int,
    number: str,
    fields: dict[str, object],
) -> None:
    """Stage a bounded manual row correction without changing row provenance."""
    if not fields or any(key not in EDITABLE_FIELDS for key in fields):
        raise RuntimeError("manual_field_invalid")
    if sha256(source) != source_hash:
        raise RuntimeError("manual_target_stale")
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{output.stem}.", suffix=".manual-staged.xlsx", dir=output.parent)
    os.close(descriptor)
    staged = Path(temporary_name)
    try:
        workbook = load_workbook(source)
        sheet = workbook[SHEET]
        if row < 4 or _row_by_number(sheet, number) != row:
            raise RuntimeError("manual_row_unavailable")
        for key, value in fields.items():
            if key in DATE_EDIT_FIELDS and value is not None and not isinstance(value, datetime):
                raise RuntimeError("manual_date_invalid")
            if key not in DATE_EDIT_FIELDS and value is not None and not isinstance(value, str):
                raise RuntimeError("manual_value_invalid")
            cell = sheet.cell(row, HEADERS[EDITABLE_FIELDS[key]])
            cell.value = value
            if isinstance(value, str):
                cell.data_type = "s"  # A manual value is never an Excel formula.
        status = sheet.cell(row, STATUS_COLUMN)
        marker = "Исправлено вручную: " + ", ".join(f"«{EDITABLE_FIELDS[key]}»" for key in fields) + "."
        old_lines = str(status.value).splitlines() if status.value not in (None, "") else []
        stale_markers = tuple(f"Не перенесено «{EDITABLE_FIELDS[key]}»:" for key in fields)
        status.value = "\n".join([line for line in old_lines if not line.startswith(stale_markers)] + [marker])
        workbook.save(staged)
        _reinject_native_conditional_formatting(source, staged)
        _reinject_extensions(source, staged)
        if sha256(source) != source_hash:
            raise RuntimeError("manual_target_stale")
        source_book, staged_book = load_workbook(source, data_only=False), load_workbook(staged, data_only=False)
        source_sheet, staged_sheet = source_book[SHEET], staged_book[SHEET]
        allowed = {source_sheet.cell(row, HEADERS[EDITABLE_FIELDS[key]]).coordinate for key in fields}
        allowed.add(source_sheet.cell(row, STATUS_COLUMN).coordinate)
        for current_row in range(4, source_sheet.max_row + 1):
            for column in range(1, source_sheet.max_column + 1):
                old, new = source_sheet.cell(current_row, column), staged_sheet.cell(current_row, column)
                old_link = old.hyperlink.target if old.hyperlink else None
                new_link = new.hyperlink.target if new.hyperlink else None
                if old.coordinate not in allowed and (old.value != new.value or old._style != new._style or old_link != new_link):
                    raise RuntimeError(f"manual_changed:{old.coordinate}")
                if isinstance(old.value, str) and old.value.startswith("=") and old.value != new.value:
                    raise RuntimeError(f"manual_formula_changed:{old.coordinate}")
        # W, Y:Z and the canonical number are intentionally not in ``allowed``.
        sheet_path = _sheet_xml_path(source, SHEET)
        if _extension_block(source, sheet_path) != _extension_block(staged, sheet_path):
            raise RuntimeError("x14_extensions_changed")
        if _standard_cf_blocks(source, sheet_path) != _standard_cf_blocks(staged, sheet_path):
            raise RuntimeError("native_conditional_formatting_changed")
        os.replace(staged, output)
    finally:
        staged.unlink(missing_ok=True)
