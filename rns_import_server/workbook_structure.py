"""Read-only workbook structural evidence for native publication."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class WorkbookStructure:
    sheet: str
    max_row: int
    max_column: int
    merges: tuple[str, ...]
    defined_names: tuple[str, ...]
    auto_filter: str | None
    formula_cells: tuple[str, ...]


def inspect_workbook(path: Path, sheet_name: str) -> WorkbookStructure:
    """Open read-only and never save; unsupported protected books fail closed."""
    book = load_workbook(path, read_only=False, data_only=False)
    try:
        if sheet_name not in book.sheetnames:
            raise RuntimeError("workbook_sheet_missing")
        sheet = book[sheet_name]
        if book.security.lockStructure or sheet.protection.sheet:
            raise RuntimeError("workbook_structure_protected")
        formulas = tuple(
            cell.coordinate for row in sheet.iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        return WorkbookStructure(
            sheet=sheet_name, max_row=sheet.max_row, max_column=sheet.max_column,
            merges=tuple(sorted(str(item) for item in sheet.merged_cells.ranges)),
            defined_names=tuple(sorted(item.name for item in book.defined_names.values())),
            auto_filter=sheet.auto_filter.ref, formula_cells=formulas,
        )
    finally:
        book.close()


def insertion_is_structurally_safe(structure: WorkbookStructure, row: int) -> bool:
    """Do not split a vertical merge; native Excel is still required afterwards."""
    if row < 1 or row > structure.max_row + 1:
        return False
    for range_text in structure.merges:
        from openpyxl.worksheet.cell_range import CellRange
        cell_range = CellRange(range_text)
        if cell_range.min_row < row <= cell_range.max_row:
            return False
    return True
