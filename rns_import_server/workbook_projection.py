"""Fail-closed, read-only workbook projections for the Wave3 planning ports.

This module deliberately has no durable-path producer.  A later gate must
inject :class:`WorkbookProjectionAuthority` only after it has proved the
target, workbook contract, sheet and template evidence together.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from enum import StrEnum
import hashlib
from io import BytesIO
import os
from pathlib import Path
import stat
from types import MappingProxyType
from typing import Protocol
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from rns_import_server.group_provisioning import GroupProvisioningProjection, ProvisioningRow
from rns_import_server.workbook_groups import SheetProjection, SheetRow


_BUSINESS_COLUMNS = tuple(range(1, 25)) + (27,)
_SUPPORTED_CELL_TYPES = (type(None), bool, int, float, str, date, datetime, time)


class WorkbookProjectionCode(StrEnum):
    OK = "ok"
    INVALID_AUTHORITY = "invalid_authority"
    UNSAFE_TARGET = "unsafe_target"
    SOURCE_UNSTABLE = "source_unstable"
    SOURCE_UNREADABLE = "source_unreadable"
    SHEET_MISMATCH = "sheet_mismatch"
    TEMPLATE_MISMATCH = "template_mismatch"
    UNSUPPORTED_CELL = "unsupported_cell"
    GROUP_OWNERSHIP_MISSING = "group_ownership_missing"
    GROUP_OWNERSHIP_CONFLICT = "group_ownership_conflict"


@dataclass(frozen=True)
class TemplateCellEvidence:
    """Exact template cell evidence supplied by the future authority producer."""

    row: int
    column: int
    value: object


@dataclass(frozen=True)
class GroupOwnershipEvidence:
    """Exact per-row ownership proof; projection never guesses this fact."""

    row: int
    owned: bool


@dataclass(frozen=True, repr=False)
class WorkbookProjectionAuthority:
    """Immutable verified input; ``_target_path`` is private implementation data.

    IDs are evidence received from an authority producer, never inferred from
    the path or workbook bytes here.
    """

    target_identity: str
    workbook_contract_id: str
    sheet_identity: str
    template_version: str
    registry_generation: int
    template_cells: tuple[TemplateCellEvidence, ...]
    group_ownership: tuple[GroupOwnershipEvidence, ...]
    _target_path: str

    @classmethod
    def verified(
        cls,
        *,
        target_path: str,
        target_identity: str,
        workbook_contract_id: str,
        sheet_identity: str,
        template_version: str,
        registry_generation: int,
        template_cells: tuple[TemplateCellEvidence, ...],
        group_ownership: tuple[GroupOwnershipEvidence, ...],
    ) -> "WorkbookProjectionAuthority":
        return cls(
            target_identity=target_identity,
            workbook_contract_id=workbook_contract_id,
            sheet_identity=sheet_identity,
            template_version=template_version,
            registry_generation=registry_generation,
            template_cells=template_cells,
            group_ownership=group_ownership,
            _target_path=target_path,
        )


class WorkbookProjectionAuthorityPort(Protocol):
    def read_authority(self) -> WorkbookProjectionAuthority: ...


@dataclass(frozen=True)
class SourceObservation:
    device: int
    inode: int
    size: int
    mtime_ns: int
    ctime_ns: int


@dataclass(frozen=True)
class WorkbookProjectionSnapshot:
    """The single stable read which backs both consumer-specific projections."""

    authority_identity: str
    workbook_contract_id: str
    pre_hash: str
    registry_generation: int
    source_observation: SourceObservation
    sheet: SheetProjection
    provisioning: GroupProvisioningProjection


@dataclass(frozen=True)
class WorkbookProjectionResult:
    code: WorkbookProjectionCode
    snapshot: WorkbookProjectionSnapshot | None = None


def _valid_scalar(value: object) -> bool:
    return type(value) in _SUPPORTED_CELL_TYPES


def _authority_code(authority: object) -> WorkbookProjectionCode | None:
    if type(authority) is not WorkbookProjectionAuthority:
        return WorkbookProjectionCode.INVALID_AUTHORITY
    strings = (
        authority.target_identity,
        authority.workbook_contract_id,
        authority.sheet_identity,
        authority.template_version,
        authority._target_path,
    )
    if any(type(value) is not str or not value for value in strings):
        return WorkbookProjectionCode.INVALID_AUTHORITY
    if type(authority.registry_generation) is not int or authority.registry_generation < 0:
        return WorkbookProjectionCode.INVALID_AUTHORITY
    if type(authority.template_cells) is not tuple or not authority.template_cells:
        return WorkbookProjectionCode.INVALID_AUTHORITY
    if type(authority.group_ownership) is not tuple or not authority.group_ownership:
        return WorkbookProjectionCode.GROUP_OWNERSHIP_MISSING
    seen: set[tuple[int, int]] = set()
    for evidence in authority.template_cells:
        if type(evidence) is not TemplateCellEvidence:
            return WorkbookProjectionCode.INVALID_AUTHORITY
        if type(evidence.row) is not int or type(evidence.column) is not int:
            return WorkbookProjectionCode.INVALID_AUTHORITY
        if evidence.row < 1 or evidence.column < 1 or not _valid_scalar(evidence.value):
            return WorkbookProjectionCode.INVALID_AUTHORITY
        key = (evidence.row, evidence.column)
        if key in seen:
            return WorkbookProjectionCode.INVALID_AUTHORITY
        seen.add(key)
    owned_rows: set[int] = set()
    for evidence in authority.group_ownership:
        if type(evidence) is not GroupOwnershipEvidence or type(evidence.row) is not int or type(evidence.owned) is not bool:
            return WorkbookProjectionCode.INVALID_AUTHORITY
        if evidence.row < 1:
            return WorkbookProjectionCode.INVALID_AUTHORITY
        if evidence.row in owned_rows:
            return WorkbookProjectionCode.GROUP_OWNERSHIP_CONFLICT
        owned_rows.add(evidence.row)
    return None


def _canonical_regular_path(path: str) -> Path | None:
    if not os.path.isabs(path) or os.path.normpath(path) != path:
        return None
    candidate = Path(path)
    try:
        for component in (candidate, *candidate.parents):
            if component.is_symlink():
                return None
        if not candidate.is_file() or candidate.is_symlink():
            return None
    except OSError:
        return None
    return candidate


def _observation(stat_result: os.stat_result) -> SourceObservation:
    return SourceObservation(
        stat_result.st_dev, stat_result.st_ino, stat_result.st_size, stat_result.st_mtime_ns, stat_result.st_ctime_ns
    )


def _read_bound_bytes(path: Path) -> tuple[SourceObservation, bytes] | None:
    """Bind one regular source descriptor, then hash and parse precisely its bytes."""
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0)
    try:
        descriptor = os.open(path, flags)
    except OSError:
        return None
    try:
        before = os.fstat(descriptor)
        if not stat.S_ISREG(before.st_mode):
            return None
        chunks: list[bytes] = []
        while True:
            chunk = os.read(descriptor, 1024 * 1024)
            if not chunk:
                break
            chunks.append(chunk)
        if os.fstat(descriptor) != before:
            return None
        return _observation(before), b"".join(chunks)
    except OSError:
        return None
    finally:
        os.close(descriptor)


def _path_matches(path: Path, observed: SourceObservation) -> bool:
    verified = _canonical_regular_path(str(path))
    if verified is None:
        return False
    try:
        return _observation(os.lstat(verified)) == observed
    except OSError:
        return False


def _same_scalar(left: object, right: object) -> bool:
    return type(left) is type(right) and left == right


def _cell_value(cell: object) -> object:
    value = cell.value  # type: ignore[attr-defined]
    if not _valid_scalar(value):
        raise TypeError("unsupported cell value")
    return value


def _project_rows(
    worksheet: object, group_ownership: dict[int, bool]
) -> tuple[tuple[SheetRow, ...], tuple[ProvisioningRow, ...]]:
    sheet_rows: list[SheetRow] = []
    provisioning_rows: list[ProvisioningRow] = []
    max_row = worksheet.max_row  # type: ignore[attr-defined]
    for number in range(1, max_row + 1):
        cells = {column: worksheet.cell(number, column) for column in _BUSINESS_COLUMNS}  # type: ignore[attr-defined]
        values = {column: _cell_value(cell) for column, cell in cells.items()}
        a_to_f = tuple(values[column] for column in range(1, 7))
        has_business = any(value not in (None, "") for value in values.values())
        preformatted = all(bool(getattr(cell, "has_style", False)) for cell in cells.values())
        is_business_row = has_business or preformatted
        sheet_rows.append(
            SheetRow(
                number, *a_to_f, is_business_row=is_business_row, is_preformatted=preformatted,
                group_ownership_proven=group_ownership[number],
            )
        )
        provisioning_rows.append(
            ProvisioningRow(number, MappingProxyType(values), is_business_row=is_business_row, is_preformatted=preformatted)
        )
    return tuple(sheet_rows), tuple(provisioning_rows)


def project_workbook(authority: WorkbookProjectionAuthority) -> WorkbookProjectionResult:
    """Read one verified workbook snapshot and produce both planning projections."""
    authority_code = _authority_code(authority)
    if authority_code is not None:
        return WorkbookProjectionResult(authority_code)
    path = _canonical_regular_path(authority._target_path)
    if path is None:
        return WorkbookProjectionResult(WorkbookProjectionCode.UNSAFE_TARGET)
    bound = _read_bound_bytes(path)
    if bound is None:
        return WorkbookProjectionResult(WorkbookProjectionCode.SOURCE_UNREADABLE)
    before, source_bytes = bound
    if not _path_matches(path, before):
        return WorkbookProjectionResult(WorkbookProjectionCode.SOURCE_UNSTABLE)
    pre_hash = hashlib.sha256(source_bytes).hexdigest()

    book = None
    try:
        book = load_workbook(BytesIO(source_bytes), read_only=True, data_only=False)
        if authority.sheet_identity not in book.sheetnames:
            return WorkbookProjectionResult(WorkbookProjectionCode.SHEET_MISMATCH)
        worksheet = book[authority.sheet_identity]
        for evidence in authority.template_cells:
            if not _same_scalar(_cell_value(worksheet.cell(evidence.row, evidence.column)), evidence.value):
                return WorkbookProjectionResult(WorkbookProjectionCode.TEMPLATE_MISMATCH)
        ownership = {item.row: item.owned for item in authority.group_ownership}
        worksheet_rows = set(range(1, worksheet.max_row + 1))
        if set(ownership) != worksheet_rows:
            return WorkbookProjectionResult(WorkbookProjectionCode.GROUP_OWNERSHIP_MISSING)
        sheet_rows, provisioning_rows = _project_rows(worksheet, ownership)
    except TypeError:
        return WorkbookProjectionResult(WorkbookProjectionCode.UNSUPPORTED_CELL)
    except (OSError, ValueError, KeyError, RuntimeError, AttributeError, BadZipFile, InvalidFileException, ParseError):
        return WorkbookProjectionResult(WorkbookProjectionCode.SOURCE_UNREADABLE)
    finally:
        if book is not None:
            book.close()
    if not _path_matches(path, before):
        return WorkbookProjectionResult(WorkbookProjectionCode.SOURCE_UNSTABLE)

    sheet = SheetProjection.from_rows(
        authority.target_identity, pre_hash, authority.registry_generation, sheet_rows
    )
    provisioning = GroupProvisioningProjection(
        authority.target_identity, pre_hash, authority.registry_generation, provisioning_rows
    )
    return WorkbookProjectionResult(
        WorkbookProjectionCode.OK,
        WorkbookProjectionSnapshot(
            authority.target_identity,
            authority.workbook_contract_id,
            pre_hash,
            authority.registry_generation,
            before,
            sheet,
            provisioning,
        ),
    )


class WorkbookProjectionAdapter:
    """Read-only port adapter; a later gate owns authority production."""

    def __init__(self, authority: WorkbookProjectionAuthorityPort) -> None:
        self._authority = authority

    def read(self) -> WorkbookProjectionResult:
        try:
            authority = self._authority.read_authority()
        except Exception:
            return WorkbookProjectionResult(WorkbookProjectionCode.INVALID_AUTHORITY)
        return project_workbook(authority)
