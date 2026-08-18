"""Fail-closed coordinator for blank fill and Windows-native middle insertion."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import shutil

from openpyxl import load_workbook

from rns_import_server.audit import sha256
from rns_import_server.excel_native import NativeExcelError, NativeInsertRequest, native_excel_available, run_native_insert
from rns_import_server.workbook_groups import MutationPlan
from rns_import_server.workbook_mutation_manifest import ALLOWLISTED_COLUMNS, manifest_for, validate_insertion
from rns_import_server.workbook_structure import inspect_workbook, insertion_is_structurally_safe


class GroupRowInsertionError(RuntimeError):
    def __init__(self, code: str, *, stage: str, cause: BaseException | None = None):
        self.code, self.stage, self.cause = code, stage, cause
        super().__init__(f"{code}@{stage}" + (f": {cause}" if cause else ""))


@dataclass(frozen=True)
class GroupRowRequest:
    plan: MutationPlan
    source: Path
    output: Path
    sheet: str
    fields: dict[int, object]
    hyperlink: str | None = None


def _write_allowlisted(path: Path, sheet_name: str, row: int, fields: dict[int, object], hyperlink: str | None) -> None:
    if any(column not in ALLOWLISTED_COLUMNS for column in fields):
        raise GroupRowInsertionError("group_row_field_not_allowed", stage="write")
    book = load_workbook(path)
    try:
        sheet = book[sheet_name]
        for column, value in fields.items():
            sheet.cell(row, column).value = value
        if hyperlink is not None:
            cell = sheet.cell(row, 23)
            cell.value, cell.hyperlink, cell.style = hyperlink.rsplit("/", 1)[-1], hyperlink, "Hyperlink"
        book.save(path)
    finally:
        book.close()


def publish_group_row(request: GroupRowRequest, *, native_script: Path, operation_directory: Path) -> dict[str, object]:
    """Publish only proven blank fills; middle insertion requires desktop Excel."""
    plan = request.plan
    if sha256(request.source) != plan.workbook_hash:
        raise GroupRowInsertionError("workbook_pre_hash_mismatch", stage="preflight")
    if plan.mode == "existing_blank":
        staged = request.output.with_suffix(".blank.staged.xlsx")
        shutil.copy2(request.source, staged)
        try:
            _write_allowlisted(staged, request.sheet, plan.target_row, request.fields, request.hyperlink)
            if sha256(request.source) != plan.workbook_hash:
                raise GroupRowInsertionError("workbook_pre_hash_mismatch", stage="pre_publish")
            shutil.move(staged, request.output)
            return {"mode": "blank_fill", "row": plan.target_row, "published": True}
        finally:
            staged.unlink(missing_ok=True)
    if plan.mode != "insert_before_header":
        raise GroupRowInsertionError("group_row_plan_invalid", stage="preflight")
    # Hosted runners must expose the portable negative contract before any
    # staging or structural work.  This also makes clear that no fallback exists.
    if not native_excel_available():
        raise GroupRowInsertionError("excel_required_for_middle_insert", stage="pre_open")
    structure = inspect_workbook(request.source, request.sheet)
    if not insertion_is_structurally_safe(structure, plan.target_row):
        raise GroupRowInsertionError("group_row_structure_unsafe", stage="preflight")
    operation_directory.mkdir(parents=True, exist_ok=True)
    control, candidate = operation_directory / "control.xlsx", operation_directory / "candidate.xlsx"
    shutil.copy2(request.source, control); shutil.copy2(request.source, candidate)
    native = NativeInsertRequest("pending", "pending", "pending", control, candidate, plan.target_row,
                                 operation_directory / "excel-lease.json", operation_directory / "lease-ack.json")
    try:
        run_native_insert(native, script=native_script)
        _write_allowlisted(candidate, request.sheet, plan.target_row, request.fields, request.hyperlink)
        control_manifest = manifest_for(control, request.sheet)
        candidate_manifest = manifest_for(candidate, request.sheet, insertion_row=plan.target_row)
        validate_insertion(control_manifest, candidate_manifest, plan.target_row)
    except NativeExcelError as error:
        raise GroupRowInsertionError(error.code, stage=error.stage, cause=error) from error
    except Exception as error:
        raise GroupRowInsertionError("group_row_oracle_failed", stage="validation", cause=error) from error
    if sha256(request.source) != plan.workbook_hash:
        raise GroupRowInsertionError("workbook_pre_hash_mismatch", stage="pre_publish")
    shutil.move(candidate, request.output)
    return {"mode": "middle_insert", "row": plan.target_row, "published": True, "manifest": candidate_manifest.digest}
