"""Deterministic, read-only semantic manifests for paired Excel outputs."""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from rns_import_server.audit import digest


ALLOWLISTED_COLUMNS = tuple(range(1, 25)) + (27,)  # A:X and AA
_REGISTRY_SHEET = "Реестр РНС"
_FIRST_REQUEST_COLUMN = 1
_LAST_REQUEST_COLUMN = 43  # AQ
_FORMULA_COLUMNS = (25, 26)  # Y:Z
_KNOWN_INSERTION_ROWS = frozenset((6, 10, 104))
_REGISTRY_TOKEN = "'Реестр РНС'!"


class MutationManifestError(RuntimeError):
    """Stable, blocking evidence failure for the native insertion gate."""

    def __init__(self, code: str, *, subject: str, field: str):
        self.code, self.subject, self.field = code, subject, field
        super().__init__(f"{code}:{subject}:{field}")


def _fail(code: str, subject: str, field: str) -> None:
    raise MutationManifestError(code, subject=subject, field=field)


@dataclass(frozen=True)
class MutationManifest:
    version: str
    sheet: str
    insertion_row: int | None
    max_row: int
    values: dict[str, object]
    formulas: dict[str, str]
    hyperlinks: dict[str, str]
    digest: str


def manifest_for(path: Path, sheet_name: str, *, insertion_row: int | None = None) -> MutationManifest:
    # ``read_only`` worksheets omit relationships, including hyperlinks.  A
    # blank-fill gate must prove those links, so this stays read-only in
    # behavior but loads a normal workbook view.
    book = load_workbook(path, read_only=False, data_only=False)
    try:
        sheet = book[sheet_name]
        values: dict[str, object] = {}
        formulas: dict[str, str] = {}
        links: dict[str, str] = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    (formulas if isinstance(cell.value, str) and cell.value.startswith("=") else values)[cell.coordinate] = cell.value
                hyperlink = getattr(cell, "hyperlink", None)
                if hyperlink and hyperlink.target:
                    links[cell.coordinate] = hyperlink.target
        payload = {"version": "native-group-row-insertion-v1", "sheet": sheet_name, "insertion_row": insertion_row,
                   "max_row": sheet.max_row, "values": values, "formulas": formulas, "hyperlinks": links}
        return MutationManifest(**payload, digest=digest(payload))
    finally:
        book.close()


def validate_insertion(control: MutationManifest, candidate: MutationManifest, row: int) -> None:
    """Prove exactly one physical row and forbid edits outside the mapped insert."""
    if control.sheet != candidate.sheet or candidate.max_row != control.max_row + 1:
        raise RuntimeError("mutation_manifest_row_count_mismatch")
    if row < 1:
        raise RuntimeError("mutation_manifest_insert_row_invalid")
    for source, target in ((control.values, candidate.values), (control.formulas, candidate.formulas), (control.hyperlinks, candidate.hyperlinks)):
        for coordinate, value in source.items():
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            column, source_row = coordinate_from_string(coordinate)
            mapped = f"{column}{source_row if source_row < row else source_row + 1}"
            expected = value
            if source is control.formulas:
                from openpyxl.formula.translate import Translator
                expected = Translator(value, origin=coordinate).translate_formula(mapped)
            if target.get(mapped) != expected and not (column == "A" and source_row >= row):
                raise RuntimeError(f"mutation_manifest_changed:{coordinate}")


def validate_control(original: MutationManifest, control: MutationManifest) -> None:
    """A control can normalize caches, but cannot change semantic workbook data."""
    if original.sheet != control.sheet or original.max_row != control.max_row:
        raise RuntimeError("control_manifest_structure_changed")
    if original.values != control.values or original.formulas != control.formulas or original.hyperlinks != control.hyperlinks:
        raise RuntimeError("control_manifest_semantic_changed")


def _row_map(values: Mapping[str, object], row: int) -> dict[int, object]:
    result: dict[int, object] = {}
    for coordinate, value in values.items():
        column, coordinate_row = coordinate_from_string(coordinate)
        if coordinate_row == row:
            result[column_index_from_string(column)] = value
    return result


def _outside_row_map(values: Mapping[str, object], row: int) -> dict[str, object]:
    return {
        coordinate: value for coordinate, value in values.items()
        if coordinate_from_string(coordinate)[1] != row
    }


def validate_blank_fill(
    control: MutationManifest,
    candidate: MutationManifest,
    *,
    target_row: int,
    fields: Mapping[int, object],
    hyperlink: str | None,
) -> None:
    """Prove a native blank-fill changed only exact trusted row fields.

    Comparison is manifest-only and read-only.  It deliberately has no repair
    path: any semantic value, formula, link, sheet, or row-count surprise
    blocks publication at the caller's validation boundary.
    """
    if target_row < 1:
        _fail("blank-fill-target-row-invalid", control.sheet, str(target_row))
    trusted = _field_map(fields)
    if hyperlink is not None and not isinstance(hyperlink, str):
        _fail("blank-fill-hyperlink-invalid", control.sheet, "hyperlink")
    if control.sheet != candidate.sheet:
        _fail("blank-fill-sheet-mismatch", control.sheet, candidate.sheet)
    if control.max_row != candidate.max_row:
        _fail("blank-fill-row-count-mismatch", control.sheet, str(target_row))
    maps = (
        ("value", control.values, candidate.values),
        ("formula", control.formulas, candidate.formulas),
        ("hyperlink", control.hyperlinks, candidate.hyperlinks),
    )
    for kind, before, after in maps:
        if _outside_row_map(before, target_row) != _outside_row_map(after, target_row):
            _fail(f"blank-fill-outside-{kind}-mismatch", control.sheet, str(target_row))

    expected_values = _row_map(control.values, target_row)
    for column, value in trusted.items():
        if value in (None, ""):
            expected_values.pop(column, None)
        else:
            expected_values[column] = value
    if _row_map(candidate.values, target_row) != expected_values:
        _fail("blank-fill-value-mismatch", control.sheet, str(target_row))
    if _row_map(candidate.formulas, target_row) != _row_map(control.formulas, target_row):
        _fail("blank-fill-formula-mismatch", control.sheet, str(target_row))

    # The request owns W only.  Existing links elsewhere in a claimed blank
    # row are not inherited: their presence makes this plan invalid.
    expected_links = {} if hyperlink is None else {23: hyperlink}
    if _row_map(candidate.hyperlinks, target_row) != expected_links:
        _fail("blank-fill-hyperlink-mismatch", control.sheet, str(target_row))


def _field_map(fields: Mapping[int, object]) -> dict[int, object]:
    if not isinstance(fields, Mapping):
        _fail("inserted-row-fields-invalid", "fields", type(fields).__name__)
    result: dict[int, object] = {}
    for column, value in fields.items():
        if isinstance(column, bool) or not isinstance(column, int) or not _FIRST_REQUEST_COLUMN <= column <= _LAST_REQUEST_COLUMN:
            _fail("inserted-row-fields-invalid", "fields", str(column))
        if column in _FORMULA_COLUMNS:
            _fail("inserted-row-formula-field-forbidden", "fields", str(column))
        if isinstance(value, str) and value.startswith("="):
            _fail("inserted-row-formula-value-forbidden", "fields", str(column))
        result[column] = value
    return result


def _style_semantics(cell: object) -> tuple[object, ...]:
    """Resolve style-table IDs into cross-workbook semantic components."""
    return (
        copy(cell.font),
        copy(cell.fill),
        copy(cell.border),
        copy(cell.alignment),
        cell.number_format,
        copy(cell.protection),
        bool(cell.quotePrefix),
        bool(cell.pivotButton),
    )


def _hyperlink_target(cell: object) -> str | None:
    hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink is None:
        return None
    target = getattr(hyperlink, "target", None)
    return target if isinstance(target, str) else None


def validate_inserted_row(
    control_path: Path,
    candidate_path: Path,
    *,
    sheet_name: str,
    insertion_row: int,
    fields: Mapping[int, object],
    hyperlink: str | None,
) -> None:
    """Prove the persisted row is exactly the native request plus Y:Z formulas.

    This is deliberately a read-only check.  It does not infer fields from a
    predecessor or repair a candidate: every persisted value has to be
    attributable to the trusted request.
    """
    if insertion_row < 2:
        _fail("inserted-row-invalid", str(sheet_name), "insertion_row")
    trusted = _field_map(fields)
    if hyperlink is not None and not isinstance(hyperlink, str):
        _fail("inserted-row-hyperlink-invalid", str(sheet_name), "hyperlink")
    control = None
    try:
        control = load_workbook(control_path, read_only=False, data_only=False)
        candidate = load_workbook(candidate_path, read_only=False, data_only=False)
    except Exception as error:
        if control is not None:
            control.close()
        _fail("inserted-row-read-failed", str(sheet_name), type(error).__name__)
    try:
        if sheet_name not in control.sheetnames or sheet_name not in candidate.sheetnames:
            _fail("inserted-row-sheet-missing", str(sheet_name), "sheet")
        source_sheet, target_sheet = control[sheet_name], candidate[sheet_name]
        format_source_row = insertion_row - 1
        if target_sheet.row_dimensions[insertion_row].height != source_sheet.row_dimensions[format_source_row].height:
            _fail("inserted-row-height-mismatch", str(sheet_name), str(insertion_row))
        for column in range(_FIRST_REQUEST_COLUMN, _LAST_REQUEST_COLUMN + 1):
            source_cell = source_sheet.cell(format_source_row, column)
            target_cell = target_sheet.cell(insertion_row, column)
            if _style_semantics(target_cell) != _style_semantics(source_cell):
                _fail("inserted-row-style-mismatch", str(sheet_name), target_cell.coordinate)
            if column in _FORMULA_COLUMNS:
                source_formula = source_cell.value
                try:
                    expected = (
                        Translator(source_formula, origin=source_cell.coordinate).translate_formula(target_cell.coordinate)
                        if isinstance(source_formula, str) and source_formula.startswith("=")
                        else None
                    )
                except Exception as error:
                    failure = MutationManifestError(
                        "inserted-row-formula-translation-invalid",
                        subject=str(sheet_name),
                        field=source_cell.coordinate,
                    )
                    raise failure from error
                if target_cell.value != expected:
                    _fail("inserted-row-formula-mismatch", str(sheet_name), target_cell.coordinate)
                continue
            if target_cell.value != trusted.get(column):
                _fail("inserted-row-value-mismatch", str(sheet_name), target_cell.coordinate)
        for column in range(_FIRST_REQUEST_COLUMN, _LAST_REQUEST_COLUMN + 1):
            cell = target_sheet.cell(insertion_row, column)
            link = _hyperlink_target(cell)
            if column != 23 and cell.hyperlink is not None:
                _fail("inserted-row-unexpected-hyperlink", str(sheet_name), cell.coordinate)
            if column == 23:
                if hyperlink is None:
                    if cell.hyperlink is not None:
                        _fail("inserted-row-hyperlink-mismatch", str(sheet_name), cell.coordinate)
                elif link != hyperlink or getattr(cell.hyperlink, "location", None) is not None:
                    _fail("inserted-row-hyperlink-mismatch", str(sheet_name), cell.coordinate)
    finally:
        control.close()
        candidate.close()


def _is_column_part(value: str) -> bool:
    return bool(value) and all("A" <= character <= "Z" for character in value)


def _cell_part(value: str) -> tuple[str, str] | None:
    cursor = 0
    if cursor < len(value) and value[cursor] == "$":
        cursor += 1
    start = cursor
    while cursor < len(value) and "A" <= value[cursor] <= "Z":
        cursor += 1
    column = value[start:cursor]
    if not _is_column_part(column):
        return None
    if cursor < len(value) and value[cursor] == "$":
        cursor += 1
    row = value[cursor:]
    if not row or not row.isdecimal():
        return None
    return column, row


def _registry_reference_end(formula: str, start: int) -> int:
    cursor = start
    while cursor < len(formula) and (formula[cursor].isalnum() or formula[cursor] in "$:"):
        cursor += 1
    return cursor


def _expanded_registry_formula(formula: str) -> str | None:
    """Return the one permitted Dashboard text change, without formula parsing."""
    pieces: list[str] = []
    cursor = 0
    changed = False
    while True:
        marker = formula.find(_REGISTRY_TOKEN, cursor)
        if marker < 0:
            pieces.append(formula[cursor:])
            break
        end = _registry_reference_end(formula, marker + len(_REGISTRY_TOKEN))
        reference = formula[marker + len(_REGISTRY_TOKEN):end]
        pieces.append(formula[cursor:marker + len(_REGISTRY_TOKEN)])
        parts = reference.split(":")
        if len(parts) != 2:
            return None
        if _is_column_part(parts[0]) and _is_column_part(parts[1]):
            pieces.append(reference)
        else:
            first, last = _cell_part(parts[0]), _cell_part(parts[1])
            if first is None or last is None or first[0] != last[0] or first[1] != "4" or last[1] != "1001":
                return None
            pieces.append(reference[:-4] + "1002")
            changed = True
        cursor = end
    return "".join(pieces) if changed else formula


def _formula_cells(book: object, registry_sheet: str) -> tuple[tuple[str, str, str], ...]:
    result: list[tuple[str, str, str]] = []
    for sheet in book.worksheets:
        if sheet.title == registry_sheet:
            continue
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    result.append((sheet.title, cell.coordinate, cell.value))
    return tuple(result)


def validate_dependent_registry_references(
    control_path: Path,
    candidate_path: Path,
    *,
    insertion_row: int,
    registry_sheet: str = _REGISTRY_SHEET,
) -> None:
    """Allow only the measured Dashboard range endpoint expansion.

    Formula locations/order are part of the contract.  This is comparison-only:
    it neither evaluates nor rewrites a workbook formula.
    """
    if insertion_row not in _KNOWN_INSERTION_ROWS:
        _fail("dependent-formula-insertion-row-invalid", registry_sheet, str(insertion_row))
    control = None
    try:
        control = load_workbook(control_path, read_only=False, data_only=False)
        candidate = load_workbook(candidate_path, read_only=False, data_only=False)
    except Exception as error:
        if control is not None:
            control.close()
        _fail("dependent-formula-read-failed", registry_sheet, type(error).__name__)
    try:
        if registry_sheet not in control.sheetnames or registry_sheet not in candidate.sheetnames:
            _fail("dependent-formula-registry-sheet-missing", registry_sheet, "sheet")
        before, after = _formula_cells(control, registry_sheet), _formula_cells(candidate, registry_sheet)
        before_keys = tuple((sheet, coordinate) for sheet, coordinate, _ in before)
        after_keys = tuple((sheet, coordinate) for sheet, coordinate, _ in after)
        if before_keys != after_keys:
            _fail("dependent-formula-cell-set-mismatch", registry_sheet, "cells")
        for (sheet, coordinate, old), (_, _, new) in zip(before, after):
            try:
                expected = _expanded_registry_formula(old)
            except Exception as error:
                failure = MutationManifestError(
                    "dependent-formula-tokenization-invalid",
                    subject=sheet,
                    field=coordinate,
                )
                raise failure from error
            if expected is None or new != expected:
                _fail("dependent-formula-reference-mismatch", sheet, coordinate)
    finally:
        control.close()
        candidate.close()
