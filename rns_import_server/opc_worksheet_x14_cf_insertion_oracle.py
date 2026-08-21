"""Read-only X14 conditional-formatting middle-row insertion oracle."""
from __future__ import annotations

from dataclasses import dataclass
import hashlib
from pathlib import Path
from typing import Iterable

from openpyxl.formula.translate import Translator

from . import opc_worksheet_x14_cf_owner_topology as x14


__all__ = (
    "OPCWorksheetX14CfInsertionOracleError",
    "validate_x14_cf_middle_insert",
)


@dataclass
class OPCWorksheetX14CfInsertionOracleError(ValueError):
    """A blocking, four-field mismatch from the X14 insertion oracle."""

    code: str
    subject: str
    field: str
    detail: str

    def __post_init__(self) -> None:
        ValueError.__init__(self, self.code, self.subject, self.field, self.detail)

    def as_tuple(self) -> tuple[str, str, str, str]:
        return (self.code, self.subject, self.field, self.detail)


def _fail(code: str, subject: str, field: str, detail: str) -> None:
    raise OPCWorksheetX14CfInsertionOracleError(code, subject, field, detail)


def _semantic_xml(element: object) -> object:
    """Prefix- and attribute-order-independent XML value for an inline DXF."""
    return (
        element.tag,
        tuple(sorted(element.attrib.items())),
        (element.text or "").strip(),
        (element.tail or "").strip(),
        tuple(_semantic_xml(child) for child in element),
    )


def _dxf_fingerprints(path: str) -> dict[str, str]:
    """Get DXF identities only after X2b has already validated their envelope."""
    result: dict[str, str] = {}
    for _, _, owners in x14._accepted(x14._path(path)):
        for _, owner in owners:
            for rule in (child for child in owner if child.tag == x14._RULE):
                rule_id = rule.attrib["id"]
                dxf = next(child for child in rule if child.tag == x14._DXF)
                fingerprint = hashlib.sha256(repr(_semantic_xml(dxf)).encode()).hexdigest()
                if rule_id in result:
                    _fail("ambiguous-x14-cf-rule-id", path, "id", rule_id)
                result[rule_id] = fingerprint
    return result


def _sheet(envelope: x14.WorkbookX14CfSqrefEnvelope, name: str):
    matches = tuple(sheet for sheet in envelope.worksheets if sheet.worksheet.name == name)
    if len(matches) != 1:
        _fail("missing-x14-cf-insertion-sheet", name, "sheet_name", str(len(matches)))
    return matches[0]


def _ranges(items: Iterable[x14.X14CfSqrefRange]) -> tuple[tuple[int, int, int, int], ...]:
    return tuple(sorted((item.min_row, item.min_column, item.max_row, item.max_column) for item in items))


def _mapped_range(item: x14.X14CfSqrefRange, insertion_row: int) -> tuple[int, int, int, int]:
    if item.min_row >= insertion_row:
        return (item.min_row + 1, item.min_column, item.max_row + 1, item.max_column)
    if item.max_row >= insertion_row - 1:
        return (item.min_row, item.min_column, item.max_row + 1, item.max_column)
    return (item.min_row, item.min_column, item.max_row, item.max_column)


def _anchor(item: tuple[int, int, int, int]) -> str:
    row, column, _, _ = item
    letters = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{row}"


def _translated(formula: str, source_anchor: str, candidate_anchor: str, subject: str) -> str:
    try:
        return Translator(f"={formula}", origin=source_anchor).translate_formula(candidate_anchor)[1:]
    except Exception as error:
        _fail("unsupported-x14-cf-formula", subject, "formula", type(error).__name__)
    raise AssertionError("unreachable")


def _flat(sheet: x14.WorksheetX14CfSqrefEnvelope):
    groups = tuple(sheet.containers)
    rules = tuple(rule for group in groups for rule in group.rules)
    by_id = {rule.rule_id: rule for rule in rules}
    if len(by_id) != len(rules):
        _fail("ambiguous-x14-cf-rule-id", sheet.worksheet.name, "id", "duplicate")
    ranges = {rule.rule_id: group.ranges for group in groups for rule in group.rules}
    return rules, by_id, ranges


def validate_x14_cf_middle_insert(
    control: str | Path,
    candidate: str | Path,
    *,
    sheet_name: str,
    insertion_row: int,
    format_source_row: int,
) -> None:
    """Block unless candidate exactly reflects one Excel-style middle-row insert.

    The function only reads package bytes.  It deliberately validates through the
    strict X2b reader before inspecting opaque inline-DXF payloads.
    """
    if insertion_row < 2 or format_source_row != insertion_row - 1:
        _fail("invalid-x14-cf-insertion-row", sheet_name, "format_source_row", str(format_source_row))
    try:
        control_path = str(control)
        candidate_path = str(candidate)
        before = _sheet(x14.read_worksheet_x14_cf_sqref_envelope(control_path), sheet_name)
        after = _sheet(x14.read_worksheet_x14_cf_sqref_envelope(candidate_path), sheet_name)
        before_dxf = _dxf_fingerprints(control_path)
        after_dxf = _dxf_fingerprints(candidate_path)
    except OPCWorksheetX14CfInsertionOracleError:
        raise
    except Exception as error:
        _fail("invalid-x14-cf-insertion-input", sheet_name, "x14", type(error).__name__)

    old_rules, old_by_id, old_ranges = _flat(before)
    new_rules, new_by_id, new_ranges = _flat(after)
    old_ids = tuple(rule.rule_id for rule in old_rules)
    new_ids = tuple(rule.rule_id for rule in new_rules)
    if old_ids != new_ids:
        _fail("x14-cf-rule-order-mismatch", sheet_name, "id", "guid-order")
    for rule_id in old_ids:
        old, new = old_by_id[rule_id], new_by_id[rule_id]
        if (old.type, old.priority, old.stop_if_true) != (new.type, new.priority, new.stop_if_true):
            _fail("x14-cf-rule-mismatch", sheet_name, "rule", rule_id)
        if before_dxf.get(rule_id) != after_dxf.get(rule_id):
            _fail("x14-cf-dxf-mismatch", sheet_name, "dxf", rule_id)
        expected_ranges = tuple(_mapped_range(item, insertion_row) for item in old_ranges[rule_id])
        actual_ranges = _ranges(new_ranges[rule_id])
        if tuple(sorted(expected_ranges)) != actual_ranges:
            _fail("x14-cf-sqref-mismatch", sheet_name, "sqref", rule_id)
        # Formula references move with the first range anchor, even when the
        # range itself expands upward from the copied source row.
        expected_formula = _translated(old.formula, old_ranges[rule_id][0].start_coordinate, _anchor(expected_ranges[0]), sheet_name)
        if new.formula != expected_formula:
            _fail("x14-cf-formula-mismatch", sheet_name, "formula", rule_id)
