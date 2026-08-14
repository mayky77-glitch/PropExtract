"""Bounded, capability-backed publication of manual current-job row edits."""
from __future__ import annotations

import os
import secrets
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

try:
    from rns_import_server.audit import sha256
    from rns_import_server.workbook import DATE_EDIT_FIELDS, EDITABLE_FIELDS, apply_manual_edit, apply_proposal, editable_field_values, iso_date
except ModuleNotFoundError:
    from audit import sha256
    from workbook import DATE_EDIT_FIELDS, EDITABLE_FIELDS, apply_manual_edit, apply_proposal, editable_field_values, iso_date


def _capability_matches(expected: object, candidate: object) -> bool:
    return isinstance(candidate, str) and candidate.isascii() and secrets.compare_digest(str(expected), candidate)


def _normalize_fields(fields: object) -> dict[str, object]:
    if not isinstance(fields, dict) or not fields:
        raise ValueError("Укажите хотя бы одно поле для исправления.")
    normalized: dict[str, object] = {}
    for key, raw_value in fields.items():
        if not isinstance(key, str) or key not in EDITABLE_FIELDS:
            raise ValueError("Недоступное поле для исправления.")
        if raw_value is None or raw_value == "":
            normalized[key] = None
        elif not isinstance(raw_value, str):
            raise ValueError("Значение поля должно быть текстом.")
        elif len(raw_value) > 32_767 or any(
            (ord(char) < 32 and char not in "\t\n\r") or 0xD800 <= ord(char) <= 0xDFFF or ord(char) in {0xFFFE, 0xFFFF}
            for char in raw_value
        ):
            raise ValueError("Значение поля содержит недопустимые символы или слишком длинное.")
        elif key in DATE_EDIT_FIELDS:
            try:
                normalized[key] = iso_date(raw_value)
            except ValueError as error:
                raise ValueError("Введите дату в формате ДД.ММ.ГГГГ или ГГГГ-ММ-ДД.") from error
        else:
            normalized[key] = raw_value.strip()
    return normalized


def publish_manual_edit(manager: Any, job_id: str, edit_id: str, capability: object, fields: object, retry: Callable[[Callable[[], Any]], Any]) -> dict[str, object]:
    """Consume one edit id only after its verified staged workbook is published."""
    normalized = _normalize_fields(fields)
    with manager._lock:
        job = manager._jobs.get(job_id)
        if not job or not _capability_matches(job.get("capability", ""), capability):
            raise ValueError("Действие недоступно. Запустите перенос заново.")
        if job.get("status") != "done":
            raise ValueError("Исправление доступно только после завершения переноса.")
        edits = dict(job.get("edits_internal", {}))
        edit = dict(edits.get(edit_id, {}))
        if not edit or edit.get("status") != "pending":
            raise ValueError("Исправление недоступно или уже выполнено.")
        expected = str(job.get("target_hash", ""))
        if not expected:
            raise RuntimeError("manual_target_stale")
        reservation_order = int(job.get("publication_order", 0)) + 1
        job["publication_order"] = reservation_order
        edit["status"] = "publishing"; edit["expected_target_hash"] = expected
        edit["reservation_order"] = reservation_order
        edit["field_labels"] = tuple(EDITABLE_FIELDS[key] for key in normalized)
        edits[edit_id] = edit; job["edits_internal"] = edits
    staged: Path | None = None
    manager._publish_lock.acquire()
    try:
        target = Path(str(job["xlsx"]))
        row, number = edit.get("row"), str(edit.get("number", ""))
        with manager._lock:
            current = manager._jobs.get(job_id)
            current_edit = dict(current.get("edits_internal", {}).get(edit_id, {})) if current else {}
            expected = str(current.get("target_hash", "")) if current else ""
            proposals = dict(current.get("proposals_internal", {})) if current else {}
        edit_order = int(edit.get("reservation_order", 0))
        changed_labels = set(edit.get("field_labels", ()))
        if current_edit.get("status") != "publishing":
            raise RuntimeError("manual_target_stale")
        if any(
            item.get("status") in {"approving", "approved"}
            and item.get("number") == number
            and item.get("field") in changed_labels
            and int(item.get("reservation_order", 0)) < edit_order
            and str(item.get("expected_target_hash", "")) == str(edit.get("expected_target_hash", ""))
            for item in proposals.values()
        ):
            raise RuntimeError("manual_target_stale")
        if not isinstance(row, int) or not expected or not target.is_file() or target.is_symlink() or sha256(target) != expected:
            raise RuntimeError("manual_target_stale")
        descriptor, name = tempfile.mkstemp(prefix=f".{target.stem}.manual-", suffix=".xlsx", dir=target.parent)
        os.close(descriptor); staged = Path(name); staged.unlink(missing_ok=True)
        apply_manual_edit(target, staged, expected, row, number, normalized)
        staged_hash = sha256(staged)
        refreshed_values = editable_field_values(staged, row, number)
        refreshed_details = _status_details(staged, row)
        backup_dir = target.parent / "Резервные копии PropExtract"; backup_dir.mkdir(parents=True, exist_ok=True)
        backup = backup_dir / f"{target.stem} — до ручного исправления {datetime.now().strftime('%Y-%m-%d_%H-%M-%S_%f')}.xlsx"
        retry(lambda: shutil.copy2(target, backup))
        if sha256(backup) != expected:
            backup.unlink(missing_ok=True); raise RuntimeError("manual_backup_invalid")
        if sha256(target) != expected:
            backup.unlink(missing_ok=True); raise RuntimeError("manual_target_stale")
        retry(lambda: os.replace(staged, target))
        with manager._lock:
            current = manager._jobs[job_id]
            edits = dict(current.get("edits_internal", {})); completed = dict(edits[edit_id])
            completed["status"] = "edited"; edits[edit_id] = completed
            cards = list(current.get("row_cards", []))
            for card in cards:
                if card.get("edit_id") == edit_id:
                    card["editable_values"] = refreshed_values; card["edited"] = True
                    card["object"] = refreshed_values.get("object") or None
                    card["details"] = refreshed_details
                    fresh_id = secrets.token_urlsafe(24)
                    edits[fresh_id] = {"row": row, "number": number, "status": "pending"}
                    card["edit_id"] = fresh_id
            proposals = dict(current.get("proposals_internal", {}))
            public_proposals = list(current.get("proposals", []))
            changed_labels = set(edit.get("field_labels", ()))
            field_keys = {label: key for key, label in EDITABLE_FIELDS.items()}
            for proposal_id, proposal in proposals.items():
                if proposal.get("status") in {"pending", "approving", "approved", "resolved_manual"} and proposal.get("number") == number and proposal.get("field") in changed_labels:
                    field_key = field_keys[str(proposal.get("field"))]
                    proposal["status"] = "resolved_manual"; proposal["manual_value"] = refreshed_values[field_key]; proposals[proposal_id] = proposal
            for public in public_proposals:
                if public.get("number") != number:
                    continue
                public["object"] = refreshed_values.get("object") or None
                field_key = field_keys.get(str(public.get("field")))
                if public.get("field") in changed_labels and field_key:
                    public["status"] = "resolved_manual"; public["action"] = "Исправлено вручную"; public["manual_value"] = refreshed_values[field_key]
            current.update(edits_internal=edits, row_cards=cards, proposals_internal=proposals, proposals=public_proposals, published=True, target_hash=staged_hash, updated_at=datetime.now().isoformat(timespec="seconds"))
        return manager.public(job_id) or {}
    except Exception:
        with manager._lock:
            current = manager._jobs.get(job_id)
            if current:
                edits = dict(current.get("edits_internal", {})); pending = dict(edits.get(edit_id, {}))
                if pending.get("status") == "publishing":
                    pending["status"] = "pending"; edits[edit_id] = pending; current["edits_internal"] = edits
        raise
    finally:
        if staged:
            staged.unlink(missing_ok=True)
        manager._publish_lock.release()


def publish_proposal(manager: Any, job_id: str, proposal_id: str, capability: object, retry: Callable[[Callable[[], Any]], Any]) -> dict[str, object]:
    """Publish one approved proposal, rebasing only non-overlapping internal work."""
    with manager._lock:
        job = manager._jobs.get(job_id)
        if not job or not _capability_matches(job.get("capability", ""), capability):
            raise ValueError("Действие недоступно. Запустите перенос заново.")
        if job.get("status") != "done":
            raise ValueError("Подтверждение доступно только после завершения переноса.")
        proposals = dict(job.get("proposals_internal", {})); proposal = dict(proposals.get(proposal_id, {}))
        if not proposal or proposal.get("status") != "pending":
            raise ValueError("Предложение недоступно или уже обработано.")
        if not job.get("target_hash"):
            raise RuntimeError("proposal_target_stale")
        proposal["status"] = "approving"; proposal["expected_target_hash"] = job["target_hash"]
        proposal["reservation_order"] = int(job.get("publication_order", 0)) + 1
        job["publication_order"] = proposal["reservation_order"]
        proposals[proposal_id] = proposal; job["proposals_internal"] = proposals
    staged: Path | None = None
    manager._publish_lock.acquire()
    try:
        target = Path(str(job["xlsx"])); proposal_label = str(proposal.get("field", "")); proposal_order = int(proposal["reservation_order"])
        with manager._lock:
            current = manager._jobs.get(job_id); current_proposal = dict(current.get("proposals_internal", {}).get(proposal_id, {})) if current else {}
            expected = str(current.get("target_hash", "")) if current else ""; edits = dict(current.get("edits_internal", {})) if current else {}
        if current_proposal.get("status") != "approving" or any(
            item.get("status") in {"publishing", "edited"} and item.get("number") == proposal.get("number")
            and proposal_label in set(item.get("field_labels", ())) and int(item.get("reservation_order", 0)) < proposal_order
            for item in edits.values()
        ):
            raise RuntimeError("proposal_target_stale")
        document = dict(job.get("documents_internal", {}).get(proposal.get("document_id"), {})); pdf = document.get("path")
        try:
            if not isinstance(pdf, Path):
                raise ValueError
            pdf.resolve().relative_to(Path(str(job["pdf_dir"])).resolve())
        except ValueError as error:
            raise RuntimeError("proposal_source_unavailable") from error
        if pdf.suffix.lower() != ".pdf" or not pdf.is_file() or pdf.is_symlink() or document.get("hash") != sha256(pdf):
            raise RuntimeError("proposal_source_unavailable")
        if not expected or not target.is_file() or target.is_symlink() or sha256(target) != expected:
            raise RuntimeError("proposal_target_stale")
        descriptor, name = tempfile.mkstemp(prefix=f".{target.stem}.proposal-", suffix=".xlsx", dir=target.parent)
        os.close(descriptor); staged = Path(name); staged.unlink(missing_ok=True)
        value: object = proposal["value"]
        if proposal_label in {"Дата выдачи", "Срок действия", "Дата последн. измен."}:
            value = iso_date(str(value))
        apply_proposal(target, staged, expected, str(proposal["number"]), proposal_label, value, pdf)
        staged_hash = sha256(staged)
        backup_dir = target.parent / "Резервные копии PropExtract"; backup_dir.mkdir(parents=True, exist_ok=True)
        backup = backup_dir / f"{target.stem} — до подтверждения {datetime.now().strftime('%Y-%m-%d_%H-%M-%S_%f')}.xlsx"
        retry(lambda: shutil.copy2(target, backup))
        if sha256(backup) != expected or sha256(target) != expected:
            backup.unlink(missing_ok=True); raise RuntimeError("proposal_target_stale")
        retry(lambda: os.replace(staged, target)); proposal["status"] = "approved"
        with manager._lock:
            current = manager._jobs[job_id]; proposals = dict(current.get("proposals_internal", {})); proposals[proposal_id] = proposal
            public = list(current.get("proposals", []))
            for item in public:
                if item.get("id") == proposal_id:
                    item["status"] = "approved"
            current.update(proposals_internal=proposals, proposals=public, published=True, target_hash=staged_hash, updated_at=datetime.now().isoformat(timespec="seconds"))
        return {"status": "approved", "backup": backup.name}
    except Exception:
        with manager._lock:
            current = manager._jobs.get(job_id)
            if current:
                proposals = dict(current.get("proposals_internal", {})); pending = dict(proposals.get(proposal_id, {}))
                if pending.get("status") == "approving":
                    pending["status"] = "pending"; proposals[proposal_id] = pending; current["proposals_internal"] = proposals
        raise
    finally:
        if staged:
            staged.unlink(missing_ok=True)
        manager._publish_lock.release()


def _status_details(target: Path, row: int) -> str:
    try:
        from rns_import_server.workbook import SHEET, STATUS_COLUMN
    except ModuleNotFoundError:
        from workbook import SHEET, STATUS_COLUMN
    from openpyxl import load_workbook
    value = load_workbook(target, data_only=False)[SHEET].cell(row, STATUS_COLUMN).value
    return str(value or "").replace("\n", "; ")
