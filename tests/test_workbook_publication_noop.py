from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from rns_import_server.audit import sha256
from rns_import_server.workbook import SHEET, STATUS_COLUMN, STATUS_HEADER, apply


def _review_workbook(path: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet.cell(3, 6).value = "Номер РНС"
    sheet.cell(3, 24).value = "Примечание"
    sheet.cell(3, STATUS_COLUMN).value = STATUS_HEADER
    sheet.cell(4, 6).value = "38-1-1-2026"
    sheet.cell(4, STATUS_COLUMN).value = (
        "Не перенесено «Застройщик»: устаревшая системная причина.\n"
        "Исправлено вручную: «Наименование объекта»."
    )
    book.save(path)


def test_stable_review_is_not_republished_and_preserves_manual_aa_marker(tmp_path: Path):
    workbook = tmp_path / "register.xlsx"
    pdf = tmp_path / "review.pdf"
    _review_workbook(workbook)
    pdf.touch()
    record = {"38-1-1-2026": {"filename": pdf.name, "pdf": pdf}}

    first = apply(record, workbook, workbook, sha256(workbook))
    assert first["published"] is True
    status = load_workbook(workbook)[SHEET].cell(4, STATUS_COLUMN).value
    assert "В документе не найдено ни одного подтверждённого поля для переноса." in status
    assert "Исправлено вручную: «Наименование объекта»." in status

    stable_hash = sha256(workbook)
    stable_mtime = workbook.stat().st_mtime_ns
    second = apply(record, workbook, workbook, stable_hash)

    assert second["published"] is False
    assert sha256(workbook) == stable_hash
    assert workbook.stat().st_mtime_ns == stable_mtime
    assert "Исправлено вручную: «Наименование объекта»." in load_workbook(workbook)[SHEET].cell(4, STATUS_COLUMN).value


def test_low_quality_semantic_noop_after_manual_edit_preserves_aa_without_publication(tmp_path: Path):
    workbook = tmp_path / "register.xlsx"
    pdf = tmp_path / "review.pdf"
    _review_workbook(workbook)
    pdf.touch()
    sheet = load_workbook(workbook)[SHEET]
    sheet["D4"] = "Подтверждённый объект"
    sheet.cell(4, STATUS_COLUMN).value = "Исправлено вручную: «Наименование объекта»."
    sheet.column_dimensions["AA"].width = 58
    sheet.parent.save(workbook)
    before_hash, before_mtime = sha256(workbook), workbook.stat().st_mtime_ns

    result = apply({"38-1-1-2026": {
        "filename": pdf.name,
        "pdf": pdf,
        "object": "Подтверждённый объект",
        "field_quality": {"object": {"status": "review", "reason": "low_confidence"}},
    }}, workbook, workbook, before_hash)

    assert result["published"] is False
    assert result["changes"][0]["physical_mutation"] is False
    assert result["changes"][0]["issues"] == []
    assert sha256(workbook) == before_hash
    assert workbook.stat().st_mtime_ns == before_mtime
    assert load_workbook(workbook)[SHEET].cell(4, STATUS_COLUMN).value == "Исправлено вручную: «Наименование объекта»."


def test_stable_review_rejects_a_stale_source_hash(tmp_path: Path):
    workbook = tmp_path / "register.xlsx"
    pdf = tmp_path / "review.pdf"
    _review_workbook(workbook)
    pdf.touch()
    record = {"38-1-1-2026": {"filename": pdf.name, "pdf": pdf}}

    with pytest.raises(RuntimeError, match="source_xlsx_changed"):
        apply(record, workbook, workbook, "stale")


def test_stable_review_materializes_a_distinct_identical_output(tmp_path: Path):
    source = tmp_path / "register.xlsx"
    output = tmp_path / "published.xlsx"
    pdf = tmp_path / "review.pdf"
    _review_workbook(source)
    pdf.touch()
    record = {"38-1-1-2026": {"filename": pdf.name, "pdf": pdf}}
    apply(record, source, source, sha256(source))

    source_hash = sha256(source)
    source_mtime = source.stat().st_mtime_ns
    result = apply(record, source, output, source_hash)

    assert result["published"] is False
    assert output.exists()
    assert sha256(output) == source_hash
    assert sha256(source) == source_hash
    assert source.stat().st_mtime_ns == source_mtime


def test_generated_merge_warning_is_replaced_without_losing_operator_note(tmp_path: Path):
    workbook = tmp_path / "register.xlsx"
    pdf = tmp_path / "review.pdf"
    _review_workbook(workbook)
    pdf.touch()
    apply({"38-1-1-2026": {"filename": pdf.name, "pdf": pdf}}, workbook, workbook, sha256(workbook))
    sheet = load_workbook(workbook)[SHEET]
    sheet.cell(4, STATUS_COLUMN).value += "\nПодтверждено оператором Ивановым."
    book = sheet.parent
    book.save(workbook)

    issuer = "Связанные изменения содержат разные значения поля «Орган выдачи»; автоматический перенос поля не выполнен."
    builder = "Связанные изменения содержат разные значения поля «Застройщик»; автоматический перенос поля не выполнен."
    base = {"filename": pdf.name, "pdf": pdf}
    first = {**base, "merge_issues": [{"message": issuer}]}
    second = {**base, "merge_issues": [{"message": builder}]}

    apply({"38-1-1-2026": first}, workbook, workbook, sha256(workbook))
    apply({"38-1-1-2026": second}, workbook, workbook, sha256(workbook))
    status = str(load_workbook(workbook)[SHEET].cell(4, STATUS_COLUMN).value)

    assert issuer not in status
    assert builder in status
    assert "Подтверждено оператором Ивановым." in status

    apply({"38-1-1-2026": base}, workbook, workbook, sha256(workbook))
    cleared = str(load_workbook(workbook)[SHEET].cell(4, STATUS_COLUMN).value)
    assert builder not in cleared
    assert "В документе не найдено ни одного подтверждённого поля для переноса." in cleared
    assert "Подтверждено оператором Ивановым." in cleared
