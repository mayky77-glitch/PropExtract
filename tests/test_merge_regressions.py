"""Synthetic regressions for amendment merge and CLI summary contracts."""
from __future__ import annotations

import json
import sys
from itertools import permutations
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from rns_import_server import app
from rns_import_server.audit import sha256
from rns_import_server.workbook import SHEET, apply


NUMBER = "RU-12345678-09-2026"


def _record(filename: str, *, end: str, changed: str, provenance: str = "ocr") -> dict[str, object]:
    return {
        "number": NUMBER,
        "filename": filename,
        "pdf": f"C:/synthetic/{filename}",
        "end": end,
        "changed": changed,
        "field_provenance": {"end": provenance, "changed": provenance},
        "warnings": [],
    }


@pytest.mark.parametrize(
    ("directive_end", "expected_end", "expected_filename"),
    [
        ("31.12.2027", "31.12.2027", "продление.pdf"),
        ("31.12.2026", "31.12.2026", "разрешение.pdf"),
        ("31.12.2025", "31.12.2026", "разрешение.pdf"),
        ("не дата", "31.12.2026", "разрешение.pdf"),
    ],
)
def test_extension_end_replaces_only_with_strictly_later_valid_deadline(
    directive_end: str, expected_end: str, expected_filename: str
):
    permit = _record("разрешение.pdf", end="31.12.2026", changed="01.01.2026")
    directive = _record("продление.pdf", end=directive_end, changed="01.01.2026", provenance="filename")

    merged = app._merge_group(NUMBER, [permit, directive], [])

    assert merged is not None
    assert merged["end"] == expected_end
    assert merged["filename"] == expected_filename


def test_winning_extension_fields_keep_their_provenance_and_primary_source():
    permit = _record("разрешение.pdf", end="31.12.2026", changed="01.01.2026")
    directive = _record("продление.pdf", end="31.12.2027", changed="02.02.2026", provenance="filename")
    directive["field_provenance"] = {"end": "filename", "changed": "ocr"}

    merged = app._merge_group(NUMBER, [permit, directive], [])

    assert merged is not None
    assert merged["pdf"] == "C:/synthetic/продление.pdf"
    assert merged["filename"] == "продление.pdf"
    assert merged["field_sources"] == {
        "end": "C:/synthetic/продление.pdf",
        "changed": "C:/synthetic/продление.pdf",
    }
    assert merged["field_provenance"]["end"] == "filename"
    assert merged["field_provenance"]["changed"] == "ocr"


def test_newer_changed_selects_its_actual_source_when_deadline_does_not_change():
    permit = _record("разрешение.pdf", end="31.12.2026", changed="01.01.2026")
    directive = _record("изменение.pdf", end="31.12.2026", changed="02.02.2026", provenance="filename")

    merged = app._merge_group(NUMBER, [permit, directive], [])

    assert merged is not None
    assert merged["pdf"] == "C:/synthetic/изменение.pdf"
    assert merged["filename"] == "изменение.pdf"
    assert merged["field_sources"]["changed"] == "C:/synthetic/изменение.pdf"
    assert merged["field_provenance"]["changed"] == "filename"


def test_primary_source_is_newest_changed_winner_for_every_amendment_order():
    permit = _record("разрешение.pdf", end="31.12.2026", changed="01.01.2026")
    changed_winner = _record("изменение-новейшее.pdf", end="31.12.2026", changed="03.03.2026")
    end_winner = _record("продление-срок.pdf", end="31.12.2028", changed="02.02.2026", provenance="filename")

    outcomes = set()
    for versions in permutations([permit, changed_winner, end_winner]):
        merged = app._merge_group(NUMBER, list(versions), [])
        assert merged is not None
        outcomes.add((
            merged["end"],
            merged["changed"],
            merged["pdf"],
            merged["filename"],
            merged["field_sources"]["end"],
            merged["field_sources"]["changed"],
            merged["field_provenance"]["end"],
            merged["field_provenance"]["changed"],
        ))

    assert outcomes == {(
        "31.12.2028",
        "03.03.2026",
        "C:/synthetic/изменение-новейшее.pdf",
        "изменение-новейшее.pdf",
        "C:/synthetic/продление-срок.pdf",
        "C:/synthetic/изменение-новейшее.pdf",
        "filename",
        "ocr",
    )}


def test_primary_source_uses_filename_tie_break_for_equal_newest_changes():
    permit = _record("разрешение.pdf", end="31.12.2026", changed="01.01.2026")
    first = _record("изменение-а.pdf", end="31.12.2026", changed="03.03.2026")
    second = _record("изменение-б.pdf", end="31.12.2026", changed="03.03.2026", provenance="filename")

    sources = {
        app._merge_group(NUMBER, list(versions), [])["filename"]
        for versions in permutations([permit, first, second])
    }

    assert sources == {"изменение-б.pdf"}


def test_valid_directive_repairs_invalid_base_dates_and_quarantines_unrepaired_dates():
    permit = _record("разрешение.pdf", end="не дата", changed="тоже не дата")
    directive = _record("продление.pdf", end="31.12.2027", changed="02.02.2026", provenance="filename")

    repaired = app._merge_group(NUMBER, [permit, directive], [])
    quarantined = app._merge_group(NUMBER, [permit], [])

    assert repaired is not None
    assert (repaired["end"], repaired["changed"]) == ("31.12.2027", "02.02.2026")
    assert repaired["field_provenance"]["end"] == "filename"
    assert quarantined is not None
    assert quarantined["end"] is None and quarantined["changed"] is None
    assert {"invalid_date:end", "invalid_date:changed"} <= set(quarantined["warnings"])


def test_directive_only_date_winners_are_deterministic_not_evidence_count_selected():
    older = _record("изменение-старое.pdf", end="31.12.2026", changed="01.01.2026")
    older["object"] = "Дополнительное поле повышает evidence count"
    newer = _record("продление-новое.pdf", end="31.12.2028", changed="02.02.2026", provenance="filename")

    outcomes = set()
    for versions in permutations([older, newer]):
        merged = app._merge_group(NUMBER, list(versions), [])
        assert merged is not None
        outcomes.add((
            merged["existing_only"],
            merged["end"],
            merged["changed"],
            merged["filename"],
            merged["field_sources"]["end"],
        ))

    assert outcomes == {(
        True,
        "31.12.2028",
        "02.02.2026",
        "продление-новое.pdf",
        "C:/synthetic/продление-новое.pdf",
    )}


@pytest.mark.parametrize("with_permit", [False, True])
def test_conflicting_missing_directive_field_is_never_selected_by_order(with_permit: bool):
    first = _record("изменение-а.pdf", end="31.12.2027", changed="02.02.2026")
    second = _record("изменение-б.pdf", end="31.12.2027", changed="02.02.2026")
    first["issuer"], second["issuer"] = "Орган A", "Орган B"
    versions = [first, second]
    if with_permit:
        permit = _record("разрешение.pdf", end="31.12.2026", changed="01.01.2026")
        permit["issuer"] = None
        versions.append(permit)

    results = []
    for order in permutations(versions):
        documents: list[dict] = []
        merged = app._merge_group(NUMBER, list(order), documents)
        assert merged is not None
        assert documents == []
        results.append((merged["issuer"], merged["merge_issues"], merged["warnings"]))

    assert {value for value, _, _ in results} == {None}
    assert all(issues == [{
        "code": "conflicting_directive_field:issuer",
        "field": "issuer",
        "message": "Связанные изменения содержат разные значения поля «Орган выдачи»; автоматический перенос поля не выполнен.",
    }] for _, issues, _ in results)
    assert all("conflicting_directive_field:issuer" in warnings for _, _, warnings in results)


def test_equivalent_directive_values_choose_deterministic_source():
    first = _record("изменение-а.pdf", end="31.12.2027", changed="02.02.2026")
    second = _record("изменение-б.pdf", end="31.12.2027", changed="02.02.2026", provenance="filename")
    first["issuer"], second["issuer"] = "Орган «А»", 'орган "а"'

    sources = set()
    for versions in permutations([first, second]):
        merged = app._merge_group(NUMBER, list(versions), [])
        assert merged is not None
        assert merged["merge_issues"] == []
        sources.add((merged["issuer"], merged["field_sources"]["issuer"]))

    assert sources == {('орган "а"', "C:/synthetic/изменение-б.pdf")}


def _blank_existing_workbook(path: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["W3"] = "Ссылка на документ"
    sheet["F4"] = NUMBER
    book.save(path)


def test_directive_only_newest_dates_reach_blank_workbook_without_invalid_date_abort(tmp_path: Path):
    source, output = tmp_path / "register.xlsx", tmp_path / "output.xlsx"
    _blank_existing_workbook(source)
    older = _record("изменение-старое.pdf", end="31.13.2026", changed="не дата")
    older["object"] = "Дополнительное поле"
    newer = _record("продление-новое.pdf", end="31.12.2028", changed="02.02.2026")
    older["pdf"] = str(tmp_path / older["filename"])
    newer["pdf"] = str(tmp_path / newer["filename"])
    merged = app._merge_group(NUMBER, [older, newer], [])
    assert merged is not None

    result = apply({NUMBER: merged}, source, output, sha256(source))
    saved = load_workbook(output)[SHEET]

    assert result["changes"][0]["outcome"] == "updated"
    assert saved["H4"].value.strftime("%d.%m.%Y") == "31.12.2028"
    assert saved["I4"].value.strftime("%d.%m.%Y") == "02.02.2026"


def test_run_report_preserves_field_provenance_and_merge_issues(monkeypatch, tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    pdf = pdf_dir / "разрешение.pdf"
    pdf.write_bytes(b"pdf")
    xlsx, output, report = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "report.json"
    xlsx.write_bytes(b"xlsx")
    record = _record(pdf.name, end="31.12.2027", changed="02.02.2026", provenance="filename")
    record["pdf"] = str(pdf)
    record["field_sources"] = {"end": str(pdf), "changed": str(pdf)}
    record["merge_issues"] = [{"code": "synthetic", "field": "issuer", "message": "Синтетическая проверка"}]

    monkeypatch.setattr(app, "discover_pdfs", lambda root: [pdf])
    monkeypatch.setattr(app, "sha256", lambda path: "stable-hash")
    monkeypatch.setattr(app, "collect", lambda *args: ({NUMBER: record}, []))
    monkeypatch.setattr(app, "apply", lambda *args: {"changes": [], "conflicts": [], "verification": {}})

    result = app.run(pdf_dir, xlsx, output)
    app.atomic_json(report, result)
    restored = json.loads(report.read_text(encoding="utf-8"))

    selected = restored["selected_records"][NUMBER]
    assert selected["field_provenance"] == {"end": "filename", "changed": "filename"}
    assert selected["field_sources"] == {"end": str(pdf), "changed": str(pdf)}
    assert selected["merge_issues"] == record["merge_issues"]


@pytest.mark.parametrize(
    ("max_pages", "retry_pages"),
    [(0, 10), (4, 4), (100, 10)],
)
def test_collect_retries_unidentified_permit_at_400_dpi_with_bounded_pages(
    monkeypatch, tmp_path: Path, max_pages: int, retry_pages: int
):
    pdf = tmp_path / "Разрешение на строительство.pdf"
    pdf.write_bytes(b"pdf")
    calls: list[tuple[int, int]] = []

    def read(source: Path, dpi: int, pages: int):
        calls.append((dpi, pages))
        return ("нет номера" if dpi == 180 else NUMBER), 14

    monkeypatch.setattr(app, "read_ocr", read)

    records, documents = app.collect(tmp_path, 180, max_pages, pdfs=[pdf])

    assert calls == [(180, max_pages), (400, retry_pages)]
    assert list(records) == [NUMBER]
    assert documents[0]["number"] == NUMBER


@pytest.mark.parametrize(
    ("filename", "dpi", "text"),
    [
        ("ГРО.pdf", 180, "нет номера"),
        ("unrelated.pdf", 180, "нет номера"),
        ("Разрешение.pdf", 400, "нет номера"),
        ("Разрешение.pdf", 180, f"{NUMBER}; RU-87654321-09-2026"),
        ("Разрешение.pdf", 180, NUMBER),
    ],
)
def test_collect_does_not_retry_unrelated_400_ambiguous_or_already_extracted(
    monkeypatch, tmp_path: Path, filename: str, dpi: int, text: str
):
    pdf = tmp_path / filename
    pdf.write_bytes(b"pdf")
    calls: list[tuple[int, int]] = []

    def read(source: Path, requested_dpi: int, pages: int):
        calls.append((requested_dpi, pages))
        return text, 3

    monkeypatch.setattr(app, "read_ocr", read)

    app.collect(tmp_path, dpi, 0, pdfs=[pdf])

    assert calls == [(dpi, 0)]


def test_collect_retry_failure_keeps_bounded_russian_error_and_exact_technical_detail(monkeypatch, tmp_path: Path):
    pdf = tmp_path / "Разрешение.pdf"
    pdf.write_bytes(b"pdf")
    calls = 0

    def read(source: Path, dpi: int, pages: int):
        nonlocal calls
        calls += 1
        if dpi == 400:
            raise RuntimeError("native retry detail")
        return "нет номера", 2

    monkeypatch.setattr(app, "read_ocr", read)

    records, documents = app.collect(tmp_path, 180, 0, pdfs=[pdf])

    assert records == {} and calls == 2 and len(documents) == 1
    assert documents[0]["warnings"] == ["identity_retry_failed"]
    assert documents[0]["error"] == "Повторное распознавание номера РНС при 400 DPI завершилось ошибкой."
    assert documents[0]["technical_error"] == "native retry detail"


def test_collect_rejects_negative_page_limit_before_first_ocr(monkeypatch, tmp_path: Path):
    pdf = tmp_path / "Разрешение.pdf"
    pdf.write_bytes(b"pdf")
    monkeypatch.setattr(app, "read_ocr", lambda *args: pytest.fail("OCR must not start"))

    with pytest.raises(ValueError, match="не может быть отрицательным"):
        app.collect(tmp_path, 180, -4, pdfs=[pdf])


def test_process_summary_is_ascii_json_that_restores_cyrillic_paths(monkeypatch, capsys):
    pdf_dir = Path("C:/синтетика/PDF")
    xlsx = Path("C:/синтетика/Реестр.xlsx")
    output = Path("C:/синтетика/Результат.xlsx")
    captured_report: list[tuple[Path, dict]] = []
    result = {"logical_records": [NUMBER], "conflicts": [], "marker": "кириллица"}

    monkeypatch.setattr(app, "run", lambda *args: result)
    monkeypatch.setattr(app, "atomic_json", lambda path, value: captured_report.append((path, value)))
    monkeypatch.setattr(
        sys,
        "argv",
        ["app.py", "process", "--pdf-dir", str(pdf_dir), "--xlsx", str(xlsx), "--output", str(output)],
    )

    app.main()

    summary = capsys.readouterr().out
    assert summary.encode("ascii")
    assert json.loads(summary) == {
        "output": str(output),
        "report": str(output.with_suffix(".json")),
        "records": 1,
        "conflicts": 0,
    }
    assert captured_report == [(output.with_suffix(".json"), result)]
