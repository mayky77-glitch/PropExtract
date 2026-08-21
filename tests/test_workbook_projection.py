from __future__ import annotations

import hashlib
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from rns_import_server.workbook_projection import (
    TemplateCellEvidence,
    GroupOwnershipEvidence,
    WorkbookProjectionAdapter,
    WorkbookProjectionAuthority,
    WorkbookProjectionCode,
    project_workbook,
)


def _hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _make_book(path: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "Реестр РНС"
    sheet["D1"] = "Стройка"
    sheet["C2"] = "123-1234567.0001"
    sheet["D2"] = "Объект"
    fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for row in (3, 4):
        for column in tuple(range(1, 25)) + (27,):
            sheet.cell(row, column).fill = fill
    book.save(path)
    book.close()


def _authority(path: Path, **changes: object) -> WorkbookProjectionAuthority:
    values: dict[str, object] = {
        "target_path": str(path),
        "target_identity": "target-identity",
        "workbook_contract_id": "publication-finalized-k3b2b-v1",
        "sheet_identity": "Реестр РНС",
        "template_version": "template-v1",
        "registry_generation": 7,
        "template_cells": (TemplateCellEvidence(1, 4, "Стройка"),),
        "group_ownership": tuple(GroupOwnershipEvidence(row, row == 3) for row in range(1, 5)),
    }
    values.update(changes)
    return WorkbookProjectionAuthority.verified(**values)  # type: ignore[arg-type]


def test_one_immutable_read_feeds_both_projections_and_preserves_source(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    copy = tmp_path / "copy.xlsx"
    _make_book(source)
    copy.write_bytes(source.read_bytes())
    before = (_hash(source), source.stat().st_mtime_ns, source.read_bytes(), _hash(copy), copy.stat().st_mtime_ns, copy.read_bytes())

    result = project_workbook(_authority(copy))

    assert result.code is WorkbookProjectionCode.OK and result.snapshot is not None
    snapshot = result.snapshot
    assert snapshot.pre_hash == before[3] == snapshot.sheet.workbook_hash == snapshot.provisioning.workbook_hash
    assert snapshot.authority_identity == snapshot.sheet.workbook_identity == snapshot.provisioning.workbook_identity
    assert snapshot.registry_generation == snapshot.sheet.registry_generation == snapshot.provisioning.registry_generation == 7
    assert snapshot.sheet.rows[1].c == "123-1234567.0001"
    assert snapshot.provisioning.rows[1].values[4] == "Объект"
    assert snapshot.sheet.rows[2].is_preformatted and snapshot.provisioning.rows[2].is_business_row
    assert (_hash(source), source.stat().st_mtime_ns, source.read_bytes(), _hash(copy), copy.stat().st_mtime_ns, copy.read_bytes()) == before


def test_rejects_template_sheet_and_path_authority_mismatches(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _make_book(path)
    assert project_workbook(_authority(path, sheet_identity="Other")).code is WorkbookProjectionCode.SHEET_MISMATCH
    assert project_workbook(_authority(path, template_cells=(TemplateCellEvidence(1, 4, "Other"),))).code is WorkbookProjectionCode.TEMPLATE_MISMATCH
    noncanonical = f"{path.parent}/./{path.name}"
    assert project_workbook(_authority(path, target_path=noncanonical)).code is WorkbookProjectionCode.UNSAFE_TARGET


def test_rejects_symlink_and_malformed_authority_without_partial_projection(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _make_book(path)
    link = tmp_path / "link.xlsx"
    try:
        link.symlink_to(path)
    except OSError:
        return
    assert project_workbook(_authority(link)).code is WorkbookProjectionCode.UNSAFE_TARGET
    assert project_workbook(_authority(path, registry_generation=True)).code is WorkbookProjectionCode.INVALID_AUTHORITY
    assert project_workbook(_authority(path, template_cells=())).snapshot is None


def test_adapter_requires_injected_authority_port(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _make_book(path)

    class Port:
        def read_authority(self) -> WorkbookProjectionAuthority:
            return _authority(path)

    assert WorkbookProjectionAdapter(Port()).read().code is WorkbookProjectionCode.OK


def test_rejects_path_replacement_after_hash_even_with_equal_size_and_mtime(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "book.xlsx"
    replacement = tmp_path / "replacement.xlsx"
    _make_book(path)
    _make_book(replacement)
    book = Workbook()
    sheet = book.active
    sheet.title = "Реестр РНС"
    sheet["D1"] = "Стройка"
    sheet["C2"] = "123-1234567.9999"
    sheet["D2"] = "Объект"
    fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for row in (3, 4):
        for column in tuple(range(1, 25)) + (27,):
            sheet.cell(row, column).fill = fill
    book.save(replacement)
    book.close()
    padding = path.stat().st_size - replacement.stat().st_size
    assert padding >= 0
    if padding:
        replacement.write_bytes(replacement.read_bytes() + b"\0" * padding)
    assert replacement.stat().st_size == path.stat().st_size
    original_stat = path.stat()
    import rns_import_server.workbook_projection as projection_module
    original_load = projection_module.load_workbook

    def replace_then_load(*args, **kwargs):
        path.write_bytes(replacement.read_bytes())
        os.utime(path, ns=(original_stat.st_atime_ns, original_stat.st_mtime_ns))
        return original_load(*args, **kwargs)

    monkeypatch.setattr(projection_module, "load_workbook", replace_then_load)
    result = project_workbook(_authority(path))
    assert result.code is WorkbookProjectionCode.SOURCE_UNSTABLE and result.snapshot is None


def test_explicit_last_group_ownership_allows_blank_plan_and_missing_malformed_conflict_fail(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _make_book(path)
    result = project_workbook(_authority(path))
    assert result.snapshot is not None
    from rns_import_server.workbook_groups import WorkbookGroupCode, resolve_workbook_group
    resolution = resolve_workbook_group(
        result.snapshot.sheet, construction_id="construction", official_name="Стройка", code_prefix="123-1234567",
        rns="RU-12345678-09-2026", official_names=("Стройка",),
    )
    assert resolution.code is WorkbookGroupCode.BLANK_ROW_PLANNED
    missing = _authority(path, group_ownership=(GroupOwnershipEvidence(1, False),))
    malformed = _authority(path, group_ownership=(GroupOwnershipEvidence(1, 1),))
    conflict = _authority(path, group_ownership=(GroupOwnershipEvidence(1, False), GroupOwnershipEvidence(1, True)))
    assert project_workbook(missing).code is WorkbookProjectionCode.GROUP_OWNERSHIP_MISSING
    assert project_workbook(malformed).code is WorkbookProjectionCode.INVALID_AUTHORITY
    assert project_workbook(conflict).code is WorkbookProjectionCode.GROUP_OWNERSHIP_CONFLICT
