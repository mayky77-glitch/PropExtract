"""HTTP acceptance coverage for capability-backed current-job row correction."""
from __future__ import annotations

import json
import socket
import threading
import time
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from rns_import_server.audit import sha256
from rns_import_server.server import create_server
from rns_import_server.workbook import SHEET, apply


NUMBER = "38-1-1-2026"


def _port() -> int:
    with socket.socket() as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def _request(port: int, method: str, path: str, payload: dict[str, object] | None = None) -> tuple[int, dict[str, object]]:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8") if payload is not None else None
    request = urllib.request.Request(f"http://127.0.0.1:{port}{path}", data=body, method=method, headers={"Content-Type": "application/json"} if body else {})
    try:
        with urllib.request.urlopen(request, timeout=5) as response:
            return response.status, json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        return error.code, json.loads(error.read().decode("utf-8"))


def _wait(port: int, job_id: str) -> dict[str, object]:
    for _ in range(100):
        status, job = _request(port, "GET", f"/api/jobs/{job_id}")
        assert status == 200
        if job.get("status") in {"done", "error"}:
            return job
        time.sleep(0.02)
    raise AssertionError("job did not finish")


def _target(path: Path, old_pdf: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["F4"] = NUMBER
    sheet["D4"] = "Старый объект"
    sheet["H4"] = datetime(2025, 12, 31)
    sheet["W3"] = "Ссылка на документ"
    sheet["W4"] = old_pdf.name
    sheet["W4"].hyperlink = old_pdf.as_uri()
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    sheet["Z4"] = '=IF(F4<>"",ROW(),"")'
    book.save(path)


def _start_review_job(tmp_path: Path) -> tuple[int, object, threading.Thread, Path, dict[str, object]]:
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    old_pdf, pdf = pdf_dir / "old.pdf", pdf_dir / "new.pdf"
    old_pdf.write_bytes(b"old")
    pdf.write_bytes(b"new")
    target = tmp_path / "register.xlsx"
    _target(target, old_pdf)

    def runner(pdf_root: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict[str, object]:
        record = {
            "number": NUMBER, "filename": pdf.name, "pdf": str(pdf), "stage": None,
            "object": "Новый объект", "issue": None, "end": None, "changed": None,
            "issuer": None, "builder": None, "region": None, "district": None, "developer": None,
        }
        result = apply({NUMBER: record}, xlsx, output, sha256(xlsx))
        result.update(
            input_hashes={"xlsx": sha256(xlsx), "pdfs": {pdf.name: sha256(pdf), old_pdf.name: sha256(old_pdf)}},
            documents=[{"file": str(pdf)}], logical_records=[NUMBER],
            selected_records={NUMBER: {**record, "field_sources": {"object": pdf.name}}},
        )
        return result

    port = _port()
    server = create_server("127.0.0.1", port, runner)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    status, started = _request(port, "POST", "/api/jobs", {"pdf_dir": str(pdf_dir), "xlsx": str(target)})
    assert status == 202
    return port, server, thread, target, _wait(port, str(started["id"]))


def _stop(server: object, thread: threading.Thread) -> None:
    server.shutdown()  # type: ignore[attr-defined]
    server.server_close()  # type: ignore[attr-defined]
    thread.join(timeout=5)


def test_http_manual_edit_current_job_replay_proposal_resolution_and_invariants(tmp_path: Path):
    port, server, thread, target, job = _start_review_job(tmp_path)
    try:
        assert job["status"] == "done"
        assert "xlsx" not in job and "target_hash" not in job and "edits_internal" not in job
        card = job["row_cards"][0]  # type: ignore[index]
        proposal = job["proposals"][0]  # type: ignore[index]
        edit_id, capability = str(card["edit_id"]), str(job["capability"])
        before = sha256(target)

        status, denied = _request(port, "POST", f"/api/jobs/{job['id']}/edits/{edit_id}", {"capability": "wrong", "fields": {"object": "Нельзя"}})
        assert status == 400 and "Проверьте" in str(denied.get("error"))
        assert sha256(target) == before

        status, updated = _request(port, "POST", f"/api/jobs/{job['id']}/edits/{edit_id}", {"capability": capability, "fields": {"object": "Исправленный объект", "end": "2027-12-31"}})
        assert status == 200
        assert updated["row_cards"][0]["editable_values"]["object"] == "Исправленный объект"  # type: ignore[index]
        assert updated["proposals"][0]["id"] == proposal["id"] and updated["proposals"][0]["status"] == "resolved_manual"  # type: ignore[index]
        saved = load_workbook(target, data_only=False)[SHEET]
        assert saved["D4"].value == "Исправленный объект"
        assert saved["H4"].value == datetime(2027, 12, 31)
        assert saved["F4"].value == NUMBER
        assert saved["W4"].hyperlink.target.endswith("/old.pdf")
        assert saved["Y4"].value == '=IF(A4<>"",ROW(),"")' and saved["Z4"].value == '=IF(F4<>"",ROW(),"")'

        status, replay = _request(port, "POST", f"/api/jobs/{job['id']}/edits/{edit_id}", {"capability": capability, "fields": {"object": "Повтор"}})
        assert status == 400 and "Проверьте" in str(replay.get("error"))
        assert load_workbook(target)[SHEET]["D4"].value == "Исправленный объект"
    finally:
        _stop(server, thread)


def test_http_manual_edit_rejects_invalid_date_and_stale_target_in_russian(tmp_path: Path):
    port, server, thread, target, job = _start_review_job(tmp_path)
    try:
        card = job["row_cards"][0]  # type: ignore[index]
        endpoint = f"/api/jobs/{job['id']}/edits/{card['edit_id']}"
        payload = {"capability": job["capability"], "fields": {"end": "31.41.2027"}}
        status, invalid = _request(port, "POST", endpoint, payload)
        assert status == 400
        assert "Проверьте" in str(invalid.get("error")) and "manual_" not in json.dumps(invalid, ensure_ascii=False)

        book = load_workbook(target)
        book[SHEET]["J4"] = "Внешнее изменение"
        book.save(target)
        status, stale = _request(port, "POST", endpoint, {"capability": job["capability"], "fields": {"object": "Поздняя запись"}})
        assert status == 500
        assert stale["error"] == "Данные изменились после запуска. Реестр не изменён."
        assert load_workbook(target)[SHEET]["J4"].value == "Внешнее изменение"
    finally:
        _stop(server, thread)
