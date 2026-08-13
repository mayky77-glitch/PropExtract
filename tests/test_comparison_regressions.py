from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from rns_import_server.audit import sha256
from rns_import_server.workbook import SHEET, apply, transfer_issue


NUMBER = "RU-12345678-09-2026"


def _record(pdf: Path, **values: object) -> dict[str, object]:
    return {
        "number": NUMBER,
        "stage": None,
        "object": None,
        "issue": None,
        "end": None,
        "changed": None,
        "issuer": None,
        "builder": None,
        "region": None,
        "district": None,
        "developer": None,
        "filename": pdf.name,
        "pdf": str(pdf),
        **values,
    }


def _existing_workbook(path: Path, field: str, value: object) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["F4"] = NUMBER
    sheet["W3"] = "Ссылка на документ"
    columns = {"Наименование объекта": "D4", "Срок действия": "H4", "Орган выдачи": "J4", "Застройщик": "K4"}
    sheet[columns[field]] = value
    book.save(path)


def _apply_existing(tmp_path: Path, field: str, existing: object, **record_values: object) -> dict[str, object]:
    source, output, pdf = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "permit.pdf"
    pdf.write_bytes(b"pdf")
    _existing_workbook(source, field, existing)
    return apply({NUMBER: _record(pdf, **record_values)}, source, output, sha256(source))


def test_quote_only_builder_difference_creates_neither_issue_nor_proposal(tmp_path: Path):
    result = _apply_existing(tmp_path, "Застройщик", 'ПАО "Газпром"', builder="ПАО «Газпром»")

    assert transfer_issue("Застройщик", 'ПАО "Газпром"', "ПАО «Газпром»") is None
    assert result["changes"][0]["outcome"] == "already_present"
    assert result["conflicts"] == []


def test_object_comparison_ignores_only_generic_bundle_preamble(tmp_path: Path):
    object_tail = (
        "Этап 13.5. Объекты внешнего электроснабжения УКПГ-45. "
        "Блочно-комплектная ПС 110 кВ УКПГ-45 с двумя силовыми "
        "трансформаторами 110/10 кВ единичной мощностью 16 МВА"
    )
    pdf_object = f"Обустройство Ковыктинского газоконденсатного месторождения. {object_tail}"
    result = _apply_existing(
        tmp_path,
        "Наименование объекта",
        object_tail,
        object=pdf_object,
    )

    assert transfer_issue("Наименование объекта", object_tail, pdf_object) is None
    assert result["changes"][0]["outcome"] == "already_present"
    assert result["conflicts"] == []


def test_numeric_engineering_unit_presentation_differences_are_comparison_only(tmp_path: Path):
    existing = "ПС Восточная 110 кВ, мощность 16 МВА"
    proposed = "ПС Восточная 110 KB, мощность 16МВА"
    source, output, pdf = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "permit.pdf"
    pdf.write_bytes(b"pdf")
    _existing_workbook(source, "Наименование объекта", existing)

    result = apply({NUMBER: _record(pdf, object=proposed)}, source, output, sha256(source))

    assert transfer_issue("Наименование объекта", existing, proposed) is None
    assert result["changes"][0]["outcome"] == "already_present"
    assert load_workbook(output)[SHEET]["D4"].value == existing


@pytest.mark.parametrize(
    "proposed",
    (
        "ПС Восточная 110 KB, мощность 17МВА",
        "ПС Восточная 110 kB, мощность 16 МВА. Этап 2",
        "ПС Восточная 110 кВ, газотурбинная мощность 16 МВА",
    ),
)
def test_engineering_unit_normalization_keeps_semantic_object_conflicts(tmp_path: Path, proposed: str):
    existing = "ПС Восточная 110 кВ, газозапорная мощность 16 МВА. Этап 1"
    result = _apply_existing(tmp_path, "Наименование объекта", existing, object=proposed)

    assert result["changes"][0]["outcome"] == "review"
    assert result["conflicts"][0]["existing"] == existing
    assert result["conflicts"][0]["pdf"] == proposed


def test_object_comparison_preserves_significant_pre_anchor_substation(tmp_path: Path):
    result = _apply_existing(
        tmp_path,
        "Наименование объекта",
        "ПС Северная. Этап 1. Линия 110 кВ",
        object="ПС Южная. Этап 1. Линия 110 кВ",
    )

    assert result["changes"][0]["outcome"] == "review"
    assert result["conflicts"] == [
        {
            "number": NUMBER,
            "cell": "D4",
            "field": "Наименование объекта",
            "existing": "ПС Северная. Этап 1. Линия 110 кВ",
            "pdf": "ПС Южная. Этап 1. Линия 110 кВ",
            "action": "Перенести изменения",
        }
    ]


@pytest.mark.parametrize(
    ("field", "existing", "record_values"),
    [
        ("Срок действия", datetime(2026, 11, 28), {"end": "28.11.2025"}),
        ("Орган выдачи", "Администрация района", {"issuer": "Служба надзора"}),
        (
            "Наименование объекта",
            "Этап 1. Линия 110 кВ ПС Восточная",
            {"object": "Этап 1. Линия 110 кВ"},
        ),
        (
            "Наименование объекта",
            "Этап 1. Линия 110 кВ ПС Восточная",
            {"object": "Этап 2. Линия 110 кВ ПС Восточная"},
        ),
        (
            "Наименование объекта",
            "Этап 1. Линия 110 кВ ПС Восточная",
            {"object": "Линия 110 кВ ПС Восточная"},
        ),
        (
            "Наименование объекта",
            "Обустройство Ковыктинского газоконденсатного месторождения. "
            "Этап 1. Линия 110 кВ",
            {
                "object": "Обустройство Ковыктинского тазоконденсатного месторождения. "
                "Этап 1. Линия 110 кВ"
            },
        ),
    ],
)
def test_substantive_date_issuer_and_object_differences_remain_proposals(
    tmp_path: Path,
    field: str,
    existing: object,
    record_values: dict[str, object],
):
    result = _apply_existing(tmp_path, field, existing, **record_values)

    assert result["changes"][0]["outcome"] == "review"
    assert result["conflicts"] == [
        {
            "number": NUMBER,
            "cell": {"Срок действия": "H4", "Орган выдачи": "J4", "Наименование объекта": "D4"}[field],
            "field": field,
            "existing": "28.11.2026" if field == "Срок действия" else str(existing),
            "pdf": next(iter(record_values.values())),
            "action": "Перенести изменения",
        }
    ]
