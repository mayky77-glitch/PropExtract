"""Synthetic contracts for capability-backed, OOXML-safe row correction."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
import json
import threading
import time

import pytest
from openpyxl import Workbook, load_workbook

from rns_import_server.audit import sha256
from rns_import_server import action_history
from rns_import_server import row_edit
from rns_import_server import server
from rns_import_server.server import JobManager
from rns_import_server.workbook import SHEET, apply


NUMBER = "38-1-1-2026"


def _wait(manager: JobManager, job_id: str) -> dict[str, object]:
    import time
    for _ in range(100):
        job = manager.get(job_id)
        if job and job["status"] in {"done", "error"}:
            return job
        time.sleep(0.02)
    raise AssertionError("job did not finish")


def _record(pdf: Path, **updates: object) -> dict[str, object]:
    return {
        "number": NUMBER, "filename": pdf.name, "pdf": str(pdf),
        "stage": None, "object": None, "issue": None, "end": None,
        "changed": None, "issuer": None, "builder": None, "region": None,
        "district": None, "developer": None, **updates,
    }


def _target(path: Path, old_pdf: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["F4"] = NUMBER
    sheet["D4"] = "Старый объект"
    sheet["H4"] = datetime(2025, 12, 31)
    sheet["W3"] = "Ссылка на документ"
    sheet["W4"] = old_pdf.name
    sheet["W4"].hyperlink = old_pdf.as_uri()
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    sheet["Z4"] = '=IF(F4<>"",ROW(),"")'
    book.save(path)


def _review_job(tmp_path: Path, *, quality: dict[str, object] | None = None) -> tuple[JobManager, dict[str, object], Path, Path]:
    pdf_dir = tmp_path / "pdf"; pdf_dir.mkdir()
    old_pdf, pdf = pdf_dir / "old.pdf", pdf_dir / "new.pdf"
    old_pdf.write_bytes(b"old"); pdf.write_bytes(b"new")
    target = tmp_path / "register.xlsx"; _target(target, old_pdf)

    def runner(pdf_root: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict[str, object]:
        record = _record(pdf, object="Новый объект")
        if quality is not None:
            record["field_quality"] = quality
        result = apply({NUMBER: record}, xlsx, output, sha256(xlsx))
        selected = {**record, "field_sources": {"object": pdf.name}}
        result.update(input_hashes={"xlsx": sha256(xlsx), "pdfs": {pdf.name: sha256(pdf), old_pdf.name: sha256(old_pdf)}}, documents=[{"file": str(pdf)}], logical_records=[NUMBER], selected_records={NUMBER: selected})
        return result

    manager = JobManager(runner, error_log=tmp_path / "error.log")
    job = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))
    assert job["status"] == "done"
    return manager, job, target, pdf


def test_review_with_no_written_cells_keeps_original_w_exact(tmp_path: Path):
    old_pdf, new_pdf = tmp_path / "old.pdf", tmp_path / "new.pdf"
    old_pdf.write_bytes(b"old"); new_pdf.write_bytes(b"new")
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _target(source, old_pdf)

    result = apply({NUMBER: _record(new_pdf, merge_issues=[{"message": "Требуется сверка."}])}, source, output, sha256(source))
    saved = load_workbook(output)[SHEET]

    assert result["changes"][0]["outcome"] == "review"
    assert result["changes"][0]["written"] == []
    assert saved["W4"].value == old_pdf.name
    assert saved["W4"].hyperlink.target == old_pdf.as_uri()


def test_review_with_safe_written_cell_updates_w_and_keeps_review_status(tmp_path: Path):
    old_pdf, new_pdf = tmp_path / "old.pdf", tmp_path / "new.pdf"
    old_pdf.write_bytes(b"old"); new_pdf.write_bytes(b"new")
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _target(source, old_pdf)
    book = load_workbook(source); book[SHEET]["H4"] = None; book.save(source)

    result = apply({NUMBER: _record(new_pdf, end="31.12.2027", merge_issues=[{"message": "Требуется сверка."}])}, source, output, sha256(source))
    saved = load_workbook(output)[SHEET]

    assert result["changes"][0]["outcome"] == "review"
    assert saved["H4"].value == datetime(2027, 12, 31)
    assert saved["W4"].hyperlink.target == new_pdf.as_uri()
    assert saved["AA4"].value == "Требуется сверка."


def test_manual_edit_is_current_job_capability_bound_atomic_and_preserves_invariants(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    public = manager.public(str(job["id"]))
    assert public is not None
    card = public["row_cards"][0]
    edit_id = card["edit_id"]
    before = sha256(target)

    with pytest.raises(ValueError, match="Действие недоступно"):
        manager.edit(str(job["id"]), edit_id, "fake", {"object": "Исправленный объект"})
    assert sha256(target) == before

    result = manager.edit(str(job["id"]), edit_id, job["capability"], {"object": "Исправленный объект", "end": "2027-12-31"})
    saved = load_workbook(target)[SHEET]
    backups = list((target.parent / "Резервные копии PropExtract").glob("*.xlsx"))

    assert result["row_cards"][0]["edited"] is True
    assert result["row_cards"][0]["object"] == "Исправленный объект"
    assert saved["D4"].value == "Исправленный объект"
    assert saved["H4"].value == datetime(2027, 12, 31)
    assert "Исправлено вручную" in saved["AA4"].value
    assert saved["W4"].hyperlink.target == (tmp_path / "pdf" / "old.pdf").as_uri()
    assert saved["Y4"].value == '=IF(A4<>"",ROW(),"")' and saved["Z4"].value == '=IF(F4<>"",ROW(),"")'
    manual_backups = [item for item in backups if "ручного исправления" in item.name]
    assert len(manual_backups) == 1 and sha256(manual_backups[0]) == before
    fresh_edit_id = result["row_cards"][0]["edit_id"]
    with pytest.raises(ValueError, match="недоступно"):
        manager.edit(str(job["id"]), edit_id, job["capability"], {"object": "Повтор"})
    assert fresh_edit_id != edit_id


def test_manual_resolution_projects_the_saved_value_not_stale_pdf_proposal(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    public = manager.public(str(job["id"]))
    edit_id, proposal_id = public["row_cards"][0]["edit_id"], public["proposals"][0]["id"]  # type: ignore[index]
    result = manager.edit(str(job["id"]), edit_id, job["capability"], {"object": "Третье ручное значение"})
    proposal = next(item for item in result["proposals"] if item.get("id") == proposal_id)
    javascript = (Path(__file__).parents[1] / "rns_import_server/static/app.js").read_text(encoding="utf-8")
    assert proposal["status"] == "resolved_manual" and proposal["manual_value"] == "Третье ручное значение"
    assert proposal["object"] == "Третье ручное значение"
    assert result["summary"]["review_rows"] == 0
    report = target.with_name(f"{target.stem} — отчет PropExtract.json")
    assert __import__("json").loads(report.read_text(encoding="utf-8"))["final_state"]["summary"]["review_rows"] == 0
    assert 'resolved ? "Исправлено" : "В документе"' in javascript
    assert "const displayedValue = resolved ? item.manual_value : item.proposed;" in javascript


def test_manual_text_that_looks_like_formula_is_serialized_as_literal(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    edit_id = manager.public(str(job["id"]))["row_cards"][0]["edit_id"]  # type: ignore[index]
    manager.edit(str(job["id"]), edit_id, job["capability"], {"object": "=HYPERLINK(\"bad\")"})
    saved = load_workbook(target, data_only=False)[SHEET]
    assert saved["D4"].value == '=HYPERLINK("bad")'
    assert saved["D4"].data_type == "s"
    assert saved["Y4"].value == '=IF(A4<>"",ROW(),"")'


def test_manual_edit_stale_target_and_invalid_date_fail_closed(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    card = manager.public(str(job["id"]))["row_cards"][0]  # type: ignore[index]
    with pytest.raises(ValueError, match="Введите дату"):
        manager.edit(str(job["id"]), card["edit_id"], job["capability"], {"end": "31.41.2027"})
    assert manager.public(str(job["id"]))["row_cards"][0]["edit_id"] == card["edit_id"]  # type: ignore[index]

    book = load_workbook(target); book[SHEET]["J4"] = "Внешнее изменение"; book.save(target)
    with pytest.raises(RuntimeError, match="manual_target_stale"):
        manager.edit(str(job["id"]), card["edit_id"], job["capability"], {"object": "Исправление"})
    assert load_workbook(target)[SHEET]["J4"].value == "Внешнее изменение"


def test_manual_edit_rejects_excel_unsafe_text_without_publish(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    card = manager.public(str(job["id"]))["row_cards"][0]  # type: ignore[index]
    before = sha256(target)
    with pytest.raises(ValueError, match="недопустимые"):
        manager.edit(str(job["id"]), card["edit_id"], job["capability"], {"object": "плохой\x00текст"})
    assert sha256(target) == before


def test_low_quality_field_has_review_card_but_no_one_click_proposal(tmp_path: Path):
    manager, job, _, _ = _review_job(tmp_path, quality={"object": {"status": "review", "reason": "low_confidence"}})
    public = manager.public(str(job["id"]))
    assert public is not None
    assert public["proposals"][0]["quality"] == "review"
    assert "id" not in public["proposals"][0]
    updated = manager.edit(
        str(job["id"]), public["row_cards"][0]["edit_id"], job["capability"], {"object": "Проверенный объект"}
    )
    assert updated["row_cards"][0]["object"] == "Проверенный объект"
    assert updated["proposals"][0]["status"] == "resolved_manual"
    assert updated["proposals"][0]["manual_value"] == "Проверенный объект"


def test_manual_edit_can_supersede_an_already_approved_field(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    public = manager.public(str(job["id"]))
    proposal_id = public["proposals"][0]["id"]  # type: ignore[index]
    edit_id = public["row_cards"][0]["edit_id"]  # type: ignore[index]

    manager.approve(str(job["id"]), proposal_id, job["capability"])
    updated = manager.edit(str(job["id"]), edit_id, job["capability"], {"object": "Финальное ручное значение"})

    assert load_workbook(target)[SHEET]["D4"].value == "Финальное ручное значение"
    proposal = next(item for item in updated["proposals"] if item.get("id") == proposal_id)
    assert proposal["status"] == "resolved_manual"
    assert proposal["manual_value"] == "Финальное ручное значение"
    report = target.with_name(f"{target.stem} — отчет PropExtract.json")
    actions = __import__("json").loads(report.read_text(encoding="utf-8"))["final_state"]["actions"]
    assert actions == [
        {"type": "proposal_approved", "row": 4, "field": "Наименование объекта", "status": "approved"},
        {"type": "manual_edit", "row": 4, "field": "Наименование объекта", "status": "edited"},
    ]


def test_post_manual_low_quality_semantic_noop_has_no_new_backup_or_changed_rows(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path, quality={"object": {"status": "review", "reason": "low_confidence"}})
    public = manager.public(str(job["id"]))
    manager.edit(str(job["id"]), public["row_cards"][0]["edit_id"], job["capability"], {"object": "Новый объект"})
    before_hash, before_mtime = sha256(target), target.stat().st_mtime_ns
    manual_aa = str(load_workbook(target)[SHEET]["AA4"].value)
    backups = target.parent / "Резервные копии PropExtract"
    before_backups = sorted(item.name for item in backups.glob("*.xlsx"))

    repeated = _wait(manager, str(manager.start(str(target.parent / "pdf"), str(target))["id"]))

    assert repeated["published"] is False
    assert repeated["backup"] is None
    assert repeated["summary"]["changed_rows"] == 0
    assert repeated["summary"]["review_rows"] == 0
    assert sha256(target) == before_hash
    assert target.stat().st_mtime_ns == before_mtime
    assert str(load_workbook(target)[SHEET]["AA4"].value) == manual_aa and "Исправлено вручную" in manual_aa
    assert sorted(item.name for item in backups.glob("*.xlsx")) == before_backups


def test_action_history_persists_after_noop_rerun(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    report = target.with_name(f"{target.stem} — отчет PropExtract.json")
    initial = json.loads(report.read_text(encoding="utf-8"))
    assert initial["final_state"]["actions"] == []

    public = manager.public(str(job["id"]))
    proposal_id = public["proposals"][0]["id"]  # type: ignore[index]
    edit_id = public["row_cards"][0]["edit_id"]  # type: ignore[index]

    manager.approve(str(job["id"]), str(proposal_id), job["capability"])
    manager.edit(str(job["id"]), edit_id, job["capability"], {"object": "Новый объект"})
    with_actions = json.loads(report.read_text(encoding="utf-8"))
    actions = with_actions["final_state"]["actions"]
    assert actions == [
        {"type": "proposal_approved", "row": 4, "field": "Наименование объекта", "status": "approved"},
        {"type": "manual_edit", "row": 4, "field": "Наименование объекта", "status": "edited"},
    ]

    noop = _wait(manager, str(manager.start(str(target.parent / "pdf"), str(target))["id"]))
    with manager._lock:
        current = manager._jobs.get(str(noop["id"])) or {}
    assert current.get("action_events_internal") == actions
    replayed = json.loads(report.read_text(encoding="utf-8"))
    assert replayed["final_state"]["actions"] == actions
    assert noop["published"] is False
    assert noop["backup"] is None


def test_action_history_rejects_stale_or_tampered_report_with_public_warning(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    report = target.with_name(f"{target.stem} — отчет PropExtract.json")
    payload = json.loads(report.read_text(encoding="utf-8"))
    payload["final_state"]["actions"] = [
        {"type": "manual_edit", "row": 4, "field": "Секретное поле", "status": "edited"}
    ]
    report.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    rejected = _wait(manager, str(manager.start(str(target.parent / "pdf"), str(target))["id"]))
    assert rejected["action_events_internal"] == []
    assert "повреждён или имеет неподдерживаемый формат" in str(rejected["warning"])
    saved = json.loads(report.read_text(encoding="utf-8"))
    assert saved["final_state"]["actions"] == []
    assert "повреждён или имеет неподдерживаемый формат" in saved["final_state"]["warning"]

    payload = saved
    payload["final_state"]["workbook_sha256"] = "0" * 64
    report.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    stale = _wait(manager, str(manager.start(str(target.parent / "pdf"), str(target))["id"]))
    assert stale["action_events_internal"] == []
    assert "Excel изменён после создания отчёта" in str(stale["warning"])


def test_action_history_does_not_follow_report_symlink(tmp_path: Path):
    manager, _, target, _ = _review_job(tmp_path)
    report = target.with_name(f"{target.stem} — отчет PropExtract.json")
    external = tmp_path / "external.json"
    external.write_text(report.read_text(encoding="utf-8"), encoding="utf-8")
    report.unlink()
    report.symlink_to(external)
    before = external.read_bytes()

    rejected = _wait(manager, str(manager.start(str(target.parent / "pdf"), str(target))["id"]))

    assert rejected["action_events_internal"] == []
    assert "не удалось безопасно прочитать отчёт" in str(rejected["warning"])
    assert external.read_bytes() == before
    assert report.is_file() and not report.is_symlink()


def test_action_history_rejects_swap_to_symlink_before_open(monkeypatch, tmp_path: Path):
    manager, _, target, _ = _review_job(tmp_path)
    report = target.with_name(f"{target.stem} — отчет PropExtract.json")
    external = tmp_path / "external.json"
    external.write_text(report.read_text(encoding="utf-8"), encoding="utf-8")
    external_before = external.read_bytes()
    original_open = action_history.open_descriptor
    swapped = False

    def swap_before_open(path):
        nonlocal swapped
        if Path(path) == report and not swapped:
            swapped = True
            report.unlink()
            report.symlink_to(external)
        return original_open(path)

    monkeypatch.setattr(action_history, "open_descriptor", swap_before_open)
    rejected = _wait(manager, str(manager.start(str(target.parent / "pdf"), str(target))["id"]))

    assert swapped is True
    assert rejected["action_events_internal"] == []
    assert "не удалось безопасно прочитать отчёт" in str(rejected["warning"])
    assert external.read_bytes() == external_before


def test_windows_report_open_rejects_same_inode_reparse_point(tmp_path: Path):
    report = tmp_path / "report.json"
    report.write_text("{}", encoding="utf-8")
    renamed = tmp_path / "renamed.json"
    report.rename(renamed)
    original_inode = renamed.stat().st_ino
    report.symlink_to(renamed)
    assert renamed.stat().st_ino == original_inode

    calls: dict[str, object] = {}

    class Function:
        def __init__(self, callback):
            self.callback = callback
            self.argtypes = None
            self.restype = None

        def __call__(self, *args):
            return self.callback(*args)

    class Kernel32:
        def __init__(self):
            self.CreateFileW = Function(self.create_file)
            self.GetFileInformationByHandle = Function(self.get_information)
            self.CloseHandle = Function(self.close_handle)

        @staticmethod
        def create_file(*args):
            calls["flags"] = args[5]
            return 123

        @staticmethod
        def get_information(handle, pointer):
            calls["inspected"] = handle
            pointer._obj.dwFileAttributes = 0x00000400
            return 1

        @staticmethod
        def close_handle(handle):
            calls["closed"] = handle
            return 1

    def must_not_convert(handle: int, flags: int) -> int:
        raise AssertionError("reparse-point handle must never become a readable descriptor")

    with pytest.raises(OSError, match="reparse point"):
        action_history.open_windows_descriptor(
            report,
            kernel32=Kernel32(),
            open_osfhandle=must_not_convert,
        )

    assert int(calls["flags"]) & 0x00200000
    assert calls["inspected"] == 123
    assert calls["closed"] == 123


def test_action_history_bounded_reader_accepts_short_regular_file_reads(monkeypatch, tmp_path: Path):
    _, _, target, _ = _review_job(tmp_path)
    original_read = server.os.read

    def short_read(descriptor: int, size: int) -> bytes:
        return original_read(descriptor, min(size, 7))

    monkeypatch.setattr(server.os, "read", short_read)
    actions, warning = action_history.load(target)

    assert actions == []
    assert warning is None


def test_action_history_rejects_oversize_report(monkeypatch, tmp_path: Path):
    manager, _, target, _ = _review_job(tmp_path)
    report = target.with_name(f"{target.stem} — отчет PropExtract.json")
    monkeypatch.setattr(action_history, "MAX_REPORT_SIZE", 128)
    report.write_bytes(b" " * 129)

    rejected = _wait(manager, str(manager.start(str(target.parent / "pdf"), str(target))["id"]))

    assert rejected["action_events_internal"] == []
    assert "повреждён или имеет неподдерживаемый формат" in str(rejected["warning"])


def test_start_waits_for_action_final_report_before_loading_action_history(monkeypatch, tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    public = manager.public(str(job["id"]))
    edit_id = public["row_cards"][0]["edit_id"]  # type: ignore[index]
    entered, release = threading.Event(), threading.Event()
    action_errors: list[Exception] = []
    start_result: list[dict[str, object]] = []
    start_errors: list[Exception] = []
    original_write = server.write_final_action_report

    def pause_report_write(*args, **kwargs):
        entered.set()
        assert release.wait(timeout=5)
        return original_write(*args, **kwargs)

    monkeypatch.setattr(server, "write_final_action_report", pause_report_write)
    action_thread = threading.Thread(
        target=lambda: _capture(action_errors, lambda: manager.edit(str(job["id"]), edit_id, job["capability"], {"object": "Новый объект"}))
    )
    action_thread.start()
    assert entered.wait(timeout=5)  # XLSX is already atomically replaced.

    def start_again() -> None:
        try:
            start_result.append(manager.start(str(target.parent / "pdf"), str(target)))
        except Exception as error:
            start_errors.append(error)

    start_thread = threading.Thread(target=start_again)
    start_thread.start()
    time.sleep(0.1)
    assert start_thread.is_alive(), "start read stale report while action publication was incomplete"
    release.set()
    action_thread.join(timeout=10); start_thread.join(timeout=10)

    assert not action_errors and not start_errors and len(start_result) == 1
    repeated = _wait(manager, str(start_result[0]["id"]))
    actions = [
        {"type": "manual_edit", "row": 4, "field": "Наименование объекта", "status": "edited"}
    ]
    assert repeated["action_events_internal"] == actions
    assert repeated["published"] is False
    assert repeated["backup"] is None
    assert repeated["warning"] is None
    payload = json.loads(target.with_name(f"{target.stem} — отчет PropExtract.json").read_text(encoding="utf-8"))
    assert payload["final_state"]["actions"] == actions


def test_manual_action_is_rejected_while_import_is_running(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    edit_id = manager.public(str(job["id"]))["row_cards"][0]["edit_id"]  # type: ignore[index]
    original_runner = manager.runner
    entered, release = threading.Event(), threading.Event()

    def slow_runner(*args, **kwargs):
        entered.set()
        assert release.wait(timeout=5)
        return original_runner(*args, **kwargs)

    manager.runner = slow_runner
    active = manager.start(str(target.parent / "pdf"), str(target))
    assert entered.wait(timeout=5)
    before = sha256(target)
    try:
        with pytest.raises(server.BusyError, match="пока выполняется импорт"):
            manager.edit(str(job["id"]), edit_id, job["capability"], {"object": "Не записывать"})
        assert sha256(target) == before
    finally:
        release.set()
    assert _wait(manager, str(active["id"]))["status"] == "done"


def test_manual_edit_refreshes_allowlisted_final_report_after_verified_publication(tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    public = manager.public(str(job["id"]))
    report = target.with_name(f"{target.stem} — отчет PropExtract.json")
    initial = __import__("json").loads(report.read_text(encoding="utf-8"))
    assert isinstance(initial["input_hashes"]["xlsx"], str)
    assert initial["verification"]["x14_preserved"] is True
    assert initial["final_state"]["actions"] == []
    manager.edit(str(job["id"]), public["row_cards"][0]["edit_id"], job["capability"], {"object": "Проверено вручную"})

    payload = __import__("json").loads(report.read_text(encoding="utf-8"))
    raw = __import__("json").dumps(payload, ensure_ascii=False)
    state = manager.get(str(job["id"]))

    final = payload["final_state"]
    assert final["schema"] == "propextract.final-action.v1"
    assert final["published"] is True
    assert final["workbook_sha256"] == sha256(target)
    assert final["actions"] == [{"type": "manual_edit", "row": 4, "field": "Наименование объекта", "status": "edited"}]
    assert payload["input_hashes"] == initial["input_hashes"] and payload["verification"] == initial["verification"]
    assert "capability" not in raw and str(target) not in raw and "Проверено вручную" not in raw
    assert state["report"] == str(report)


def test_report_failure_after_manual_edit_is_a_safe_warning_and_keeps_workbook(monkeypatch, tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    public = manager.public(str(job["id"]))
    monkeypatch.setattr(server, "write_final_action_report", lambda *_: (_ for _ in ()).throw(OSError("private/path")))

    updated = manager.edit(str(job["id"]), public["row_cards"][0]["edit_id"], job["capability"], {"object": "Сохранено"})

    assert load_workbook(target)[SHEET]["D4"].value == "Сохранено"
    assert updated["published"] is True
    assert updated["warning"].startswith("Excel обновлён, но отчёт не записан. Причина:")
    assert "локальный путь" in updated["warning"] or "private/path" in updated["warning"]
    assert "private/path" not in __import__("json").dumps(updated, ensure_ascii=False)


@pytest.mark.parametrize("tamper", ["replace", "delete", "symlink"])
def test_action_report_rebuilds_from_server_safe_base_not_tampered_disk(tamper: str, tmp_path: Path):
    manager, job, target, _ = _review_job(tmp_path)
    report = target.with_name(f"{target.stem} — отчет PropExtract.json")
    external = tmp_path / "external.json"
    external.write_text('{"private":"external-secret"}', encoding="utf-8")
    if tamper == "replace":
        report.write_text('{"private":"disk-secret"}', encoding="utf-8")
    elif tamper == "delete":
        report.unlink()
    else:
        report.unlink()
        report.symlink_to(external)

    public = manager.public(str(job["id"]))
    assert "report_base_internal" not in public
    manager.edit(str(job["id"]), public["row_cards"][0]["edit_id"], job["capability"], {"object": "Сохранено"})

    payload = __import__("json").loads(report.read_text(encoding="utf-8"))
    raw = __import__("json").dumps(payload, ensure_ascii=False)
    assert report.is_symlink() is False
    assert payload["final_state"]["workbook_sha256"] == sha256(target)
    assert "disk-secret" not in raw and "external-secret" not in raw
    assert external.read_text(encoding="utf-8") == '{"private":"external-secret"}'


@pytest.mark.parametrize("first", ["approve", "edit"])
def test_concurrent_approval_and_manual_edit_use_the_reserved_target_generation(monkeypatch, tmp_path: Path, first: str):
    """Both actions reserve old SHA; only the first publisher may replace it."""
    manager, job, target, _ = _review_job(tmp_path)
    public = manager.public(str(job["id"]))
    proposal_id = public["proposals"][0]["id"]  # type: ignore[index]
    edit_id = public["row_cards"][0]["edit_id"]  # type: ignore[index]
    entered, release = threading.Event(), threading.Event()
    errors: list[Exception] = []
    original = row_edit.apply_proposal if first == "approve" else row_edit.apply_manual_edit

    def pause_first(*args, **kwargs):
        entered.set()
        assert release.wait(timeout=5)
        return original(*args, **kwargs)

    if first == "approve":
        monkeypatch.setattr(row_edit, "apply_proposal", pause_first)
        first_call = lambda: manager.approve(str(job["id"]), proposal_id, job["capability"])
        second_call = lambda: manager.edit(str(job["id"]), edit_id, job["capability"], {"object": "Ручное"})
    else:
        monkeypatch.setattr(row_edit, "apply_manual_edit", pause_first)
        first_call = lambda: manager.edit(str(job["id"]), edit_id, job["capability"], {"object": "Ручное"})
        second_call = lambda: manager.approve(str(job["id"]), proposal_id, job["capability"])

    first_thread = threading.Thread(target=lambda: _capture(errors, first_call))
    second_thread = threading.Thread(target=lambda: _capture(errors, second_call))
    first_thread.start(); assert entered.wait(timeout=5)
    second_thread.start()
    for _ in range(100):
        state = manager.get(str(job["id"])) or {}
        proposal_state = state.get("proposals_internal", {}).get(proposal_id, {}).get("status")
        edit_state = state.get("edits_internal", {}).get(edit_id, {}).get("status")
        if proposal_state == "approving" and edit_state == "publishing":
            break
        time.sleep(0.01)
    else:
        raise AssertionError("both actions did not reserve their target generation")
    release.set(); first_thread.join(timeout=10); second_thread.join(timeout=10)

    assert len(errors) == 1 and "target_stale" in str(errors[0])
    saved = load_workbook(target)[SHEET]
    assert saved["D4"].value == ("Новый объект" if first == "approve" else "Ручное")
    state = manager.get(str(job["id"])) or {}
    assert state["proposals_internal"][proposal_id]["status"] == ("approved" if first == "approve" else "resolved_manual")
    assert state["edits_internal"][edit_id]["status"] == ("edited" if first == "edit" else "pending")


def _capture(errors: list[Exception], operation) -> None:
    try:
        operation()
    except Exception as error:
        errors.append(error)


def test_field_quality_blocks_empty_and_new_row_cells_but_keeps_actionable_values(tmp_path: Path):
    old_pdf, pdf = tmp_path / "old.pdf", tmp_path / "new.pdf"
    old_pdf.write_bytes(b"old"); pdf.write_bytes(b"new")
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _target(source, old_pdf)
    book = load_workbook(source); book[SHEET]["D4"] = None; book.save(source)
    record = _record(pdf, object="Подтверждённый объект", developer="Сомнительный разработчик", field_quality={"object": {"status": "actionable"}, "developer": {"status": "review"}})
    result = apply({NUMBER: record}, source, output, sha256(source))
    saved = load_workbook(output)[SHEET]
    assert result["changes"][0]["outcome"] == "review"
    assert saved["D4"].value == "Подтверждённый объект" and saved["N4"].value is None
    assert "«Разработчик ПД»" in saved["AA4"].value

    new_number = "38-2-2-2026"
    new_record = _record(pdf, number=new_number, object="Подтверждённый объект", developer="Сомнительный разработчик", field_quality={"object": {"status": "actionable"}, "developer": {"status": "review"}})
    new_output = tmp_path / "new-output.xlsx"
    apply({new_number: new_record}, source, new_output, sha256(source))
    new_sheet = load_workbook(new_output)[SHEET]
    assert new_sheet["F5"].value == new_number and new_sheet["D5"].value == "Подтверждённый объект"
    assert new_sheet["N5"].value is None and "требует ручной сверки" in new_sheet["AA5"].value
