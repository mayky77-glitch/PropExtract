from pathlib import Path

import pytest
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Font, PatternFill

import rns_import_server.workbook_mutation_manifest as mutation_manifest
from rns_import_server.workbook_mutation_manifest import (
    MutationManifestError,
    manifest_for,
    validate_blank_fill,
    validate_dependent_registry_references,
    validate_inserted_row,
    validate_insertion,
)


def _save(path: Path, inserted: bool = False) -> None:
    book = Workbook(); sheet = book.active; sheet.title = "Реестр РНС"
    sheet["A4"] = "1"; sheet["Y4"] = "=A4"; sheet["W4"] = "old"; sheet["W4"].hyperlink = "https://example.test/old"
    if inserted:
        sheet["A5"] = ""; sheet["A6"] = "2"; sheet["Y6"] = "=A6"; sheet["W6"] = "old"; sheet["W6"].hyperlink = "https://example.test/old"
    else:
        sheet["A5"] = "2"; sheet["Y5"] = "=A5"; sheet["W5"] = "old"; sheet["W5"].hyperlink = "https://example.test/old"
    book.save(path); book.close()


def test_manifest_proves_exact_one_insert(tmp_path: Path) -> None:
    control, candidate = tmp_path / "control.xlsx", tmp_path / "candidate.xlsx"; _save(control); _save(candidate, True)
    validate_insertion(manifest_for(control, "Реестр РНС"), manifest_for(candidate, "Реестр РНС", insertion_row=5), 5)


def test_manifest_rejects_unmapped_change(tmp_path: Path) -> None:
    control, candidate = tmp_path / "control.xlsx", tmp_path / "candidate.xlsx"; _save(control); _save(candidate, True)
    from openpyxl import load_workbook
    book = load_workbook(candidate); book.active["Y6"] = "=A1"; book.save(candidate); book.close()
    with pytest.raises(RuntimeError, match="mutation_manifest_changed"):
        validate_insertion(manifest_for(control, "Реестр РНС"), manifest_for(candidate, "Реестр РНС", insertion_row=5), 5)


def _blank_fill_books(path: Path) -> tuple[Path, Path, dict[int, object], str]:
    control, candidate = path / "control.xlsx", path / "candidate.xlsx"
    fields, link = {6: "RU-00000000-00-2026", 23: "Документ"}, "https://example.test/document"
    book = Workbook(); sheet = book.active; sheet.title = "Реестр РНС"
    sheet["A4"] = "outside"; sheet["A4"].hyperlink = "https://example.test/outside"; sheet["F5"] = "old"; sheet["W5"] = "old document"; sheet["W5"].hyperlink = "https://example.test/old"
    sheet["Y5"] = "=A5"; sheet["Z5"] = "=Y5+1"; book.save(control); book.close()
    book = load_workbook(control); sheet = book.active
    for column, value in fields.items(): sheet.cell(5, column).value = value
    sheet["W5"].hyperlink = link; book.save(candidate); book.close()
    return control, candidate, fields, link


def test_blank_fill_manifest_accepts_only_exact_trusted_row_changes(tmp_path: Path) -> None:
    control, candidate, fields, link = _blank_fill_books(tmp_path)
    before, after = manifest_for(control, "Реестр РНС"), manifest_for(candidate, "Реестр РНС")
    validate_blank_fill(before, after, target_row=5, fields=fields, hyperlink=link)
    assert (before.max_row, after.max_row, before.formulas["Y5"], after.formulas["Y5"]) == (5, 5, "=A5", "=A5")
    assert after.hyperlinks["W5"] == {"target": link, "location": None}


@pytest.mark.parametrize(
    ("change", "code"),
    [
        (lambda sheet: setattr(sheet["A4"], "value", "changed"), "blank-fill-outside-value-mismatch"),
        (lambda sheet: setattr(sheet["G5"], "value", "untrusted"), "blank-fill-value-mismatch"),
        (lambda sheet: setattr(sheet["Y5"], "value", "=A1"), "blank-fill-formula-mismatch"),
        (lambda sheet: setattr(sheet["W5"], "hyperlink", "https://example.test/untrusted"), "blank-fill-hyperlink-mismatch"),
        (lambda sheet: setattr(sheet["W5"].hyperlink, "location", "Sheet1!A1"), "blank-fill-hyperlink-mismatch"),
        (lambda sheet: setattr(sheet["A4"].hyperlink, "location", "Sheet1!A1"), "blank-fill-outside-hyperlink-mismatch"),
        (lambda sheet: sheet.insert_rows(1), "blank-fill-row-count-mismatch"),
    ],
)
def test_blank_fill_manifest_rejects_outside_target_formula_link_or_count_changes(tmp_path: Path, change, code: str) -> None:
    control, candidate, fields, link = _blank_fill_books(tmp_path)
    book = load_workbook(candidate); change(book.active); book.save(candidate); book.close()
    with pytest.raises(MutationManifestError) as captured:
        validate_blank_fill(manifest_for(control, "Реестр РНС"), manifest_for(candidate, "Реестр РНС"), target_row=5, fields=fields, hyperlink=link)
    assert captured.value.code == code


def test_blank_fill_requires_trusted_w_display_for_requested_hyperlink(tmp_path: Path) -> None:
    control, candidate, fields, link = _blank_fill_books(tmp_path)
    with pytest.raises(MutationManifestError) as captured:
        validate_blank_fill(manifest_for(control, "Реестр РНС"), manifest_for(candidate, "Реестр РНС"), target_row=5, fields={6: fields[6]}, hyperlink=link)
    assert captured.value.code == "blank-fill-hyperlink-display-required"


def _gate_books(path: Path, insertion_row: int) -> tuple[Path, Path, dict[int, object], str]:
    control, candidate = path / "control.xlsx", path / "candidate.xlsx"
    fields, link = {6: "RU-00000000-00-2026", 23: "Документ"}, "https://example.test/document"
    book = Workbook(); registry = book.active; registry.title = "Реестр РНС"; dashboard = book.create_sheet("Дашборд")
    source_row = insertion_row - 1
    registry.row_dimensions[source_row].height = 27
    for column in range(1, 44):
        cell = registry.cell(source_row, column)
        cell._style = copy(registry.cell(1, 1)._style)
        cell.font = Font(name="Arial", bold=column == 1)
        cell.fill = PatternFill("solid", fgColor="AABBCC")
    registry.cell(source_row, 1).value = "predecessor"
    registry.cell(source_row, 23).value = "old document"
    registry.cell(source_row, 23).hyperlink = "https://example.test/old"
    registry.cell(source_row, 25).value = f"=A{source_row}"
    registry.cell(source_row, 26).value = f"=Y{source_row}+1"
    dashboard["B2"] = "=SUM('Реестр РНС'!F4:F1001)"
    dashboard["C2"] = "=INDEX('Реестр РНС'!C:C,1)"
    dashboard["D2"] = "=SUM('Реестр РНС'!F4:F1001,'Реестр РНС'!R4:R1001)"
    book.save(control); book.close()
    book = load_workbook(control); registry, dashboard = book["Реестр РНС"], book["Дашборд"]
    registry.insert_rows(insertion_row)
    for column in range(1, 44):
        source, target = registry.cell(source_row, column), registry.cell(insertion_row, column)
        target._style = copy(source._style)
    registry.row_dimensions[insertion_row].height = registry.row_dimensions[source_row].height
    registry.cell(insertion_row, 25).value = Translator(registry.cell(source_row, 25).value, origin=f"Y{source_row}").translate_formula(f"Y{insertion_row}")
    registry.cell(insertion_row, 26).value = Translator(registry.cell(source_row, 26).value, origin=f"Z{source_row}").translate_formula(f"Z{insertion_row}")
    for column, value in fields.items(): registry.cell(insertion_row, column).value = value
    registry.cell(insertion_row, 23).hyperlink = link
    dashboard["B2"] = "=SUM('Реестр РНС'!F4:F1002)"
    dashboard["D2"] = "=SUM('Реестр РНС'!F4:F1002,'Реестр РНС'!R4:R1002)"
    book.save(candidate); book.close()
    return control, candidate, fields, link


@pytest.mark.parametrize("insertion_row", [6, 10, 104])
def test_inserted_row_and_dashboard_formula_gate_accepts_exact_native_semantics(tmp_path: Path, insertion_row: int) -> None:
    control, candidate, fields, link = _gate_books(tmp_path, insertion_row)
    validate_inserted_row(control, candidate, sheet_name="Реестр РНС", insertion_row=insertion_row, fields=fields, hyperlink=link)
    validate_dependent_registry_references(control, candidate, insertion_row=insertion_row)


@pytest.mark.parametrize(
    ("change", "code"),
    [
        (lambda book, row: setattr(book["Реестр РНС"].cell(row, 6), "value", "untrusted"), "inserted-row-value-mismatch"),
        (lambda book, row: setattr(book["Реестр РНС"].cell(row, 25), "value", "=A1"), "inserted-row-formula-mismatch"),
        (lambda book, row: setattr(book["Реестр РНС"].row_dimensions[row], "height", 13), "inserted-row-height-mismatch"),
        (lambda book, row: setattr(book["Реестр РНС"].cell(row, 23), "hyperlink", "https://example.test/untrusted"), "inserted-row-hyperlink-mismatch"),
    ],
)
def test_inserted_row_gate_rejects_untrusted_persisted_semantics(tmp_path: Path, change, code: str) -> None:
    control, candidate, fields, link = _gate_books(tmp_path, 6)
    book = load_workbook(candidate); change(book, 6); book.save(candidate); book.close()
    with pytest.raises(MutationManifestError) as captured:
        validate_inserted_row(control, candidate, sheet_name="Реестр РНС", insertion_row=6, fields=fields, hyperlink=link)
    assert captured.value.code == code


def test_inserted_row_gate_resolves_same_style_id_across_workbooks(tmp_path: Path) -> None:
    control, candidate, fields, link = _gate_books(tmp_path, 6)
    book = load_workbook(control); control_style_id = book["Реестр РНС"]["A5"].style_id; book.close()
    book = load_workbook(candidate); cell = book["Реестр РНС"]["A6"]
    assert cell.style_id == control_style_id
    book._fonts[cell._style.fontId] = Font(name="Same ID, different resolved font")
    book.save(candidate); book.close()
    book = load_workbook(candidate); assert book["Реестр РНС"]["A6"].style_id == control_style_id; book.close()
    with pytest.raises(MutationManifestError) as captured:
        validate_inserted_row(control, candidate, sheet_name="Реестр РНС", insertion_row=6, fields=fields, hyperlink=link)
    assert captured.value.code == "inserted-row-style-mismatch"


def test_dependent_formula_gate_rejects_changed_registry_token(tmp_path: Path) -> None:
    control, candidate, _fields, _link = _gate_books(tmp_path, 6)
    book = load_workbook(candidate); book["Дашборд"]["C2"] = "=INDEX('Реестр РНС'!D:D,1)"; book.save(candidate); book.close()
    with pytest.raises(MutationManifestError) as captured:
        validate_dependent_registry_references(control, candidate, insertion_row=6)
    assert captured.value.code == "dependent-formula-reference-mismatch"


def test_dependent_formula_gate_rejects_unchanged_multi_reference_1001(tmp_path: Path) -> None:
    control, candidate, _fields, _link = _gate_books(tmp_path, 6)
    book = load_workbook(candidate)
    book["Дашборд"]["D2"] = "=SUM('Реестр РНС'!F4:F1001,'Реестр РНС'!R4:R1001)"
    book.save(candidate); book.close()
    with pytest.raises(MutationManifestError) as captured:
        validate_dependent_registry_references(control, candidate, insertion_row=6)
    assert captured.value.code == "dependent-formula-reference-mismatch"


def test_dependent_formula_tokenization_failure_is_typed(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    control, candidate, _fields, _link = _gate_books(tmp_path, 6)
    monkeypatch.setattr(mutation_manifest, "_expanded_registry_formula", lambda _formula: (_ for _ in ()).throw(ValueError("invalid token")))
    with pytest.raises(MutationManifestError) as captured:
        validate_dependent_registry_references(control, candidate, insertion_row=6)
    assert captured.value.code == "dependent-formula-tokenization-invalid"
    assert isinstance(captured.value.__cause__, ValueError)
