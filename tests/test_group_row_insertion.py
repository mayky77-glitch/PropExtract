from contextlib import nullcontext
from dataclasses import replace
from pathlib import Path
import threading
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink

import rns_import_server.group_row_insertion as insertion
import rns_import_server.workbook_mutation_manifest as mutation_manifest
from rns_import_server.audit import sha256
from rns_import_server.group_row_insertion import GroupRowInsertionError, GroupRowRequest, PublicationContext, publish_group_row, recover_group_row
from rns_import_server.opc_worksheet_x14_cf_insertion_oracle import OPCWorksheetX14CfInsertionOracleError
from rns_import_server.opc_workbook_filter_database_insertion_oracle import OPCWorkbookFilterDatabaseInsertionOracleError
from rns_import_server.opc_worksheet_structure_insertion_oracle import OPCWorksheetStructureInsertionOracleError
from rns_import_server.workbook_groups import MutationPlan
from rns_import_server.workbook_mutation_manifest import MutationManifestError
from rns_import_server.registry_storage import RegistryConflictError


class Journal:
    def __init__(self): self.calls = []; self.phase = "planned"; self.operation = None
    def get(self, operation_id):
        if self.operation is not None and self.operation["operation_id"] == operation_id:
            return SimpleNamespace(values=self.operation)
        return None
    def create(self, **kwargs):
        self.operation = {**kwargs, "phase": "planned"}; self.calls.append(("create", kwargs)); return object()
    def reserve(self, *, nonce_factory, **kwargs):
        existing = self.get(kwargs["operation_id"])
        if existing is not None:
            return existing, False
        owner_id, pair_nonce = nonce_factory()
        self.create(owner_id=owner_id, pair_nonce=pair_nonce, **kwargs)
        return self.get(kwargs["operation_id"]), True
    def transition(self, operation_id, *, expected_phase, next_phase, **kwargs):
        assert self.phase == expected_phase; self.phase = next_phase
        if self.operation is not None:
            self.operation["phase"] = next_phase
            self.operation.update(kwargs.get("hashes", {}))
        self.calls.append((next_phase, kwargs)); return object()
    def record_post_hash(self, operation_id, *, expected_phase, post_hash):
        assert self.phase == expected_phase
        if self.operation is not None: self.operation["post_hash"] = post_hash
        self.calls.append(("post", post_hash)); return object()
    def finalize_flag(self, operation_id, flag): self.calls.append((flag, {})); return object()


class RacingJournal(Journal):
    """Minimal CAS double for the loser branch of one concurrent create."""

    def __init__(self):
        super().__init__(); self._create_lock = threading.Lock(); self._create_barrier = threading.Barrier(2); self.create_attempts = 0

    def reserve(self, *, nonce_factory, **kwargs):
        self._create_barrier.wait(timeout=3)
        with self._create_lock:
            self.create_attempts += 1
            if self.operation is not None:
                return self.get(kwargs["operation_id"]), False
            owner_id, pair_nonce = nonce_factory()
            super().create(owner_id=owner_id, pair_nonce=pair_nonce, **kwargs)
            return self.get(kwargs["operation_id"]), True


def _book(path: Path) -> None:
    book = Workbook(); sheet = book.active; sheet.title = "Реестр РНС"; sheet.cell(5, 25).value = "=A5"; book.save(path); book.close()


def _context(plan, journal):
    def native(request, _script):
        book = load_workbook(request.candidate); sheet = book[request.sheet]
        for column, value in request.fields.items(): sheet.cell(request.insertion_row, column).value = value
        if request.hyperlink is not None:
            cell = sheet.cell(request.insertion_row, 23)
            cell._hyperlink = Hyperlink(ref=cell.coordinate, target=request.hyperlink)
        book.save(request.candidate); book.close()
        return {"lease": {"excel_adapter": "mock", "excel_pid": 1, "excel_hwnd": 2, "excel_process_started_at": "now", "excel_build": "mock"}}
    return PublicationContext(
        lambda: nullcontext(), lambda current: current, journal, plan.registry_generation, plan.workbook_identity,
        native_runner=native,
        operation_id="00000000-0000-4000-8000-000000000001",
        idempotency_key="group-row-idempotency-1",
        consumer_id="construction-routing",
        operation_kind="new_row",
    )


def test_blank_fill_requires_context() -> None:
    plan = MutationPlan("existing_blank", 5, "book", "hash", 1, "construction", "RU-00000000-00-2026")
    with pytest.raises(GroupRowInsertionError, match="publication_authority_required"):
        publish_group_row(GroupRowRequest(plan, Path("source"), Path("out"), "Реестр РНС", {}), native_script=Path("x"), operation_directory=Path("ops"))


def test_mocked_blank_lifecycle_is_locked_journaled_and_published(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("existing_blank", 5, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal = Journal(); result = publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=_context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert result["published"] is True and load_workbook(output, read_only=True)["Реестр РНС"].cell(5, 6).value == plan.canonical_rns
    assert [name for name, _ in journal.calls if name in {"staged", "native", "validated", "backup_verified", "published", "finalized"}] == ["staged", "native", "validated", "backup_verified", "published", "finalized"]


def test_blank_fill_manifest_failure_blocks_fsync_backup_and_replace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("existing_blank", 5, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal, calls = Journal(), []
    failure = MutationManifestError("blank-fill-value-mismatch", subject="Реестр РНС", field="5")
    monkeypatch.setattr(insertion, "validate_blank_fill", lambda *_args, **_kwargs: (_ for _ in ()).throw(failure))
    fsync, replace_file = insertion._fsync, insertion.os.replace
    monkeypatch.setattr(insertion, "_fsync", lambda path: calls.append(path.name) if path.name in {"candidate.xlsx", "backup.xlsx"} else fsync(path))
    monkeypatch.setattr(insertion.os, "replace", lambda source, destination: calls.append("replace") or replace_file(source, destination))
    with pytest.raises(GroupRowInsertionError) as captured:
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=_context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (captured.value.code, captured.value.stage, captured.value.cause) == (failure.code, "validate", failure)
    assert calls == ["candidate.xlsx"] and not output.exists() and not list((tmp_path / "ops").rglob("backup.xlsx"))


def test_blank_fill_hyperlink_without_trusted_w_display_blocks_replace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("existing_blank", 5, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal, calls = Journal(), []
    replace_file = insertion.os.replace
    monkeypatch.setattr(insertion.os, "replace", lambda source, destination: calls.append("replace") or replace_file(source, destination))
    with pytest.raises(GroupRowInsertionError) as captured:
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, "file:///document", _context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (captured.value.code, captured.value.stage) == ("blank-fill-hyperlink-display-required", "validate")
    assert calls == [] and not output.exists() and not list((tmp_path / "ops").rglob("backup.xlsx"))


def test_v2_authority_and_evidence_are_stable_distinct_and_journaled(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("existing_blank", 5, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    request = GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns, 23: "Документ"}, "file:///one", _context(plan, Journal()))
    baseline = insertion._evidence(context=request.context, request=request, plan=plan, mode="blank_fill")  # type: ignore[arg-type]
    assert baseline == insertion._evidence(context=request.context, request=request, plan=plan, mode="blank_fill")  # type: ignore[arg-type]
    assert baseline[0] != plan.workbook_hash and baseline[1] != plan.workbook_hash and baseline[0] != baseline[1]
    assert baseline[0] != insertion._evidence(context=request.context, request=replace(request, fields={6: "changed"}), plan=plan, mode="blank_fill")[0]  # type: ignore[arg-type]
    assert baseline[0] != insertion._evidence(context=request.context, request=replace(request, hyperlink="file:///changed"), plan=plan, mode="blank_fill")[0]  # type: ignore[arg-type]
    changed_plan = replace(plan, target_row=6, workbook_hash="different-pre-hash")
    assert baseline[1] != insertion._evidence(context=request.context, request=replace(request, plan=changed_plan), plan=changed_plan, mode="blank_fill")[1]  # type: ignore[arg-type]
    journal = Journal(); result = publish_group_row(replace(request, context=_context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    created = next(values for name, values in journal.calls if name == "create")
    assert result["operation_id"] == "00000000-0000-4000-8000-000000000001"
    assert (created["operation_id"], created["idempotency_key"], created["consumer_id"], created["operation_kind"]) == (
        "00000000-0000-4000-8000-000000000001", "group-row-idempotency-1", "construction-routing", "new_row",
    )
    assert (created["intent_version"], created["manifest_version"], created["intent_digest"], created["manifest_digest"]) == (
        "group-row-intent-v2", "group-row-manifest-v2", *baseline,
    )


@pytest.mark.parametrize("fields", [{6: float("nan")}, {6: object()}])
def test_invalid_v2_intent_value_blocks_before_journal_directory_or_native(tmp_path: Path, fields: dict[int, object]) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("existing_blank", 5, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal = Journal()
    with pytest.raises(GroupRowInsertionError) as captured:
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", fields, context=_context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (captured.value.code, captured.value.stage) == ("publication_intent_value_invalid", "authorize")
    assert not journal.calls and not (tmp_path / "ops").exists() and not output.exists()


def test_exact_replay_enters_recovery_without_second_directory_or_native_mutation(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("existing_blank", 5, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal, calls = Journal(), []
    context = _context(plan, journal)
    def native(request, _script):
        calls.append(request.operation_id)
        book = load_workbook(request.candidate); book[request.sheet].cell(request.insertion_row, 6).value = plan.canonical_rns; book.save(request.candidate); book.close()
        return {"lease": {"excel_adapter": "mock", "excel_pid": 1, "excel_hwnd": 2, "excel_process_started_at": "now", "excel_build": "mock"}}
    context = replace(context, native_runner=native)
    request = GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=context)
    publish_group_row(request, native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    directories = tuple((tmp_path / "ops").iterdir())
    replay = publish_group_row(request, native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (replay["published"], replay["operation_id"], replay["recovery"], calls, tuple((tmp_path / "ops").iterdir())) == (
        False, "00000000-0000-4000-8000-000000000001", "manual_repair", ["00000000-0000-4000-8000-000000000001"], directories,
    )


def test_exact_replay_uses_planned_generation_after_independent_registry_bump(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("existing_blank", 5, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal, calls = Journal(), []
    context = _context(plan, journal)
    def native(request, _script):
        calls.append(request.operation_id)
        book = load_workbook(request.candidate); book[request.sheet].cell(request.insertion_row, 6).value = plan.canonical_rns; book.save(request.candidate); book.close()
        return {"lease": {"excel_adapter": "mock", "excel_pid": 1, "excel_hwnd": 2, "excel_process_started_at": "now", "excel_build": "mock"}}
    request = GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=replace(context, native_runner=native))
    publish_group_row(request, native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    journal.phase = "staged"; journal.operation["phase"] = "staged"
    replay = publish_group_row(replace(request, context=replace(request.context, generation=2)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")  # type: ignore[arg-type]
    assert (replay["published"], replay["recovery"], calls, journal.operation["expected_generation"]) == (False, "re_resolve_required", ["00000000-0000-4000-8000-000000000001"], 1)


def test_concurrent_create_loser_reloads_authority_without_native_or_manual_repair(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("existing_blank", 5, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal, native_started, release, calls = RacingJournal(), threading.Event(), threading.Event(), []
    generated = []
    def counted_uuid4():
        value = insertion.UUID(int=len(generated) + 1)
        generated.append(str(value))
        return value
    monkeypatch.setattr(insertion, "uuid4", counted_uuid4)
    def first_native(request, _script):
        calls.append((request.operation_id, request.owner_nonce, request.pair_nonce)); native_started.set()
        assert release.wait(timeout=3)
        book = load_workbook(request.candidate); book[request.sheet].cell(request.insertion_row, 6).value = plan.canonical_rns; book.save(request.candidate); book.close()
        return {"lease": {"excel_adapter": "mock", "excel_pid": 1, "excel_hwnd": 2, "excel_process_started_at": "now", "excel_build": "mock"}}
    first_context = replace(_context(plan, journal), native_runner=first_native)
    request = GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=first_context)
    result, failures = [], []
    def run(name: str, current: GroupRowRequest) -> None:
        try:
            result.append((name, publish_group_row(current, native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")))
        except BaseException as error:
            failures.append(error)
    worker_one = threading.Thread(target=run, args=("one", request), daemon=True)
    worker_two = threading.Thread(target=run, args=("two", request), daemon=True)
    worker_one.start(); worker_two.start(); assert native_started.wait(timeout=3), failures
    release.set(); worker_one.join(timeout=5); worker_two.join(timeout=5)
    assert not worker_one.is_alive() and not worker_two.is_alive() and not failures
    outcomes = {name: value for name, value in result}
    assert sorted(value["published"] for value in outcomes.values()) == [False, True]
    assert [value["recovery"] for value in outcomes.values() if value["published"] is False] == ["in_progress"]
    assert len(calls) == 1 and len(tuple((tmp_path / "ops").iterdir())) == 1
    assert journal.create_attempts == 2
    assert len(generated) == 2
    assert journal.operation["owner_id"] == calls[0][1] and journal.operation["pair_nonce"] == calls[0][2]
    assert "manual_repair" not in [name for name, _ in journal.calls]


def test_conflicting_replay_and_non_new_row_kind_fail_before_files_or_native(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("existing_blank", 5, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal = Journal(); context = _context(plan, journal)
    context = replace(context, journal=journal)
    request = GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=context)
    publish_group_row(request, native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    directories = tuple((tmp_path / "ops").iterdir())
    with pytest.raises(GroupRowInsertionError) as conflict:
        publish_group_row(replace(request, fields={6: "conflict"}), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (conflict.value.code, conflict.value.stage, tuple((tmp_path / "ops").iterdir())) == ("publication_intent_conflict", "recovery", directories)
    with pytest.raises(GroupRowInsertionError) as mismatch:
        publish_group_row(replace(request, context=replace(context, operation_kind="group_provision")), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "other-ops")
    assert (mismatch.value.code, mismatch.value.stage) == ("publication_operation_kind_mismatch", "preflight")
    assert not (tmp_path / "other-ops").exists()


def test_legacy_or_invalid_publication_authority_fails_closed_before_work(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("existing_blank", 5, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal = Journal(); context = _context(plan, journal)
    journal.operation = {
        "operation_id": context.operation_id, "intent_version": "group-row-intent-v1", "manifest_version": "group-row-manifest-v1",
        "owner_id": "owner", "pair_nonce": "pair", "idempotency_key": context.idempotency_key, "consumer_id": context.consumer_id,
    }
    with pytest.raises(GroupRowInsertionError) as legacy:
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=context), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (legacy.value.code, legacy.value.stage) == ("legacy_publication_authority_invalid", "recovery")
    with pytest.raises(GroupRowInsertionError) as invalid:
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=replace(context, operation_id="not-a-uuid")), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "invalid-ops")
    assert (invalid.value.code, invalid.value.stage) == ("publication_identity_invalid", "authorize")
    assert not (tmp_path / "ops").exists() and not (tmp_path / "invalid-ops").exists()


@pytest.mark.parametrize("header", [6, 10, 104])
def test_middle_insert_no_excel_is_typed_prepublication_failure(tmp_path: Path, header: int) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("insert_before_header", header, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal = Journal(); context = PublicationContext(
        lambda: nullcontext(), lambda current: current, journal, 1, "book",
        operation_id="00000000-0000-4000-8000-000000000002",
        idempotency_key="group-row-idempotency-2", consumer_id="construction-routing", operation_kind="new_row",
    )
    with pytest.raises(GroupRowInsertionError, match="excel_required_for_middle_insert"):
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {}, context=context), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert source.exists() and not output.exists()


def _patch_middle_insert_pre_oracle(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(insertion, "insertion_is_structurally_safe", lambda *_: True)
    monkeypatch.setattr(insertion, "manifest_for", lambda *_args, **_kwargs: SimpleNamespace(digest="manifest"))
    monkeypatch.setattr(insertion, "validate_control", lambda *_: None)
    monkeypatch.setattr(insertion, "validate_insertion", lambda *_: None)
    monkeypatch.setattr(insertion, "validate_inserted_row", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(insertion, "validate_dependent_registry_references", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(insertion, "validate_filter_database_middle_insert", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(insertion, "validate_worksheet_structure_middle_insert", lambda *_args, **_kwargs: None)


def test_middle_insert_validators_precede_fsync_backup_and_replace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    plan = MutationPlan("insert_before_header", 6, "book", sha256(source), 1, "construction", "RU-00000000-00-2026")
    journal = Journal(); _patch_middle_insert_pre_oracle(monkeypatch)
    calls = []
    monkeypatch.setattr(insertion, "validate_control", lambda *_: calls.append("generic-control"))
    monkeypatch.setattr(insertion, "validate_insertion", lambda *_: calls.append("generic-insertion"))
    monkeypatch.setattr(insertion, "validate_inserted_row", lambda *_args, **_kwargs: calls.append("inserted-row"))
    monkeypatch.setattr(insertion, "validate_dependent_registry_references", lambda *_args, **_kwargs: calls.append("dependents"))
    x14_call = []
    monkeypatch.setattr(insertion, "validate_x14_cf_middle_insert", lambda *args, **kwargs: (calls.append("x14"), x14_call.append((args, kwargs))))
    monkeypatch.setattr(insertion, "validate_filter_database_middle_insert", lambda *_args, **_kwargs: calls.append("filter"))
    monkeypatch.setattr(insertion, "validate_worksheet_structure_middle_insert", lambda *_args, **_kwargs: calls.append("structure"))
    fsync = insertion._fsync
    def spy_fsync(path):
        if path.name == "candidate.xlsx": calls.append("fsync")
        if path.name == "backup.xlsx": calls.append("backup")
        fsync(path)
    monkeypatch.setattr(insertion, "_fsync", spy_fsync)
    replace = insertion.os.replace
    def spy_replace(source, destination):
        calls.append("replace")
        replace(source, destination)
    monkeypatch.setattr(insertion.os, "replace", spy_replace)
    result = publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=_context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert result["published"] is True and output.exists()
    assert calls[calls.index("generic-control"):] == ["generic-control", "generic-insertion", "inserted-row", "dependents", "x14", "filter", "structure", "fsync", "backup", "replace"]
    assert len(x14_call) == 1
    (control, candidate), kwargs = x14_call[0]
    assert (control.name, candidate.name, control.parent == candidate.parent) == ("control.xlsx", "candidate.xlsx", True)
    assert kwargs == {"sheet_name": "Реестр РНС", "insertion_row": 6, "format_source_row": 5}


def test_middle_insert_x14_oracle_failure_blocks_publication_and_requests_manual_repair(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    source_hash = sha256(source)
    plan = MutationPlan("insert_before_header", 6, "book", source_hash, 1, "construction", "RU-00000000-00-2026")
    journal = Journal(); _patch_middle_insert_pre_oracle(monkeypatch)
    failure = OPCWorksheetX14CfInsertionOracleError("x14-cf-sqref-mismatch", "Реестр РНС", "sqref", "rule")
    monkeypatch.setattr(insertion, "validate_x14_cf_middle_insert", lambda *_args, **_kwargs: (_ for _ in ()).throw(failure))
    with pytest.raises(GroupRowInsertionError) as captured:
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=_context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (captured.value.code, captured.value.stage, captured.value.cause) == (failure.code, "validate", failure)
    assert sha256(source) == source_hash and not output.exists()
    assert not list((tmp_path / "ops").rglob("backup.xlsx"))
    assert "published" not in [name for name, _ in journal.calls] and journal.calls[-1][0] == "manual_repair"


def test_inserted_row_gate_failure_blocks_publication_before_x14_backup_and_replace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    source_hash = sha256(source)
    plan = MutationPlan("insert_before_header", 6, "book", source_hash, 1, "construction", "RU-00000000-00-2026")
    journal = Journal(); _patch_middle_insert_pre_oracle(monkeypatch)
    failure = MutationManifestError("inserted-row-value-mismatch", subject="Реестр РНС", field="F6")
    monkeypatch.setattr(insertion, "validate_inserted_row", lambda *_args, **_kwargs: (_ for _ in ()).throw(failure))
    with pytest.raises(GroupRowInsertionError) as captured:
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=_context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (captured.value.code, captured.value.stage, captured.value.cause) == (failure.code, "validate", failure)
    assert sha256(source) == source_hash and not output.exists() and not list((tmp_path / "ops").rglob("backup.xlsx"))
    assert "published" not in [name for name, _ in journal.calls] and journal.calls[-1][0] == "manual_repair"


def test_malformed_source_formula_is_typed_validate_failure_without_publication(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    book = load_workbook(source); book["Реестр РНС"]["Y5"] = '="unterminated'; book.save(source); book.close()
    source_hash = sha256(source)
    plan = MutationPlan("insert_before_header", 6, "book", source_hash, 1, "construction", "RU-00000000-00-2026")
    journal = Journal(); _patch_middle_insert_pre_oracle(monkeypatch)
    monkeypatch.setattr(insertion, "validate_inserted_row", mutation_manifest.validate_inserted_row)
    with pytest.raises(GroupRowInsertionError) as captured:
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=_context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (captured.value.code, captured.value.stage) == ("inserted-row-formula-translation-invalid", "validate")
    assert isinstance(captured.value.cause, MutationManifestError) and captured.value.cause.__cause__ is not None
    assert sha256(source) == source_hash and not output.exists() and not list((tmp_path / "ops").rglob("backup.xlsx"))
    assert "published" not in [name for name, _ in journal.calls] and journal.calls[-1][0] == "manual_repair"


def test_middle_insert_filter_database_oracle_failure_after_x14_blocks_publication(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    source_hash = sha256(source)
    plan = MutationPlan("insert_before_header", 6, "book", source_hash, 1, "construction", "RU-00000000-00-2026")
    journal = Journal(); _patch_middle_insert_pre_oracle(monkeypatch)
    calls = []
    monkeypatch.setattr(insertion, "validate_x14_cf_middle_insert", lambda *_args, **_kwargs: calls.append("x14"))
    failure = OPCWorkbookFilterDatabaseInsertionOracleError("filter-database-range-mismatch", "Реестр РНС", "range", "last-row")
    monkeypatch.setattr(insertion, "validate_filter_database_middle_insert", lambda *_args, **_kwargs: (_ for _ in ()).throw(failure))
    with pytest.raises(GroupRowInsertionError) as captured:
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=_context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (captured.value.code, captured.value.stage, captured.value.cause) == (failure.code, "validate", failure)
    assert calls == ["x14"]
    assert sha256(source) == source_hash and not output.exists()
    assert not list((tmp_path / "ops").rglob("backup.xlsx"))
    assert "published" not in [name for name, _ in journal.calls] and journal.calls[-1][0] == "manual_repair"


def test_middle_insert_structure_oracle_runs_after_x14_and_filter_and_blocks_publication(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "out.xlsx"; _book(source)
    source_hash = sha256(source)
    plan = MutationPlan("insert_before_header", 6, "book", source_hash, 1, "construction", "RU-00000000-00-2026")
    journal = Journal(); _patch_middle_insert_pre_oracle(monkeypatch)
    calls = []
    monkeypatch.setattr(insertion, "validate_x14_cf_middle_insert", lambda *_args, **_kwargs: calls.append("x14"))
    monkeypatch.setattr(insertion, "validate_filter_database_middle_insert", lambda *_args, **_kwargs: calls.append("filter"))
    failure = OPCWorksheetStructureInsertionOracleError("worksheet-structure-range-mismatch", "Реестр РНС", "dimension", "geometry")
    monkeypatch.setattr(insertion, "validate_worksheet_structure_middle_insert", lambda *_args, **_kwargs: calls.append("structure") or (_ for _ in ()).throw(failure))
    with pytest.raises(GroupRowInsertionError) as captured:
        publish_group_row(GroupRowRequest(plan, source, output, "Реестр РНС", {6: plan.canonical_rns}, context=_context(plan, journal)), native_script=tmp_path / "helper.ps1", operation_directory=tmp_path / "ops")
    assert (captured.value.code, captured.value.stage, captured.value.cause) == (failure.code, "validate", failure)
    assert calls == ["x14", "filter", "structure"]
    assert sha256(source) == source_hash and not output.exists() and not list((tmp_path / "ops").rglob("backup.xlsx"))


def test_third_hash_recovery_is_manual_repair(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"; _book(source); journal = Journal()
    journal.phase = "staged"; context = PublicationContext(lambda: nullcontext(), lambda current: current, journal, 1, "book")
    assert recover_group_row(context=context, operation={"operation_id": "op", "phase": "staged", "pre_hash": "old", "post_hash": "new"}, source=source) == "manual_repair"
    assert journal.calls[-1][0] == "manual_repair"
