"""Fail-closed raster field-quality regressions."""
from __future__ import annotations

import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

from rns_import_server.audit import sha256
from rns_import_server.ocr import OCRLine, OCRText, OCRWord
from rns_import_server.rns_adapter import extract
from rns_import_server.server import JobManager
from rns_import_server.workbook import SHEET, apply


NUMBER = "RU-12345678-09-2026"
GARBAGE_DISTRICT = "а Иркутской области разрешения"


def _line(page: int, text: str) -> OCRLine:
    return OCRLine(
        page=page,
        page_width=1000,
        page_height=1400,
        words=tuple(
            OCRWord(token, 20 + index * 45, 30, 35, 16, 96.1)
            for index, token in enumerate(text.split())
        ),
    )


def _garbage_record(pdf: Path) -> dict[str, object]:
    content = "\n".join(
        (
            f"Номер разрешения на строительство: {NUMBER}",
            f"Муниципальный район: {GARBAGE_DISTRICT}",
        )
    )
    record = extract(pdf, OCRText(content, (_line(1, content.splitlines()[0]), _line(1, content.splitlines()[1])), source="raster"))
    assert record is not None
    record["field_sources"] = {"district": str(pdf.resolve())}
    return record


def _district_record(pdf: Path, line: str) -> dict[str, object]:
    content = "\n".join((f"Номер разрешения на строительство: {NUMBER}", line))
    record = extract(
        pdf,
        OCRText(
            content,
            (_line(1, content.splitlines()[0]), _line(1, content.splitlines()[1])),
            source="raster",
        ),
    )
    assert record is not None
    return record


def _target(path: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["F4"] = NUMBER
    sheet["M4"] = "Жигаловский район"
    sheet["W3"] = "Ссылка на документ"
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    sheet["Z4"] = '=IF(F4<>"",ROW(),"")'
    book.save(path)


def _wait(manager: JobManager, job_id: str) -> dict[str, object]:
    for _ in range(100):
        job = manager.get(job_id)
        if job and job["status"] in {"done", "error"}:
            return job
        time.sleep(0.02)
    raise AssertionError("job did not finish")


def test_sanitized_district_garbage_stays_review_only_through_workbook_and_admin(tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    pdf = pdf_dir / "permit.pdf"
    pdf.write_bytes(b"synthetic")
    target = tmp_path / "register.xlsx"
    _target(target)

    parsed = _garbage_record(pdf)
    assert parsed["district"] == GARBAGE_DISTRICT
    assert parsed["field_quality"] == {"district": {"status": "review", "reason": "invalid_field_grammar"}}

    staged = tmp_path / "staged.xlsx"
    result = apply({NUMBER: parsed}, target, staged, sha256(target))
    assert result["changes"][0]["outcome"] == "review"
    assert load_workbook(staged)[SHEET]["M4"].value == "Жигаловский район"
    assert result["conflicts"][0]["field"] == "Муниципальный р-н"

    def runner(pdf_root: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict[str, object]:
        record = _garbage_record(pdf)
        import_result = apply({NUMBER: record}, xlsx, output, sha256(xlsx))
        import_result.update(
            input_hashes={"xlsx": sha256(xlsx), "pdfs": {pdf.name: sha256(pdf)}},
            documents=[{"file": str(pdf), "outcome": "processed_rns"}],
            logical_records=[NUMBER],
            selected_records={NUMBER: record},
        )
        return import_result

    manager = JobManager(runner, error_log=tmp_path / "error.log")
    job = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))
    assert job["status"] == "done"
    public = manager.public(str(job["id"]))
    assert public is not None
    proposal = public["proposals"][0]
    assert proposal["quality"] == "review"
    assert "id" not in proposal and "action" not in proposal
    assert load_workbook(target)[SHEET]["M4"].value == "Жигаловский район"


def test_district_before_an_empty_label_is_not_actionable(tmp_path: Path):
    pdf = tmp_path / "permit.pdf"
    pdf.write_bytes(b"synthetic")

    prefixed = _district_record(pdf, "Жигаловский район справочно Муниципальный район:")
    genuine = _district_record(pdf, "Муниципальный район: Жигаловский район")

    assert prefixed["district"] == "Жигаловский район"
    assert prefixed["field_quality"]["district"]["status"] == "review"
    assert prefixed["field_quality"]["district"]["reason"] == "missing_contiguous_value_evidence"
    assert genuine["district"] == "Жигаловский район"
    assert genuine["field_quality"]["district"]["status"] == "actionable"
