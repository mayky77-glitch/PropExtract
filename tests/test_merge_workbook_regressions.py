"""Workbook-side contracts for quarantined cross-document merge conflicts."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from rns_import_server.audit import sha256
from rns_import_server.workbook import SHEET, apply


NUMBER = "RU-12345678-09-2026"
MESSAGE = (
    "Связанные изменения содержат разные значения поля «Орган выдачи»; "
    "автоматический перенос поля не выполнен."
)


def _workbook(path: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["F4"] = NUMBER
    sheet["W3"] = "Ссылка на документ"
    book.save(path)


def _record(pdf: Path, *, with_deadline: bool) -> dict[str, object]:
    return {
        "number": NUMBER,
        "filename": pdf.name,
        "pdf": str(pdf),
        "stage": None,
        "object": None,
        "issue": None,
        "end": "31.12.2028" if with_deadline else None,
        "changed": None,
        "issuer": None,
        "builder": None,
        "region": None,
        "district": None,
        "developer": None,
        "merge_issues": [
            {
                "code": "conflicting_directive_field:issuer",
                "field": "issuer",
                "message": MESSAGE,
            }
        ],
    }


def test_merge_conflict_is_visible_in_status_while_safe_deadline_is_written(tmp_path: Path):
    source, output, pdf = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "change.pdf"
    pdf.write_bytes(b"pdf")
    _workbook(source)

    result = apply({NUMBER: _record(pdf, with_deadline=True)}, source, output, sha256(source))
    sheet = load_workbook(output)[SHEET]

    assert result["changes"][0]["outcome"] == "review"
    assert result["changes"][0]["issues"] == [MESSAGE]
    assert result["conflicts"] == []
    assert sheet["H4"].value.strftime("%d.%m.%Y") == "31.12.2028"
    assert sheet["J4"].value is None
    assert sheet["AA4"].value == MESSAGE


def test_merge_conflict_without_transferable_value_is_still_visible_in_status(tmp_path: Path):
    source, output, pdf = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "change.pdf"
    pdf.write_bytes(b"pdf")
    _workbook(source)

    result = apply({NUMBER: _record(pdf, with_deadline=False)}, source, output, sha256(source))
    sheet = load_workbook(output)[SHEET]

    assert result["changes"][0]["outcome"] == "review"
    assert result["changes"][0]["issues"] == [MESSAGE]
    assert result["conflicts"] == []
    assert sheet["J4"].value is None
    assert sheet["AA4"].value == MESSAGE
