from __future__ import annotations

import json
import hashlib
import socket
import subprocess
import threading
import time
import urllib.error
import urllib.request
import zipfile
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from rns_import_server.audit import sha256
from rns_import_server import app, mapping, ocr, picker, rns_adapter, server
from rns_import_server.normalization import canonical_rns_identities, normalize_text
from rns_import_server.files import discover_pdfs
from rns_import_server.ocr import (
    LANGUAGE_HASHES,
    bundled_language_status,
    project_windows_tool,
    tesseract_environment,
)
from rns_import_server.runtime import _is_supported_tesseract_version, runtime_status
from rns_import_server.server import JobManager, create_server, error_hint, retry_file_operation, user_path, validated_job_paths
from rns_import_server.workbook import SHEET, _change_outcome, _row_by_number, _validate, apply, transfer_issue
from scripts.build_windows_python_runtime import build as build_windows_python_runtime


def _fake_runner(pdf_dir: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict:
    assert dpi == 180 and max_pages == 0
    progress(30, "Распознаём PDF", "sample.pdf")
    progress(80, "Переносим данные в Excel", None)
    output.write_bytes(xlsx.read_bytes() + b"-updated")
    return {
        "input_hashes": {"xlsx": sha256(xlsx)},
        "documents": [{"file": str(pdf_dir / "sample.pdf")}],
        "logical_records": ["00-00-00-0000"],
        "changes": [{
            "new": False,
            "row": 42,
            "issues": [],
            "outcome": "added",
            "document": "sample.pdf",
        }],
        "conflicts": [],
    }


def _wait(manager: JobManager, job_id: str) -> dict:
    for _ in range(100):
        job = manager.get(job_id)
        if job and job["status"] in {"done", "error"}:
            return job
        time.sleep(0.02)
    raise AssertionError("job did not finish")


def _proposal_job(tmp_path: Path, values: dict[str, str]) -> tuple[JobManager, dict, Path, Path]:
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    pdf = pdf_dir / "sample.pdf"
    pdf.write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["W3"] = "Ссылка на документ"
    sheet["F4"] = "38-1-1-2026"
    sheet["D4"] = "Старый объект"
    sheet["H4"] = datetime(2025, 12, 31)
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    sheet["Z4"] = '=IF(F4<>"",ROW(),"")'
    book.save(target)

    def runner(pdf_root: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict:
        record = _synthetic_record("38-1-1-2026", pdf)
        record.update(values)
        result = apply({"38-1-1-2026": record}, xlsx, output, sha256(xlsx))
        result.update({
            "input_hashes": {"xlsx": sha256(xlsx), "pdfs": {pdf.name: sha256(pdf)}},
            "documents": [{"file": str(pdf)}],
            "logical_records": ["38-1-1-2026"],
            "selected_records": {
                "38-1-1-2026": {
                    **values,
                    "filename": pdf.name,
                    "field_sources": {field: pdf.name for field in values},
                }
            },
        })
        return result

    manager = JobManager(runner, error_log=tmp_path / "error.log")
    finished = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))
    assert finished["status"] == "done"
    return manager, finished, target, pdf


def test_job_replaces_target_only_after_verified_backup(tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    (pdf_dir / "sample.pdf").write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")
    manager = JobManager(_fake_runner)

    started = manager.start(str(pdf_dir), str(target))
    finished = _wait(manager, str(started["id"]))

    assert finished["status"] == "done"
    assert target.read_bytes() == b"original-updated"
    backup = Path(str(finished["backup"]))
    assert backup.read_bytes() == b"original"
    assert finished["summary"] == {
        "pdf_count": 1,
        "failed_pdf_count": 0,
        "record_count": 1,
        "changed_rows": 1,
        "new_rows": 0,
        "already_present_count": 0,
        "already_present_files": [],
        "already_present_rows": [],
        "conflicts": 0,
        "issue_count": 0,
        "rows_with_issues": [],
        "row_numbers": [42],
        "new_row_numbers": [],
    }


def test_merge_review_issue_is_public_and_ui_groups_it_for_operator_decision(tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    pdf = pdf_dir / "изменение.pdf"
    pdf.write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")
    message = (
        "Связанные изменения содержат разные значения поля «Орган выдачи»; "
        "автоматический перенос поля не выполнен."
    )

    def runner(pdf_root: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict:
        output.write_bytes(xlsx.read_bytes() + b"-review")
        return {
            "input_hashes": {"xlsx": sha256(xlsx), "pdfs": {pdf.name: sha256(pdf)}},
            "documents": [{"file": str(pdf)}],
            "logical_records": ["38-1-1-2026"],
            "changes": [{
                "number": "38-1-1-2026",
                "new": True,
                "row": 42,
                "issues": [message],
                "outcome": "review",
                "document": pdf.name,
            }],
            "conflicts": [],
            "selected_records": {
                "38-1-1-2026": {
                    "filename": pdf.name,
                    "pdf": str(pdf),
                    "object": "Этап 1. Проверочный объект",
                    "merge_issues": [{"field": "issuer", "message": message}],
                }
            },
        }

    manager = JobManager(runner)
    finished = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))
    public = manager.public(str(finished["id"]))

    assert public is not None and public["status"] == "done"
    assert public["proposals"] == []
    assert public["row_cards"] == [{
        "row": 42,
        "number": "38-1-1-2026",
        "object": "Этап 1. Проверочный объект",
        "details": message,
        "outcome": "review",
        "needs_review": True,
        "filename": pdf.name,
        "document_id": public["documents"][0]["id"],
    }]
    javascript = (Path(__file__).parents[1] / "rns_import_server/static/app.js").read_text(encoding="utf-8")
    assert "function renderReviewCard" in javascript
    assert ".filter(item => item.needs_review" in javascript
    assert "item.details ||" in javascript
    assert "Строки для проверки" in javascript


def test_merge_review_details_remain_visible_when_same_row_has_a_proposal(tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    pdf = pdf_dir / "изменение.pdf"
    pdf.write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")
    message = (
        "Связанные изменения содержат разные значения поля «Орган выдачи»; "
        "автоматический перенос поля не выполнен."
    )

    def runner(pdf_root: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict:
        output.write_bytes(xlsx.read_bytes() + b"-mixed-review")
        return {
            "input_hashes": {"xlsx": sha256(xlsx), "pdfs": {pdf.name: sha256(pdf)}},
            "documents": [{"file": str(pdf)}],
            "logical_records": ["38-1-1-2026"],
            "changes": [{
                "number": "38-1-1-2026",
                "new": False,
                "row": 42,
                "issues": [message, "Не перенесено «Наименование объекта»: значения различаются."],
                "outcome": "review",
                "document": pdf.name,
            }],
            "conflicts": [{
                "number": "38-1-1-2026",
                "cell": "D42",
                "field": "Наименование объекта",
                "existing": "Старый объект",
                "pdf": "Новый объект",
                "action": "Перенести изменения",
            }],
            "selected_records": {
                "38-1-1-2026": {
                    "filename": pdf.name,
                    "pdf": str(pdf),
                    "object": "Новый объект",
                    "field_sources": {"object": str(pdf)},
                    "merge_issues": [{"field": "issuer", "message": message}],
                }
            },
        }

    manager = JobManager(runner)
    finished = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))
    public = manager.public(str(finished["id"]))

    assert public is not None and len(public["proposals"]) == 1
    assert public["proposals"][0]["review_details"] == message
    assert public["row_cards"][0]["needs_review"] is True
    javascript = (Path(__file__).parents[1] / "rns_import_server/static/app.js").read_text(encoding="utf-8")
    assert "Связанные документы требуют проверки" in javascript
    assert "item.review_details" in javascript
    assert 'item.status !== "approved" || item.review_details' in javascript
    assert 'item.status === "approved" && !item.review_details' in javascript
    assert "Поле перенесено — нужна проверка" in javascript


def test_approved_field_keeps_unresolved_merge_review_visible_and_in_excel_status(tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    pdf = pdf_dir / "изменение.pdf"
    pdf.write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["W3"] = "Ссылка на документ"
    sheet["F4"] = "38-1-1-2026"
    sheet["D4"] = "Старый объект"
    book.save(target)
    message = (
        "Связанные изменения содержат разные значения поля «Орган выдачи»; "
        "автоматический перенос поля не выполнен."
    )

    def runner(pdf_root: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict:
        record = _synthetic_record("38-1-1-2026", pdf)
        record.update({
            "object": "Новый объект",
            "merge_issues": [{"field": "issuer", "message": message}],
        })
        result = apply({"38-1-1-2026": record}, xlsx, output, sha256(xlsx))
        result.update({
            "input_hashes": {"xlsx": sha256(xlsx), "pdfs": {pdf.name: sha256(pdf)}},
            "documents": [{"file": str(pdf)}],
            "logical_records": ["38-1-1-2026"],
            "selected_records": {
                "38-1-1-2026": {
                    "filename": pdf.name,
                    "pdf": str(pdf),
                    "object": "Новый объект",
                    "field_sources": {"object": str(pdf)},
                    "merge_issues": [{"field": "issuer", "message": message}],
                }
            },
        })
        return result

    manager = JobManager(runner)
    finished = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))
    assert finished["status"] == "done" and len(finished["proposals"]) == 1
    proposal = finished["proposals"][0]

    manager.approve(str(finished["id"]), str(proposal["id"]), str(finished["capability"]))
    public = manager.public(str(finished["id"]))
    saved = load_workbook(target)[SHEET]

    assert public is not None
    assert public["proposals"][0]["status"] == "approved"
    assert public["proposals"][0]["review_details"] == message
    assert saved["D4"].value == "Новый объект"
    assert saved["AA4"].value == message


def test_job_refuses_external_edit_after_verified_backup(monkeypatch, tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    (pdf_dir / "sample.pdf").write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")
    original_copy = server.shutil.copy2

    def copy_then_user_edit(source: Path, destination: Path):
        copied = original_copy(source, destination)
        Path(source).write_bytes(b"user-edit-after-backup")
        return copied

    monkeypatch.setattr(server.shutil, "copy2", copy_then_user_edit)
    manager = JobManager(_fake_runner, error_log=tmp_path / "error.log")
    finished = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))

    assert finished["status"] == "error"
    assert target.read_bytes() == b"user-edit-after-backup"
    assert not list((target.parent / "Резервные копии PropExtract").glob("*.xlsx"))
    assert not list(target.parent.glob(f".{target.stem}.propextract-*.xlsx"))


def test_import_uses_shared_publication_lock(tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    (pdf_dir / "sample.pdf").write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")
    manager = JobManager(_fake_runner, error_log=tmp_path / "error.log")

    manager._publish_lock.acquire()
    try:
        job_id = str(manager.start(str(pdf_dir), str(target))["id"])
        for _ in range(100):
            current = manager.get(job_id) or {}
            if current.get("stage") == "Создаём резервную копию":
                break
            time.sleep(0.01)
        else:
            raise AssertionError("import did not reach publication")
        assert target.read_bytes() == b"original"
        assert current["status"] == "running"
    finally:
        manager._publish_lock.release()

    finished = _wait(manager, job_id)
    assert finished["status"] == "done"
    assert target.read_bytes() == b"original-updated"


def test_job_with_only_already_present_changes_keeps_workbook_unchanged(tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    (pdf_dir / "sample.pdf").write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")
    source_hash = sha256(target)
    source_mtime_ns = target.stat().st_mtime_ns

    def noop_runner(pdf_root: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict:
        output.write_bytes(b"would-have-been-published")
        return {
            "input_hashes": {"xlsx": sha256(xlsx)},
            "documents": [{"file": str(pdf_root / "sample.pdf")}],
            "logical_records": ["00-00-00-0000"],
            "changes": [{
                "new": False,
                "row": 42,
                "issues": [],
                "outcome": "already_present",
                "document": "sample.pdf",
            }],
            "conflicts": [],
        }

    manager = JobManager(noop_runner, error_log=tmp_path / "error.log")
    finished = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))

    report = target.with_name(f"{target.stem} — отчет PropExtract.json")
    assert finished["status"] == "done"
    assert finished["target_hash"] == source_hash
    assert target.read_bytes() == b"original"
    assert sha256(target) == source_hash
    assert target.stat().st_mtime_ns == source_mtime_ns
    backup_dir = target.parent / "Резервные копии PropExtract"
    assert not backup_dir.exists()
    assert list(backup_dir.glob("*.xlsx")) == []
    assert not list(target.parent.glob(f".{target.stem}.propextract-*.xlsx"))
    assert report.exists()
    report_data = json.loads(report.read_text(encoding="utf-8"))
    assert report_data["input_hashes"]["xlsx"] == source_hash
    assert report_data["backup"] is None
    assert finished["backup"] is None


def test_disk_report_redacts_source_paths_while_server_retains_capability_path(tmp_path: Path):
    pdf_dir = tmp_path / "private-pdfs"
    pdf_dir.mkdir()
    pdf = pdf_dir / "permit.pdf"
    pdf.write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")

    def runner(pdf_root: Path, xlsx: Path, output: Path, dpi: int, max_pages: int, progress=None) -> dict:
        output.write_bytes(b"updated")
        return {
            "input_hashes": {"xlsx": sha256(xlsx), "pdfs": {pdf.name: sha256(pdf)}},
            "documents": [{"file": str(pdf), "ocr_text": "private OCR text"}],
            "logical_records": ["00-00-00-0000"],
            "selected_records": {"00-00-00-0000": {"pdf": str(pdf), "field_sources": {"object": str(pdf)}}},
            "changes": [{"new": False, "row": 42, "issues": [], "outcome": "added", "document": pdf.name}],
            "conflicts": [],
        }

    manager = JobManager(runner, error_log=tmp_path / "error.log")
    finished = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))
    report_data = json.loads(Path(str(finished["report"])).read_text(encoding="utf-8"))

    assert finished["status"] == "done"
    assert str(pdf) not in json.dumps(report_data)
    assert "private OCR text" not in json.dumps(report_data)
    internal = finished["documents_internal"]
    assert next(iter(internal.values()))["path"] == pdf


def test_partial_pdf_failure_is_reported_without_stopping_valid_records(tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    (pdf_dir / "sample.pdf").write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")

    def partial_runner(*args, **kwargs):
        result = _fake_runner(*args, **kwargs)
        result["documents"].append({"file": str(pdf_dir / "broken.pdf"), "error": "pdfinfo_failed"})
        return result

    manager = JobManager(partial_runner, error_log=tmp_path / "error.log")
    finished = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))

    assert finished["status"] == "done"
    assert finished["summary"]["pdf_count"] == 1
    assert finished["summary"]["failed_pdf_count"] == 1
    assert finished["warning"] == "PDF пропущено: 1. Причины сохранены в отчёте."


def test_mapping_is_disabled_by_default_and_rejects_invalid_model_output(monkeypatch):
    record = {"object": "Исходное значение", "mapping_candidates": [{"id": "candidate-1", "value": "Кандидат", "allowed_targets": ["object"]}]}

    assert mapping.map_extracted_record(record) is record

    monkeypatch.setenv("RNS_MAPPING_LLM_ENDPOINT", "http://127.0.0.1:11434/api/chat")
    monkeypatch.setenv("RNS_MAPPING_LLM_MODEL", "local-test")
    monkeypatch.setattr(mapping, "_request", lambda *args: {"assignments": [{"candidate_id": "unknown", "target": "object"}]})

    assert mapping.map_extracted_record(record) is record


def test_mapping_accepts_only_existing_candidates_and_known_targets(monkeypatch):
    record = {
        "mapping_candidates": [
            {"id": "candidate-a", "value": "Объект из PDF", "allowed_targets": ["object"]},
            {"id": "candidate-b", "value": "Орган из PDF", "allowed_targets": ["issuer"]},
        ],
        "warnings": ["object", "issuer", "end"],
    }
    monkeypatch.setenv("RNS_MAPPING_LLM_ENDPOINT", "http://127.0.0.1:11434/api/chat")
    monkeypatch.setenv("RNS_MAPPING_LLM_MODEL", "local-test")
    monkeypatch.setattr(mapping, "_request", lambda *args: {"assignments": [
        {"candidate_id": "candidate-a", "target": "object"},
        {"candidate_id": "candidate-b", "target": "issuer"},
    ]})

    assert mapping.map_extracted_record(record) == {
        **record, "object": "Объект из PDF", "issuer": "Орган из PDF",
        "field_provenance": {"object": "mapping_llm", "issuer": "mapping_llm"}, "warnings": ["end"],
    }


def test_mapping_rejects_overwrite_duplicate_or_disallowed_assignments(monkeypatch):
    record = {
        "object": "Детерминированное значение",
        "mapping_candidates": [
            {"id": "candidate-a", "value": "Новое", "allowed_targets": ["object", "issuer"]},
            {"id": "candidate-b", "value": "Орган", "allowed_targets": ["issuer"]},
        ],
    }
    monkeypatch.setenv("RNS_MAPPING_LLM_ENDPOINT", "http://127.0.0.1:11434/api/chat")
    monkeypatch.setenv("RNS_MAPPING_LLM_MODEL", "local-test")
    monkeypatch.setattr(mapping, "_request", lambda *args: {"assignments": [
        {"candidate_id": "candidate-a", "target": "issuer"},
        {"candidate_id": "candidate-a", "target": "object"},
    ]})

    assert mapping.map_extracted_record(record) is record


def test_mapping_request_is_bounded_and_endpoint_rejects_credentials(monkeypatch):
    monkeypatch.setenv("RNS_MAPPING_LLM_ENDPOINT", "http://user:password@127.0.0.1:11434/api/chat")
    monkeypatch.setenv("RNS_MAPPING_LLM_MODEL", "local-test")
    assert mapping._endpoint() is None

    captured = {}
    class Response:
        headers = {"Content-Length": "2"}
        def __enter__(self): return self
        def __exit__(self, *args): return None
        def read(self, size):
            captured["read_size"] = size
            return b"{}"
    def open_request(request, timeout):
        captured["payload"] = json.loads(request.data)
        captured["timeout"] = timeout
        return Response()
    assert isinstance(mapping._LOOPBACK_PROXY_HANDLER, urllib.request.ProxyHandler)
    assert mapping._LOOPBACK_PROXY_HANDLER.proxies == {}
    assert not any(
        isinstance(handler, urllib.request.ProxyHandler)
        for handler in mapping._LOOPBACK_OPENER.handlers
    )
    monkeypatch.setattr(mapping._LOOPBACK_OPENER, "open", open_request)

    assert mapping._request("http://127.0.0.1:11434/api/chat", "local-test", [{"id": "a", "value": "x", "allowed_targets": ["object"]}]) is None
    assert captured["timeout"] == 3 and captured["read_size"] == 65537
    assert captured["payload"]["options"] == {"temperature": 0, "num_ctx": 1024, "num_predict": 128}
    assert captured["payload"]["keep_alive"] == "0" and captured["payload"]["think"] is False


def test_mapping_candidate_limits_and_invalid_target_types_fail_closed(monkeypatch):
    def candidate(*, candidate_id: str = "candidate", value: str = "value", targets=None):
        return {"id": candidate_id, "value": value, "allowed_targets": ["object"] if targets is None else targets}

    exact_count = [candidate(candidate_id=f"candidate-{index}") for index in range(mapping._MAX_CANDIDATES)]
    assert len(mapping._candidates({"mapping_candidates": exact_count})) == mapping._MAX_CANDIDATES
    assert mapping._candidates({"mapping_candidates": exact_count + [candidate(candidate_id="candidate-extra")]}) == []

    exact_id = "i" * mapping._MAX_CANDIDATE_ID_CHARS
    assert mapping._candidates({"mapping_candidates": [candidate(candidate_id=exact_id)]})
    assert mapping._candidates({"mapping_candidates": [candidate(candidate_id="i" * (mapping._MAX_CANDIDATE_ID_CHARS + 1))]}) == []

    exact_value = "v" * mapping._MAX_CANDIDATE_VALUE_CHARS
    assert mapping._candidates({"mapping_candidates": [candidate(value=exact_value)]})
    assert mapping._candidates({"mapping_candidates": [candidate(value="v" * (mapping._MAX_CANDIDATE_VALUE_CHARS + 1))]}) == []

    all_targets = sorted(mapping.TARGET_FIELDS)
    assert mapping._candidates({"mapping_candidates": [candidate(targets=all_targets)]})
    assert mapping._candidates({"mapping_candidates": [candidate(targets=all_targets + ["object"])]}) == []
    assert mapping._candidates({"mapping_candidates": [candidate(targets=[[]])]}) == []
    assert mapping._candidates({"mapping_candidates": [candidate(targets=[{"target": "object"}])]}) == []

    record = {"mapping_candidates": [candidate()]}
    accepted = mapping._candidates(record)
    serialized_size = len(json.dumps(accepted, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
    monkeypatch.setattr(mapping, "_MAX_CANDIDATES_BYTES", serialized_size)
    assert mapping._candidates(record)
    monkeypatch.setattr(mapping, "_MAX_CANDIDATES_BYTES", serialized_size - 1)
    assert mapping._candidates(record) == []


@pytest.mark.parametrize("body", [b"[]", b'{"message":[]}', b'{"message":{"content":"[]"}}', b'{"message":{"content":"null"}}'])
def test_mapping_request_rejects_valid_json_with_wrong_top_level_or_message_shape(monkeypatch, body):
    class Response:
        headers = {"Content-Length": str(len(body))}

        def __enter__(self): return self
        def __exit__(self, *args): return None
        def read(self, size): return body

    monkeypatch.setattr(mapping._LOOPBACK_OPENER, "open", lambda *args, **kwargs: Response())
    assert mapping._request("http://127.0.0.1:11434/api/chat", "local-test", [{"id": "a", "value": "x", "allowed_targets": ["object"]}]) is None


def test_mapping_request_payload_limit_rejects_before_network(monkeypatch):
    captured = {"calls": 0}

    class Response:
        headers = {"Content-Length": "2"}

        def __enter__(self): return self
        def __exit__(self, *args): return None
        def read(self, size): return b"{}"

    def open_request(request, timeout):
        captured["calls"] += 1
        captured["payload"] = request.data
        return Response()

    candidates = [{"id": "a", "value": "x", "allowed_targets": ["object"]}]
    monkeypatch.setattr(mapping._LOOPBACK_OPENER, "open", open_request)
    assert mapping._request("http://127.0.0.1:11434/api/chat", "local-test", candidates) is None
    payload_size = len(captured["payload"])
    monkeypatch.setattr(mapping, "_MAX_REQUEST_BYTES", payload_size)
    assert mapping._request("http://127.0.0.1:11434/api/chat", "local-test", candidates) is None
    assert captured["calls"] == 2
    monkeypatch.setattr(mapping, "_MAX_REQUEST_BYTES", payload_size - 1)
    assert mapping._request("http://127.0.0.1:11434/api/chat", "local-test", candidates) is None
    assert captured["calls"] == 2


def test_failed_job_leaves_target_unchanged(tmp_path: Path):
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")
    (tmp_path / "sample.pdf").write_bytes(b"pdf")

    def fail(*args, **kwargs):
        raise RuntimeError("test failure")

    error_log = tmp_path / "propextract-error.log"
    manager = JobManager(fail, error_log=error_log)
    finished = _wait(manager, str(manager.start(str(tmp_path), str(target))["id"]))
    assert finished["status"] == "error"
    assert target.read_bytes() == b"original"
    assert not (tmp_path / "Резервные копии PropExtract").exists()
    assert finished["error_hint"] == "Исправьте указанную причину и повторите запуск. Исходный Excel не изменён."
    assert finished["error_log"] == str(error_log)
    assert "RuntimeError: test failure" in error_log.read_text(encoding="utf-8")


def test_new_record_style_is_allowed_but_existing_row_style_change_is_rejected(tmp_path: Path):
    source = tmp_path / "register.xlsx"
    output = tmp_path / "output.xlsx"
    pdf = tmp_path / "new-record.pdf"
    pdf.write_bytes(b"pdf")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET
    sheet["W3"] = "Ссылка на документ"
    sheet["A4"] = 1
    sheet["F4"] = "38-1-1-2026"
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    sheet["Z4"] = '=IF(F4<>"",ROW(),"")'
    sheet["A4"].fill = PatternFill("solid", fgColor="00FF00")
    sheet["A5"].fill = PatternFill("solid", fgColor="FFFF00")
    sheet["Y5"] = '=IF(A5<>"",ROW(),"")'
    sheet["Z5"] = '=IF(F5<>"",ROW(),"")'
    workbook.save(source)
    record = {
        "stage": "13.5",
        "object": "Новый объект",
        "issue": "10.08.2026",
        "end": "10.08.2027",
        "changed": "10.08.2026",
        "issuer": "Администрация",
        "builder": "Застройщик",
        "region": "Иркутская область",
        "district": "Иркутский район",
        "developer": "Разработчик",
        "filename": pdf.name,
        "pdf": str(pdf),
    }

    result = apply({"38-2-2-2026": record}, source, output, sha256(source))

    saved_book = load_workbook(output)
    saved = saved_book[SHEET]
    assert result["changes"][0]["row"] == 5
    assert saved["F5"].value == "38-2-2-2026"
    assert saved["A5"]._style == saved["A4"]._style
    assert saved["Y5"].value == '=IF(A5<>"",ROW(),"")'
    assert saved["Z5"].value == '=IF(F5<>"",ROW(),"")'
    saved["W5"] = "Старая ссылка"
    saved["W5"].hyperlink = "https://example.invalid/old.pdf"
    saved["AA5"] = "Старый статус"
    saved_book.save(output)
    repeated = apply({"38-2-2-2026": record}, output, tmp_path / "repeated.xlsx", sha256(output))
    assert repeated["changes"][0]["outcome"] == "already_present"
    repeated_sheet = load_workbook(tmp_path / "repeated.xlsx")[SHEET]
    assert repeated_sheet["W5"].value == "Старая ссылка"
    assert repeated_sheet["W5"].hyperlink.target == "https://example.invalid/old.pdf"
    assert repeated_sheet["AA5"].value == "Старый статус"

    saved["A4"].fill = PatternFill("solid", fgColor="FF0000")
    saved_book.save(output)
    with pytest.raises(RuntimeError, match=r"^style_changed:A4$"):
        _validate(
            source,
            output,
            {"38-2-2-2026": record},
            {"38-2-2-2026": result["changes"][0]["status"]},
            {"38-2-2-2026": "added"},
        )


def test_empty_native_ocr_stdout_is_safe_text(monkeypatch, tmp_path: Path):
    image = tmp_path / "blank-page.png"
    image.write_bytes(b"png")
    monkeypatch.setattr(
        ocr,
        "_run",
        lambda *args, **kwargs: subprocess.CompletedProcess(args[0], 0, None, None),
    )

    assert ocr._ocr_image(image, "tesseract") == ""
    assert ocr._captured_text(None) == ""
    assert ocr._captured_text("текст") == "текст"


def test_native_process_output_is_always_decoded_as_utf8(monkeypatch):
    captured: dict[str, object] = {}

    def completed(argv, **kwargs):
        captured.update(kwargs)
        return subprocess.CompletedProcess(argv, 0, "текст", "")

    monkeypatch.setattr(ocr.subprocess, "run", completed)
    result = ocr._run(["tesseract"], timeout=1)

    assert result.stdout == "текст"
    assert captured["encoding"] == "utf-8"
    assert captured["errors"] == "replace"


def test_pdf_discovery_is_case_insensitive_and_skips_symlinks(tmp_path: Path):
    lower = tmp_path / "lower.pdf"
    upper = tmp_path / "UPPER.PDF"
    lower.write_bytes(b"pdf")
    upper.write_bytes(b"pdf")
    (tmp_path / "notes.txt").write_text("not a pdf", encoding="utf-8")
    try:
        (tmp_path / "linked.pdf").symlink_to(lower)
    except OSError:
        pass

    assert discover_pdfs(tmp_path) == [lower, upper]


def test_collect_keeps_valid_pdf_when_another_pdf_fails(monkeypatch, tmp_path: Path):
    broken = tmp_path / "broken.pdf"
    valid = tmp_path / "VALID.PDF"
    broken.write_bytes(b"broken")
    valid.write_bytes(b"valid")

    def read(pdf: Path, dpi: int, max_pages: int):
        if pdf == broken:
            raise RuntimeError("pdfinfo_failed")
        return "38-1-1-2026", 1

    def extract(pdf: Path, text: str):
        return {
            "number": "38-1-1-2026",
            "changed": None,
            "end": None,
            "filename": pdf.name,
            "pdf": str(pdf),
            "warnings": [],
        }

    monkeypatch.setattr(app, "read_ocr", read)
    monkeypatch.setattr(app, "extract", extract)
    records, documents = app.collect(tmp_path, 180, 0)

    assert list(records) == ["38-1-1-2026"]
    assert len(documents) == 2
    assert next(item for item in documents if item["file"] == str(broken))["error"] == "pdfinfo_failed"


@pytest.mark.parametrize(
    ("filename", "text", "expected"),
    [
        ("permit.pdf", "Разрешение № RU-12345678-09-2026", "RU-12345678-09-2026"),
        ("permit.pdf", "Разрешение №RU-12345678-09-2026", "RU-12345678-09-2026"),
        ("РY 12345678_09_2026.pdf", "скан без номера", "RU-12345678-09-2026"),
        ("legacy.pdf", "№ 3В–7–2–2026", "38-7-2-2026"),
        ("legacy.pdf", "№38-7-2-2026", "38-7-2-2026"),
        ("3807122026.pdf", "скан без номера", "38-07-12-2026"),
    ],
)
def test_rns_identity_accepts_sanitized_modern_ocr_and_legacy_variants(filename: str, text: str, expected: str):
    assert rns_adapter.norm(Path(filename), text) == expected


def test_rns_extracts_multiline_label_blocks_and_rejects_ambiguous_identity():
    text = """№ РУ-12345678-09-2026
Дата выдачи
05.01.2026
Срок действия
до 05.01.2027
Дата последнего изменения
06.02.2026
Наименование объекта
Синтетический объект\nв две строки
Орган выдачи
Тестовый муниципальный орган
Застройщик
ООО «Синтетический застройщик»
Разработчик ПД
ООО «Синтетический проектировщик»
Субъект РФ
Тестовая область
Муниципальный район
Тестовый район
"""
    record = rns_adapter.extract(Path("form.pdf"), text)

    assert record is not None
    assert {field: record[field] for field in ("number", "issue", "end", "changed")} == {
        "number": "RU-12345678-09-2026",
        "issue": "05.01.2026",
        "end": "05.01.2027",
        "changed": "06.02.2026",
    }
    assert record["object"] == "Синтетический объект в две строки"
    assert record["issuer"] == "Тестовый муниципальный орган"
    assert record["builder"] == "ООО «Синтетический застройщик»"
    assert record["developer"] == "ООО «Синтетический проектировщик»"
    assert record["region"] == "Тестовая область"
    assert record["district"] == "Тестовый район"
    assert rns_adapter.norm(Path("none.pdf"), "без идентификатора") is None
    assert rns_adapter.norm(Path("ambiguous.pdf"), "RU-12345678-09-2026; RU-87654321-09-2026") is None


def test_ambiguous_body_identities_block_filename_fallback():
    text = "RU-12345678-09-2026; RU-87654321-09-2026"
    pdf = Path("RU-99999999-09-2026.pdf")

    assert canonical_rns_identities(text) == ("RU-12345678-09-2026", "RU-87654321-09-2026")
    assert rns_adapter.norm(pdf, text) is None
    assert rns_adapter.extract(pdf, text) is None


def test_collect_enriches_permit_only_with_same_explicit_id(monkeypatch, tmp_path: Path):
    permit, amendment = tmp_path / "permit.pdf", tmp_path / "изменение.pdf"
    permit.write_bytes(b"pdf")
    amendment.write_bytes(b"pdf")
    records = {
        permit: {"number": "RU-12345678-09-2026", "issue": "05.01.2026", "object": "Синтетический объект", "changed": None},
        amendment: {"number": "RU-12345678-09-2026", "issue": None, "object": None, "changed": "06.02.2026", "end": "05.01.2027"},
    }

    monkeypatch.setattr(app, "read_ocr", lambda pdf, *args: (pdf.name, 1))
    monkeypatch.setattr(
        app,
        "extract",
        lambda pdf, text: {
            "filename": pdf.name,
            "pdf": str(pdf),
            "warnings": [],
            **{field: None for field in app.EVIDENCE_FIELDS},
            **records[pdf],
        },
    )
    selected, documents = app.collect(tmp_path, 180, 0, pdfs=[permit, amendment])

    assert documents[0]["number"] == "RU-12345678-09-2026"
    assert selected["RU-12345678-09-2026"]["changed"] == "06.02.2026"
    assert selected["RU-12345678-09-2026"]["end"] == "05.01.2027"
    assert selected["RU-12345678-09-2026"]["source_files"] == ["permit.pdf", "изменение.pdf"]
    assert selected["RU-12345678-09-2026"]["field_sources"]["object"] == str(permit)
    assert selected["RU-12345678-09-2026"]["field_sources"]["changed"] == str(amendment)
    assert selected["RU-12345678-09-2026"]["field_sources"]["end"] == str(amendment)


def test_collect_keeps_idless_directive_unlinked_with_diagnostic(monkeypatch, tmp_path: Path):
    directive = tmp_path / "распоряжение-продление.pdf"
    directive.write_bytes(b"pdf")
    monkeypatch.setattr(app, "read_ocr", lambda pdf, *args: ("продление", 1))
    monkeypatch.setattr(
        app,
        "extract",
        lambda pdf, text: {"number": None, "filename": pdf.name, "pdf": str(pdf), "warnings": [], **{field: None for field in app.EVIDENCE_FIELDS}},
    )

    selected, documents = app.collect(tmp_path, 180, 0, pdfs=[directive])

    assert selected == {}
    diagnostic = documents[-1]
    assert diagnostic["warnings"] == ["unlinked_directive"]
    assert "не связано" in diagnostic["error"].casefold()


def _synthetic_record(number: str, pdf: Path) -> dict[str, object]:
    return {
        "stage": None, "object": None, "issue": None, "end": None, "changed": None,
        "issuer": None, "builder": None, "region": None, "district": None, "developer": None,
        "filename": pdf.name, "pdf": str(pdf), "number": number,
    }


def test_decorated_workbook_identity_updates_only_unambiguous_row_and_preserves_formulas(tmp_path: Path):
    source, output, pdf = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "permit.pdf"
    pdf.write_bytes(b"pdf")
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["W3"] = "Ссылка на документ"
    sheet["F4"] = "№ RU-12345678-09-2026 от 05.01.2026\nизм от 06.02.2026"
    sheet["B4"] = "1.0"
    sheet["D4"] = "Синтетический объект"
    sheet["G4"] = "05.01.2026"
    sheet["H4"] = "05.01.2027"
    sheet["I4"] = "06.02.2026"
    sheet["J4"] = "Тестовый орган"
    sheet["K4"] = "Тестовый застройщик"
    sheet["L4"] = "Тестовая область"
    sheet["M4"] = "Тестовый район"
    sheet["N4"] = "Тестовый разработчик"
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    sheet["Z4"] = '=IF(F4<>"",ROW(),"")'
    book.save(source)

    record = _synthetic_record("RU-12345678-09-2026", pdf)
    record.update({
        "stage": "1.0", "object": "Синтетический объект", "issue": "05.01.2026",
        "end": "05.01.2027", "changed": "06.02.2026", "issuer": "Тестовый орган",
        "builder": "Тестовый застройщик", "region": "Тестовая область",
        "district": "Тестовый район", "developer": "Тестовый разработчик",
    })
    result = apply({"RU-12345678-09-2026": record}, source, output, sha256(source))
    saved = load_workbook(output)[SHEET]

    assert result["changes"][0]["outcome"] == "already_present"
    assert saved["F4"].value == "№ RU-12345678-09-2026 от 05.01.2026\nизм от 06.02.2026"
    assert saved["Y4"].value == '=IF(A4<>"",ROW(),"")'
    assert saved["Z4"].value == '=IF(F4<>"",ROW(),"")'
    assert _row_by_number(saved, "RU-12345678-09-2026") == 4


def test_duplicate_or_ambiguous_workbook_identity_is_conflict_without_publish_mutation(tmp_path: Path):
    source, output, pdf = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "permit.pdf"
    pdf.write_bytes(b"pdf")
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["W3"] = "Ссылка на документ"
    sheet["F4"] = "RU-12345678-09-2026"
    sheet["F5"] = "№ RU-12345678-09-2026 от 05.01.2026"
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    sheet["Z5"] = '=IF(F5<>"",ROW(),"")'
    book.save(source)

    result = apply({"RU-12345678-09-2026": _synthetic_record("RU-12345678-09-2026", pdf)}, source, output, sha256(source))
    saved = load_workbook(output)[SHEET]

    assert result["changes"][0]["outcome"] == "review_conflict"
    assert result["conflicts"][0]["action"] == "review_conflict"
    assert saved["F4"].value == "RU-12345678-09-2026"
    assert saved["F5"].value == "№ RU-12345678-09-2026 от 05.01.2026"
    assert saved["Y4"].value == '=IF(A4<>"",ROW(),"")'
    assert saved["Z5"].value == '=IF(F5<>"",ROW(),"")'
    assert saved["W4"].value is None and saved["W5"].value is None


def test_apply_preserves_synthetic_x14_extension_through_staged_publication(tmp_path: Path):
    source, output, pdf = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "permit.pdf"
    pdf.write_bytes(b"pdf")
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["W3"] = "Ссылка на документ"
    sheet["F4"] = "38-1-1-2026"
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    book.save(source)
    with zipfile.ZipFile(source) as archive:
        payload = {entry.filename: archive.read(entry.filename) for entry in archive.infolist()}
    sheet_path = "xl/worksheets/sheet1.xml"
    extension = (
        b'<extLst><ext uri="{A1B2C3D4-E5F6-47A8-9B0C-D1E2F3A4B5C6}">'
        b'<x14:conditionalFormattings xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"/>'
        b"</ext></extLst>"
    )
    payload[sheet_path] = payload[sheet_path].replace(b"</worksheet>", extension + b"</worksheet>")
    with zipfile.ZipFile(source, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in payload.items():
            archive.writestr(name, data)

    result = apply({"38-2-2-2026": _synthetic_record("38-2-2-2026", pdf)}, source, output, sha256(source))
    with zipfile.ZipFile(source) as archive:
        before = archive.read(sheet_path)
    with zipfile.ZipFile(output) as archive:
        after = archive.read(sheet_path)

    assert result["verification"]["x14_preserved"] is True
    assert extension in before and extension in after
    assert load_workbook(output)[SHEET]["Y4"].value == '=IF(A4<>"",ROW(),"")'


def test_directive_only_updates_one_existing_row_but_never_appends(tmp_path: Path):
    source, output, absent = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "absent.xlsx"
    pdf = tmp_path / "изменение.pdf"
    pdf.write_bytes(b"pdf")
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["W3"] = "Ссылка на документ"
    sheet["F4"] = "№ RU-12345678-09-2022 от 24.10.2022"
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    sheet["Z4"] = '=IF(F4<>"",ROW(),"")'
    book.save(source)
    record = _synthetic_record("RU-12345678-09-2022", pdf)
    record.update({"existing_only": True, "end": "25.02.2027", "changed": "10.08.2026"})

    result = apply({"RU-12345678-09-2022": record}, source, output, sha256(source))
    saved = load_workbook(output)[SHEET]
    assert result["changes"][0]["outcome"] == "updated"
    assert saved["H4"].value.strftime("%d.%m.%Y") == "25.02.2027"
    assert saved["I4"].value.strftime("%d.%m.%Y") == "10.08.2026"

    absent_result = apply({"RU-12345678-09-2022": record}, source, absent, sha256(source))
    assert absent_result["changes"][0]["outcome"] == "updated"
    # Use a workbook without the identity to prove no append.
    blank = tmp_path / "blank.xlsx"
    book = Workbook(); sheet = book.active; sheet.title = SHEET; sheet["W3"] = "Ссылка на документ"; sheet["Y4"] = '=IF(A4<>"",ROW(),"")'; book.save(blank)
    missing = apply({"RU-12345678-09-2022": record}, blank, tmp_path / "missing.xlsx", sha256(blank))
    assert missing["changes"][0]["outcome"] == "review_conflict"
    assert load_workbook(tmp_path / "missing.xlsx")[SHEET].max_row == 4


def test_number_only_record_is_review_not_already_present(tmp_path: Path):
    source, output, pdf = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "permit.pdf"
    pdf.write_bytes(b"pdf")
    book = Workbook(); sheet = book.active; sheet.title = SHEET; sheet["W3"] = "Ссылка на документ"; sheet["F4"] = "38-1-1-2026"; book.save(source)

    result = apply({"38-1-1-2026": _synthetic_record("38-1-1-2026", pdf)}, source, output, sha256(source))
    assert result["changes"][0]["outcome"] == "review"
    assert result["changes"][0]["issues"] == [
        "В документе не найдено ни одного подтверждённого поля для переноса."
    ]
    assert load_workbook(output)[SHEET]["W4"].value is None


def test_rns_extracts_sanitized_modern_table_dates():
    text = """RU-12345678-09-2022
1. Дата выдачи разрешения | 24.10.2022 | 0
2. Срок действия | 25.02.2027 | 0
3. Дата последнего изменения | 01.01.2024 | 0
3. Дата последнего изменения | 10.08.2026 | 0
"""
    record = rns_adapter.extract(Path("table.pdf"), text)
    assert record is not None
    assert {field: record[field] for field in ("issue", "end", "changed")} == {
        "issue": "24.10.2022", "end": "25.02.2027", "changed": "10.08.2026",
    }
    assert {field: record["field_provenance"][field] for field in ("issue", "end", "changed")} == {
        "issue": "ocr", "end": "ocr", "changed": "ocr",
    }


def test_run_does_not_publish_when_no_rns_record_is_found(monkeypatch, tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    pdf = pdf_dir / "blank.PDF"
    pdf.write_bytes(b"pdf")
    xlsx = tmp_path / "register.xlsx"
    xlsx.write_bytes(b"xlsx")
    output = tmp_path / "output.xlsx"
    monkeypatch.setattr(
        app,
        "collect",
        lambda *args, **kwargs: ({}, [{"file": str(pdf), "error": "Не найден номер РНС"}]),
    )
    monkeypatch.setattr(app, "apply", lambda *args, **kwargs: pytest.fail("Excel must not be published"))

    with pytest.raises(RuntimeError, match="ни одной записи РНС"):
        app.run(pdf_dir, xlsx, output)
    assert not output.exists()


def test_transient_file_lock_is_retried_with_a_bound(monkeypatch):
    attempts: list[int] = []
    delays: list[float] = []

    def operation():
        attempts.append(1)
        if len(attempts) < 3:
            raise PermissionError("[WinError 32] used by another process")
        return "ready"

    monkeypatch.setattr(server.time, "sleep", delays.append)

    assert retry_file_operation(operation) == "ready"
    assert len(attempts) == 3
    assert delays == [0.2, 0.4]


def test_error_hint_mentions_excel_only_for_access_errors():
    generic = error_hint(RuntimeError("unexpected data"))
    denied = error_hint(PermissionError("[WinError 32] used by another process"))
    javascript = (Path(__file__).parents[1] / "rns_import_server/static/app.js").read_text(encoding="utf-8")

    assert "Закройте" not in generic
    assert "Система запретила запись" in denied
    assert "Проверьте пути, закройте Excel" not in javascript


def test_public_failed_job_redacts_absolute_pdf_path_but_keeps_traceback_in_technical_log(tmp_path: Path):
    """Operator JSON must not disclose a local path or a raw internal exception."""
    pdf_dir = tmp_path / "секретные PDF"
    pdf_dir.mkdir()
    pdf = pdf_dir / "внутренний.pdf"
    pdf.write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")
    secret = "internal_parser_token"

    def fail(*args, **kwargs):
        progress = args[5]
        progress(30, "Распознаём PDF", str(pdf))
        raise RuntimeError(secret)

    log = tmp_path / "technical.log"
    manager = JobManager(fail, error_log=log)
    job = _wait(manager, str(manager.start(str(pdf_dir), str(target))["id"]))
    public = manager.public(str(job["id"]))

    assert public is not None
    assert str(pdf_dir) not in json.dumps(public, ensure_ascii=False)
    assert secret not in json.dumps(public, ensure_ascii=False)
    assert str(pdf) in log.read_text(encoding="utf-8")
    assert secret in log.read_text(encoding="utf-8")


def test_apply_preserves_exact_native_cf_bytes_and_loadability_after_openpyxl_rewrite(tmp_path: Path):
    source, output, pdf = tmp_path / "register.xlsx", tmp_path / "output.xlsx", tmp_path / "permit.pdf"
    pdf.write_bytes(b"pdf")
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    sheet["W3"] = "Ссылка на документ"
    sheet["F4"] = "38-1-1-2026"
    sheet["Y4"] = '=IF(A4<>"",ROW(),"")'
    book.save(source)
    sheet_path = "xl/worksheets/sheet1.xml"
    cf = b'<conditionalFormatting sqref="A4"><cfRule type="expression" priority="1"><formula>A4=1</formula></cfRule></conditionalFormatting>'
    with zipfile.ZipFile(source) as archive:
        payload = {entry.filename: archive.read(entry.filename) for entry in archive.infolist()}
    payload[sheet_path] = payload[sheet_path].replace(b"</worksheet>", cf + b"</worksheet>")
    with zipfile.ZipFile(source, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in payload.items():
            archive.writestr(name, data)

    record = _synthetic_record("38-2-2-2026", pdf)
    record["stage"] = "1"
    result = apply({"38-2-2-2026": record}, source, output, sha256(source))
    with zipfile.ZipFile(output) as archive:
        saved_xml = archive.read(sheet_path)

    assert result["verification"]["native_cf_preserved"] is True
    assert cf in saved_xml
    assert load_workbook(output)[SHEET]["F5"].value == "38-2-2-2026"


def test_approved_date_proposal_updates_exact_row_link_and_verified_backup(tmp_path: Path):
    manager, job, target, pdf = _proposal_job(tmp_path, {"end": "31.12.2026"})
    proposal = job["proposals"][0]
    before_approval = sha256(target)

    with pytest.raises(ValueError):
        manager.approve(str(job["id"]), str(proposal["id"]), "поддельное-разрешение")
    assert sha256(target) == before_approval

    approved = manager.approve(str(job["id"]), str(proposal["id"]), job["capability"])
    saved = load_workbook(target)[SHEET]
    backup = target.parent / "Резервные копии PropExtract" / str(approved["backup"])
    backed_up = load_workbook(backup)[SHEET]

    assert saved["H4"].value == datetime(2026, 12, 31)
    assert saved["W4"].hyperlink.target == pdf.as_uri()
    assert saved["Y4"].value == '=IF(A4<>"",ROW(),"")'
    assert saved["AA4"].value is None
    assert backed_up["H4"].value == datetime(2025, 12, 31)
    assert manager.public(str(job["id"]))["proposals"][0]["status"] == "approved"


def test_distinct_proposals_are_serialized_without_lost_update(tmp_path: Path):
    manager, job, target, _ = _proposal_job(
        tmp_path,
        {"object": "Новый объект", "end": "31.12.2026"},
    )
    proposals = list(job["proposals"])
    errors: list[Exception] = []

    def approve(proposal: dict) -> None:
        try:
            manager.approve(str(job["id"]), str(proposal["id"]), job["capability"])
        except Exception as error:  # Captured for the assertion in the parent thread.
            errors.append(error)

    threads = [threading.Thread(target=approve, args=(proposal,)) for proposal in proposals]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=10)

    saved = load_workbook(target)[SHEET]
    assert errors == []
    assert len(proposals) == 2
    assert saved["D4"].value == "Новый объект"
    assert saved["H4"].value == datetime(2026, 12, 31)
    assert saved["AA4"].value is None
    assert all(item["status"] == "approved" for item in manager.public(str(job["id"]))["proposals"])


def test_stale_target_rejects_proposal_without_overwriting_user_change(tmp_path: Path):
    manager, job, target, _ = _proposal_job(tmp_path, {"object": "Новый объект"})
    proposal = job["proposals"][0]
    changed = load_workbook(target)
    changed[SHEET]["J4"] = "Изменено пользователем"
    changed.save(target)

    with pytest.raises(RuntimeError, match="proposal_target_stale"):
        manager.approve(str(job["id"]), str(proposal["id"]), job["capability"])

    saved = load_workbook(target)[SHEET]
    assert saved["D4"].value == "Старый объект"
    assert saved["J4"].value == "Изменено пользователем"


def _unused_port() -> int:
    with socket.socket() as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def _get(url: str) -> tuple[int, str, dict[str, str]]:
    with urllib.request.urlopen(url, timeout=5) as response:
        return response.status, response.read().decode("utf-8"), dict(response.headers)


def _post(url: str, payload: dict, headers: dict[str, str] | None = None) -> tuple[int, dict]:
    request_headers = {"Content-Type": "application/json"}
    request_headers.update(headers or {})
    request = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers=request_headers,
    )
    with urllib.request.urlopen(request, timeout=5) as response:
        return response.status, json.loads(response.read().decode("utf-8"))


def test_http_serves_admin_help_health_and_security_headers():
    port = _unused_port()
    server = create_server("127.0.0.1", port, _fake_runner)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        status, body, headers = _get(f"http://127.0.0.1:{port}/")
        assert status == 200 and "Перенести данные" in body and "Остановить" in body
        assert headers["X-Frame-Options"] == "DENY"
        status, body, _ = _get(f"http://127.0.0.1:{port}/help")
        assert status == 200 and "Инструкция оператора" in body
        status, body, _ = _get(f"http://127.0.0.1:{port}/health")
        assert status == 200 and json.loads(body) == {"status": "ok", "service": "rns-import"}
        status, body, _ = _get(f"http://127.0.0.1:{port}/api/system")
        assert status == 200 and "commands" in json.loads(body)
        with pytest.raises(urllib.error.HTTPError) as removed:
            _post(f"http://127.0.0.1:{port}/process", {"pdf_dir": "x", "xlsx": "y"})
        assert removed.value.code == 404
        assert "Метод не найден" in removed.value.read().decode("utf-8")
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def test_invalid_job_request_is_rejected():
    port = _unused_port()
    server = create_server("127.0.0.1", port, _fake_runner)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        request = urllib.request.Request(
            f"http://127.0.0.1:{port}/api/jobs",
            data=b"{}",
            method="POST",
            headers={"Content-Type": "application/json"},
        )
        try:
            urllib.request.urlopen(request, timeout=5)
            raise AssertionError("invalid request must fail")
        except urllib.error.HTTPError as error:
            assert error.code == 400
            assert "PDF" in error.read().decode("utf-8")
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def test_explorer_paths_with_quotes_are_normalized_before_job(tmp_path: Path):
    pdf_dir = tmp_path / "PDF документы"
    pdf_dir.mkdir()
    (pdf_dir / "sample.pdf").write_bytes(b"pdf")
    target = tmp_path / "Реестр РНС.xlsx"
    target.write_bytes(b"original")

    normalized_pdf, normalized_xlsx = validated_job_paths(f'"{pdf_dir}"', f'«{target}»')

    assert normalized_pdf == pdf_dir
    assert normalized_xlsx == target
    assert user_path(pdf_dir.as_uri()) == pdf_dir


def test_manual_paths_report_which_value_is_invalid(tmp_path: Path):
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"xlsx")
    with pytest.raises(ValueError, match="Папка с PDF не найдена"):
        validated_job_paths(str(tmp_path / "missing"), str(target))
    with pytest.raises(ValueError, match="указан файл"):
        validated_job_paths(str(target), str(target))


def test_native_picker_endpoint_returns_selected_path(monkeypatch, tmp_path: Path):
    selected = tmp_path / "register.xlsx"
    monkeypatch.setattr("rns_import_server.server.select_path", lambda kind: str(selected) if kind == "xlsx" else None)
    port = _unused_port()
    server = create_server("127.0.0.1", port, _fake_runner)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        status, payload = _post(f"http://127.0.0.1:{port}/api/picker", {"kind": "xlsx"})
        assert status == 200
        assert payload == {"path": str(selected), "cancelled": False}
        status, payload = _post(f"http://127.0.0.1:{port}/api/picker", {"kind": "directory"})
        assert status == 200
        assert payload == {"path": None, "cancelled": True}
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def test_shutdown_endpoint_stops_idle_server():
    port = _unused_port()
    server = create_server("127.0.0.1", port, _fake_runner)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        status, payload = _post(
            f"http://127.0.0.1:{port}/api/shutdown",
            {},
            {"X-PropExtract-Action": "shutdown"},
        )
        assert status == 202
        assert payload == {"status": "stopping"}
        thread.join(timeout=5)
        assert not thread.is_alive()
    finally:
        server.server_close()


def test_shutdown_is_rejected_while_excel_job_is_running(tmp_path: Path):
    release = threading.Event()
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    (pdf_dir / "sample.pdf").write_bytes(b"pdf")
    target = tmp_path / "register.xlsx"
    target.write_bytes(b"original")

    def slow_runner(*args, **kwargs):
        release.wait(timeout=5)
        return _fake_runner(*args, **kwargs)

    port = _unused_port()
    server = create_server("127.0.0.1", port, slow_runner)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    server.job_manager.start(str(pdf_dir), str(target))  # type: ignore[attr-defined]
    try:
        with pytest.raises(urllib.error.HTTPError) as caught:
            _post(
                f"http://127.0.0.1:{port}/api/shutdown",
                {},
                {"X-PropExtract-Action": "shutdown"},
            )
        assert caught.value.code == 409
        assert "идёт перенос данных" in caught.value.read().decode("utf-8")
    finally:
        release.set()
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def test_macos_picker_uses_native_osascript_and_handles_cancel(monkeypatch, tmp_path: Path):
    selected = tmp_path / "pdf"
    selected.mkdir()
    calls: list[list[str]] = []

    def completed(argv, **kwargs):
        calls.append(argv)
        return subprocess.CompletedProcess(argv, 0, f"{selected}\n", "")

    monkeypatch.setattr(picker.sys, "platform", "darwin")
    monkeypatch.setattr(picker.subprocess, "run", completed)
    assert picker.choose("directory") == str(selected)
    assert calls[0][0] == "/usr/bin/osascript"

    monkeypatch.setattr(
        picker.subprocess,
        "run",
        lambda argv, **kwargs: subprocess.CompletedProcess(argv, 1, "", "execution error: cancelled (-128)"),
    )
    assert picker.choose("xlsx") == ""


def test_windows_picker_uses_exact_system_powershell(monkeypatch, tmp_path: Path):
    system_directory = tmp_path / "Windows/System32"
    powershell = system_directory / "WindowsPowerShell/v1.0/powershell.exe"
    selected = tmp_path / "pdf"
    powershell.parent.mkdir(parents=True)
    powershell.write_bytes(b"")
    selected.mkdir()
    calls: list[list[str]] = []

    def completed(argv, **kwargs):
        calls.append(argv)
        assert kwargs["encoding"] == "utf-8"
        assert kwargs["timeout"] == 120
        return subprocess.CompletedProcess(argv, 0, f"{selected}\n", "")

    monkeypatch.setattr(picker.sys, "platform", "win32")
    monkeypatch.setattr(picker, "_windows_system_directory", lambda: system_directory)
    monkeypatch.setattr(picker.subprocess, "run", completed)

    assert picker.choose("directory") == str(selected.resolve())
    assert calls[0][0] == str(powershell)
    assert "FolderBrowserDialog" in picker._windows_dialog_script("directory")
    assert "OpenFileDialog" in picker._windows_dialog_script("xlsx")
    assert "$Owner.StartPosition = 'CenterScreen'" in calls[0][-1]
    assert "$Owner.Opacity = 0.01" in calls[0][-1]


def test_transfer_issue_explains_missing_and_conflicting_values():
    assert transfer_issue("Орган выдачи", None, None) == (
        "Не перенесено «Орган выдачи»: значение не найдено в PDF."
    )
    assert transfer_issue("Орган выдачи", "Администрация", None) == (
        "Не подтверждено «Орган выдачи»: значение не найдено в PDF; "
        "значение Excel «Администрация» сохранено."
    )
    assert transfer_issue("Срок действия", "28.11.2026", "28.11.2025") == (
        "Не перенесено «Срок действия»: в Excel — «28.11.2026», "
        "в PDF — «28.11.2025»; значение Excel сохранено."
    )
    assert transfer_issue("Срок действия", "28.11.2025", "28.11.2025") is None
    assert transfer_issue("Застройщик", 'ПАО "Газпром"', "ПАО «Газпром»") is None
    assert transfer_issue("Орган выдачи", "СЛУЖБА НАДЗОРА", "Служба надзора") is None
    assert _change_outcome(False, [], []) == "already_present"
    assert _change_outcome(False, ["D42"], []) == "updated"
    assert _change_outcome(False, [], ["расхождение"]) == "review"
    assert _change_outcome(True, ["D43"], []) == "added"
    assert _change_outcome(True, ["D43"], ["спор документов"]) == "review"


def test_document_optimizer_text_normalization_contract_is_preserved():
    assert normalize_text("  Монтаж\u00a0  ТРУБ Ёлка ") == "монтаж труб елка"
    assert normalize_text("  ПАО «Газпром»  ", casefold=False) == "ПАО «Газпром»"


def test_bundled_ocr_models_are_verified_and_forced():
    status = bundled_language_status()
    assert set(status) == set(LANGUAGE_HASHES)
    assert all(item["valid"] for item in status.values())
    assert Path(tesseract_environment()["TESSDATA_PREFIX"]).name == "tessdata"


def test_windows_installer_uses_relative_tessdata_prefix_for_native_probes_and_runtime():
    root = Path(__file__).resolve().parents[1]
    installer = (root / "install_windows.ps1").read_text(encoding="utf-8")

    assert installer.count('"TESSDATA_PREFIX"') >= 2
    assert '(Join-Path $Root "rns_import_server\\tessdata")' not in installer
    assert '[Environment]::SetEnvironmentVariable("TESSDATA_PREFIX", "rns_import_server\\tessdata", "Process")' not in installer
    assert '"rns_import_server\\tessdata"' in installer


def test_runtime_reports_bundled_ocr_models():
    status = runtime_status()
    assert set(status["models"]) == {"rus", "eng"}


def test_tesseract_version_formats_used_by_portable_and_system_builds():
    assert _is_supported_tesseract_version("tesseract 5.5.1")
    assert _is_supported_tesseract_version("tesseract v5.5.3.20260724")
    assert not _is_supported_tesseract_version("tesseract 4.1.1")


def test_one_command_installers_cover_required_runtime():
    root = Path(__file__).resolve().parents[1]
    windows = (root / "install_windows.ps1").read_text(encoding="utf-8")
    windows_start = (root / "start_windows.ps1").read_text(encoding="utf-8")
    windows_stop = (root / "stop_windows.ps1").read_text(encoding="utf-8")
    linux = (root / "install_linux.sh").read_text(encoding="utf-8")
    lock = json.loads((root / "windows-runtime.lock.json").read_text(encoding="utf-8"))
    assert lock["architectures"] == ["x64", "arm64-x64-emulation"]
    for artifact in lock["artifacts"].values():
        path = root / artifact["path"]
        assert path.is_file()
        assert sha256(path) == artifact["sha256"]
    for package in lock["pythonTree"]["packages"]:
        path = root / package["path"]
        assert path.is_file()
        assert sha256(path) == package["sha256"]
    assert "WinGet, Microsoft Store, network downloads" in windows
    assert "Test-PropExtractFileSha256" in windows
    assert "Test-PropExtractPythonTree" in windows
    assert "Assert-SupportedWindows" in windows
    assert 'GetEnvironmentVariable("PYTHONDONTWRITEBYTECODE", "Process")' in windows
    assert 'GetEnvironmentVariable("TESSDATA_PREFIX", "Process")' in windows
    assert "sys.maxsize > 2**32" in windows
    assert "struct.calcsize" not in windows
    assert "sys.maxsize > 2**32" in windows_start
    assert "struct.calcsize" not in windows_start
    assert windows_start.isascii() and windows_stop.isascii()
    assert "/api/shutdown" in windows_stop
    assert "X-PropExtract-Action" in windows_stop
    assert (root / "Запустить PropExtract.cmd").is_file()
    assert (root / "Остановить PropExtract.cmd").is_file()
    assert 'function Invoke-NativeProbe' in windows
    assert '$ErrorActionPreference = "Continue"' in windows
    assert 'tesseract v?5\\.' in windows
    assert "-Verb RunAs" not in windows
    assert "EncodedCommand" not in windows
    assert "winget source reset" not in windows.lower()
    assert "Invoke-WebRequest" not in windows
    assert "WebClient" not in windows
    assert "PEP-514" not in windows
    assert "-Verb RunAs" not in (root / "install_windows.cmd").read_text(encoding="utf-8")
    assert "raw/codex/admin-ui" not in json.dumps(lock)
    assert "IO.Compression.ZipFile" in windows
    for package in ("python3-venv", "poppler-utils", "tesseract-ocr"):
        assert package in linux
    assert "-m rns_import_server.runtime" in windows
    assert "-m rns_import_server.runtime" in linux


def test_windows_python_tree_is_reproducible(tmp_path: Path):
    root = Path(__file__).resolve().parents[1]
    lock = json.loads((root / "windows-runtime.lock.json").read_text(encoding="utf-8"))
    digest, count = build_windows_python_runtime(
        root / "packages",
        root / lock["pythonTree"]["pthTemplatePath"],
        tmp_path / "python-runtime",
    )

    assert count == lock["pythonTree"]["files"]
    assert digest == lock["pythonTree"]["sha256"]


def test_project_windows_tool_finds_versioned_portable_runtime(tmp_path: Path):
    tesseract = tmp_path / ".runtime/windows/native-5.5.3/tesseract/tesseract.exe"
    poppler = tmp_path / ".runtime/windows/native-5.5.3/poppler/release/Library/bin/pdfinfo.exe"
    invalid = tmp_path / ".runtime/windows/native-9.invalid.20260810/tesseract/tesseract.exe"
    staging = tmp_path / ".runtime/windows/native-staging-test/tesseract/tesseract.exe"
    tesseract.parent.mkdir(parents=True)
    poppler.parent.mkdir(parents=True)
    invalid.parent.mkdir(parents=True)
    staging.parent.mkdir(parents=True)
    tesseract.write_bytes(b"tesseract")
    poppler.write_bytes(b"pdfinfo")
    runtime = tesseract.parents[1]
    entries = sorted(
        (path.relative_to(runtime).as_posix(), sha256(path))
        for path in runtime.rglob("*")
        if path.is_file()
    )
    digest = hashlib.sha256(
        "".join(f"{item_hash}  {relative}\n" for relative, item_hash in entries).encode()
    ).hexdigest()
    (tmp_path / "windows-runtime.lock.json").write_text(
        json.dumps(
            {
                "runtime": "5.5.3",
                "nativeTree": {
                    "files": len(entries),
                    "sha256": digest,
                    "tesseractPath": "tesseract/tesseract.exe",
                    "popplerBinPath": "poppler/release/Library/bin",
                },
            }
        ),
        encoding="utf-8",
    )
    invalid.write_bytes(b"")
    staging.write_bytes(b"")

    assert project_windows_tool("tesseract", tmp_path) == str(tesseract)
    assert project_windows_tool("pdfinfo", tmp_path) == str(poppler)
    assert project_windows_tool("unknown", tmp_path) is None


def test_project_windows_tool_rejects_a_tampered_runtime(tmp_path: Path):
    runtime = tmp_path / ".runtime/windows/native-1"
    executable = runtime / "tesseract/tesseract.exe"
    executable.parent.mkdir(parents=True)
    executable.write_bytes(b"tampered")
    (tmp_path / "windows-runtime.lock.json").write_text(
        json.dumps(
            {
                "runtime": "1",
                "nativeTree": {
                    "files": 1,
                    "sha256": "0" * 64,
                    "tesseractPath": "tesseract/tesseract.exe",
                    "popplerBinPath": "poppler/bin",
                },
            }
        ),
        encoding="utf-8",
    )

    assert project_windows_tool("tesseract", tmp_path) is None
